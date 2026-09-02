"""Build read-only HTML table views directly from the BT34 companion workbook.

Excel stays the SSOT for these sheets; this script is the only thing allowed
to turn that data into HTML, so the two never drift out of sync by hand-edit.
Re-run after any change to QPS_BT34_SUPPLEMENTARY_DETAIL_v1.xlsx to regenerate
every page in excel_view/.

Pattern follows the existing 02_SCRIPTS/ convention found in the build-script
audit: read source -> json.dumps() embedded in <script> -> vanilla JS handles
client-side sort/filter. Departs from that audit's Location A in one way,
adopting Location B's (cryoplant-project/web/) better precedent instead: an
external, shared stylesheet (qps_shared.css) rather than a per-page inline
<style> block, so every generated page and the PCA/BT navigators share one
visual language.
"""
from __future__ import annotations

import hashlib
import html
import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = ROOT / "QPS_BT34_SUPPLEMENTARY_DETAIL_v1.xlsx"
OUT_DIR = ROOT / "excel_view"

# Sheets worth a rich navigable view: the 6 green "FUNCTION" (BT ranking /
# method) sheets plus the 3 red analysis sheets that drive next-action TODOs.
# (README, blue reference sheets, and purple governance sheets stay
# Excel-only for now -- no navigation value added by an HTML mirror yet.)
SHEETS = [
    "BT_Method",
    "BT_Config",
    "BT_Clusters",
    "BT_Input",
    "STATIC_BT",
    "STATIC_BT_RTM",
    "MERGE_CANDIDATES",
    "CROSSWALK_TO_V24",
    "EXECUTION_PRIORITY",
]


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def sheet_to_rows(ws) -> tuple[list[str], list[list[str]]]:
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows: list[list[str]] = []
    for row in rows_iter:
        if all(c is None for c in row):
            continue
        rows.append(["" if c is None else str(c) for c in row])
    return header, rows


def render_page(sheet_name: str, header: list[str], rows: list[list[str]],
                 source_hash: str, generated_at: str) -> str:
    header_json = json.dumps(header)
    rows_json = json.dumps(rows)
    thead_cells = "".join(f"<th data-col=\"{i}\">{html.escape(h)}</th>" for i, h in enumerate(header))
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{html.escape(sheet_name)} — QPS BT34 Excel view</title>
<link rel="stylesheet" href="qps_shared.css">
</head>
<body>
<header>
  <h1>{html.escape(sheet_name)}</h1>
  <p>Read-only HTML mirror of one sheet in <code>QPS_BT34_SUPPLEMENTARY_DETAIL_v1.xlsx</code>.
     Excel is the source of truth — this page is regenerated from it, never edited directly.
     <a href="index.html">&larr; all sheets</a></p>
  <p class="provenance">source sha256 <code>{source_hash[:16]}&hellip;</code> ·
     generated {generated_at} by <code>build_html_from_excel.py</code></p>
</header>
<main>
  <div class="controls">
    <div class="search-box">
      <input type="text" id="filter" placeholder="Filter rows&hellip;" aria-label="Filter rows">
    </div>
    <span class="row-count" id="row-count"></span>
  </div>
  <div class="table-wrap">
    <table id="data-table">
      <thead><tr>{thead_cells}</tr></thead>
      <tbody></tbody>
    </table>
  </div>
</main>
<footer>Generated from <code>{html.escape(SOURCE_XLSX.name)}</code>, sheet <code>{html.escape(sheet_name)}</code>. Not hand-edited.</footer>
<script>
const HEADER = {header_json};
const ROWS = {rows_json};
const tbody = document.querySelector("#data-table tbody");
const rowCountEl = document.getElementById("row-count");
let sortCol = null, sortDir = 1;
let filterText = "";

function renderRows() {{
  let data = ROWS.map((r, i) => ({{ r, i }}));
  if (filterText) {{
    const q = filterText.toLowerCase();
    data = data.filter(({{r}}) => r.some(c => c.toLowerCase().includes(q)));
  }}
  if (sortCol !== null) {{
    data.sort((a, b) => {{
      const av = a.r[sortCol], bv = b.r[sortCol];
      const an = parseFloat(av), bn = parseFloat(bv);
      const bothNumeric = !isNaN(an) && !isNaN(bn) && av !== "" && bv !== "";
      const cmp = bothNumeric ? (an - bn) : av.localeCompare(bv);
      return cmp * sortDir;
    }});
  }}
  tbody.innerHTML = "";
  const frag = document.createDocumentFragment();
  data.forEach(({{r}}) => {{
    const tr = document.createElement("tr");
    r.forEach(c => {{
      const td = document.createElement("td");
      td.textContent = c;
      tr.appendChild(td);
    }});
    frag.appendChild(tr);
  }});
  tbody.appendChild(frag);
  rowCountEl.textContent = data.length + " / " + ROWS.length + " rows";
}}

document.querySelectorAll("th[data-col]").forEach(th => {{
  th.addEventListener("click", () => {{
    const col = parseInt(th.dataset.col, 10);
    if (sortCol === col) {{ sortDir *= -1; }} else {{ sortCol = col; sortDir = 1; }}
    document.querySelectorAll("th[data-col]").forEach(t => t.classList.remove("sorted-asc", "sorted-desc"));
    th.classList.add(sortDir === 1 ? "sorted-asc" : "sorted-desc");
    renderRows();
  }});
}});

document.getElementById("filter").addEventListener("input", (e) => {{
  filterText = e.target.value;
  renderRows();
}});

renderRows();
</script>
</body>
</html>
"""


def render_index(pages: list[dict], source_hash: str, generated_at: str) -> str:
    items = "".join(
        f'<li><a href="{html.escape(p["file"])}">{html.escape(p["sheet"])}</a>'
        f' <span class="row-count">{p["rows"]} rows &times; {p["cols"]} cols</span></li>'
        for p in pages
    )
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>QPS BT34 Excel view — index</title>
<link rel="stylesheet" href="qps_shared.css">
</head>
<body>
<header>
  <h1>QPS BT34 Excel view</h1>
  <p>HTML built directly from <code>QPS_BT34_SUPPLEMENTARY_DETAIL_v1.xlsx</code> — Excel is the SSOT,
     these pages are a generated, navigable mirror of the BT-ranking "FUNCTION" sheets and the
     3 analysis sheets. Regenerate with <code>build_html_from_excel.py</code> after any workbook change.</p>
  <p class="provenance">source sha256 <code>{source_hash}</code> · generated {generated_at}</p>
</header>
<main>
  <ul style="list-style:none; padding:0; display:grid; gap:8px;">
    {items}
  </ul>
</main>
<footer>See also: <a href="../QPS_PCA_Navigator.html">QPS_PCA_Navigator.html</a>,
  <a href="../QPS_RTM_BT_Navigator_v22.html">QPS_RTM_BT_Navigator_v22.html</a>.</footer>
</body>
</html>
"""


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"source workbook not found: {SOURCE_XLSX}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    source_hash = sha256_of(SOURCE_XLSX)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    pages = []
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet_name}' not found in workbook, skipping")
            continue
        ws = wb[sheet_name]
        header, rows = sheet_to_rows(ws)
        page_html = render_page(sheet_name, header, rows, source_hash, generated_at)
        out_file = OUT_DIR / f"{sheet_name}.html"
        out_file.write_text(page_html, encoding="utf-8")
        pages.append({"sheet": sheet_name, "file": out_file.name, "rows": len(rows), "cols": len(header)})
        print(f"wrote {out_file.name}  ({len(rows)} rows x {len(header)} cols)")

    index_html = render_index(pages, source_hash, generated_at)
    (OUT_DIR / "index.html").write_text(index_html, encoding="utf-8")
    print(f"wrote index.html ({len(pages)} sheets)")


if __name__ == "__main__":
    main()
