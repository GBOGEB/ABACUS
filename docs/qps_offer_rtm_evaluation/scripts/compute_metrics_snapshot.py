"""
compute_metrics_snapshot.py -- DMAIC metric-history log: the "missing link"
gap identified in KNOWLEDGE_TAXONOMY_MAPPING.md (node 04, STATS). Point-in-
time stats were always strong in this project (DOMAIN_SUMMARY, QUALITY_CHECKS,
Navigator Focus Score charts); what was missing was the SAME metrics tracked
ACROSS build versions, so a reviewer can see the evaluation's own progress
over time, not just its current state.

Usage:
    python3 compute_metrics_snapshot.py QPS_OFFER_Evaluation_FULL_v19.xlsx
        -> appends one entry to METRIC_HISTORY.json, keyed by version,
           overwriting any existing entry for that same version (idempotent
           re-run-safe: re-running against the same file produces the same
           entry, doesn't duplicate).

    python3 compute_metrics_snapshot.py --backfill
        -> runs the same extraction against every QPS_OFFER_Evaluation_FULL_vN.xlsx
           found in the working directory (v5 through the current canonical),
           building the full historical trend in one pass. This is a one-time
           backfill against files that already exist on disk -- it does NOT
           fabricate any number, every value is read live from that version's
           own workbook.

Every future FULL_vN build should end with a call to this script (see
build_workbook_v20.py's closing section) so the trend never has a version gap.
"""
import sys, os, re, glob, json
import warnings
warnings.filterwarnings("ignore")
import openpyxl

HISTORY_FILE = "METRIC_HISTORY.json"


def extract_metrics(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    m = {"file": os.path.basename(path)}

    # ---- RTM_RANKING-derived metrics (present from early versions onward)
    if "RTM_RANKING" in wb.sheetnames:
        rr = wb["RTM_RANKING"]
        hdr = [c.value for c in rr[5]] if rr.max_row >= 5 else []
        def col(name):
            return hdr.index(name) + 1 if name in hdr else None
        c_tier = col("Tier"); c_s = col("Weighted S"); c_domain = col("Domain")
        c_gate = col("Gate"); c_cluster = col("Cluster")
        n = 0; tiers = {}; s_vals = []; domains = set(); gated = 0; clustered = 0
        for row in rr.iter_rows(min_row=6, max_row=rr.max_row):
            rid = row[1].value if len(row) > 1 else None
            if not rid:
                continue
            n += 1
            if c_tier and row[c_tier-1].value:
                t = row[c_tier-1].value
                tiers[t] = tiers.get(t, 0) + 1
            if c_s and isinstance(row[c_s-1].value, (int, float)):
                s_vals.append(row[c_s-1].value)
            if c_domain and row[c_domain-1].value:
                domains.add(row[c_domain-1].value)
            if c_gate and row[c_gate-1].value == "Yes":
                gated += 1
            if c_cluster and row[c_cluster-1].value and str(row[c_cluster-1].value).strip().lower() != "not linked":
                clustered += 1
        m["rtm_count"] = n
        m["rtm_tier_distribution"] = tiers
        m["rtm_domain_count"] = len(domains)
        m["rtm_gate_count"] = gated
        m["rtm_sum_weighted_s"] = round(sum(s_vals), 1) if s_vals else None
        m["rtm_avg_weighted_s"] = round(sum(s_vals) / len(s_vals), 2) if s_vals else None
        m["rtm_crosswalk_linked_count"] = clustered
        m["rtm_crosswalk_linked_pct"] = round(clustered / n * 100, 1) if n else None

    # ---- OFFER_RANKING-derived metrics
    if "OFFER_RANKING" in wb.sheetnames:
        orr = wb["OFFER_RANKING"]
        hdr = [c.value for c in orr[5]] if orr.max_row >= 5 else []
        def ocol(name):
            return hdr.index(name) + 1 if name in hdr else None
        c_flag = ocol("Review flag")
        n = 0; flags = {}
        for row in orr.iter_rows(min_row=6, max_row=orr.max_row):
            oid = row[1].value if len(row) > 1 else None
            if not oid:
                continue
            n += 1
            if c_flag and row[c_flag-1].value:
                f = str(row[c_flag-1].value).strip()
                flags[f] = flags.get(f, 0) + 1
        m["offer_count"] = n
        m["offer_review_flag_distribution"] = flags
        ok_like = sum(v for k, v in flags.items() if k.upper().startswith("OK"))
        m["offer_review_ok_pct"] = round(ok_like / n * 100, 1) if n and flags else None

    # ---- RTM_REVIEW_QUEUE-derived: review-completion rollup
    if "RTM_REVIEW_QUEUE" in wb.sheetnames:
        rq = wb["RTM_REVIEW_QUEUE"]
        hdr = [c.value for c in rq[5]] if rq.max_row >= 5 else []
        c_disp = None
        for cand in ("Disposition",):
            if cand in hdr:
                c_disp = hdr.index(cand) + 1
        n = 0; decided = 0
        if c_disp:
            for row in rq.iter_rows(min_row=6, max_row=rq.max_row):
                rid = row[1].value if len(row) > 1 else None
                if not rid:
                    continue
                n += 1
                v = row[c_disp-1].value
                if v and str(v).strip() and str(v).strip().upper() not in ("TBD", ""):
                    decided += 1
        m["review_queue_count"] = n
        m["review_queue_decided_count"] = decided
        m["review_queue_decided_pct"] = round(decided / n * 100, 1) if n else None

    # ---- sheet-count / workbook shape
    m["sheet_count"] = len(wb.sheetnames)

    wb.close()
    return m


def version_key(fname):
    mm = re.search(r"_v(\d+[a-z]?)\.xlsx$", fname)
    return mm.group(1) if mm else fname


def main():
    history = {}
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE) as f:
            history = json.load(f)

    if len(sys.argv) > 1 and sys.argv[1] == "--backfill":
        files = sorted(
            glob.glob("QPS_OFFER_Evaluation_FULL_v*.xlsx"),
            key=lambda p: (len(version_key(p)), version_key(p)),
        )
        # only the canonical vN.xlsx files, not _qa/_qa2/etc scratch copies
        files = [f for f in files if re.match(r"QPS_OFFER_Evaluation_FULL_v\d+[a-z]?\.xlsx$", os.path.basename(f))]
        print(f"Backfilling from {len(files)} historical workbook versions:")
        for f in files:
            v = version_key(f)
            print(f"  v{v}: {f}")
            history[v] = extract_metrics(f)
    else:
        if len(sys.argv) < 2:
            print("Usage: compute_metrics_snapshot.py <workbook.xlsx> | --backfill")
            sys.exit(1)
        path = sys.argv[1]
        v = version_key(os.path.basename(path))
        history[v] = extract_metrics(path)
        print(f"Recorded metrics for v{v} from {path}")

    # keep in version-sorted order when writing back out
    def sortkey(k):
        m = re.match(r"(\d+)([a-z]?)", k)
        return (int(m.group(1)), m.group(2)) if m else (999, k)
    ordered = {k: history[k] for k in sorted(history, key=sortkey)}

    with open(HISTORY_FILE, "w") as f:
        json.dump(ordered, f, indent=2)
    print(f"Wrote {HISTORY_FILE}: {len(ordered)} version(s) tracked")


if __name__ == "__main__":
    main()
