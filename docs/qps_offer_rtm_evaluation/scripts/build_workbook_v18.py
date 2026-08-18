"""
build_workbook_v18.py -- RTM verbatim-text visibility fix.

GBO pointed at OFFER_CANONICAL's full_verbatim_offer column (wide, bulleted
sub-requirements clearly visible) and asked for "a verbatim RTM (where there
are multiple bullets to some requirements and they have subrequirements)".

Investigation: this data ALREADY EXISTS -- RTM_RANKING!N "Full verbatim
requirement" (and RTM_LOOKUP's mirrored row 20) already carries the full
bulleted text with newlines/bullets for sub-requirements, built the same way
as OFFER_CANONICAL's column. The gap wasn't missing data, it was visibility:
RTM_RANKING!M and N were only 13 characters wide (vs. OFFER_CANONICAL's
verbatim columns at 46-60 wide) with a flat, non-content-aware row height --
so a requirement with 10+ bullets and ~2000 characters rendered as a wall of
near-illegible wrapped text in a narrow column, effectively invisible even
though the data was there. Same root cause as the earlier nav-row-clipping
fix (v14) and DELIVERABLES!F fix (v8): a column sized for something else,
reused for much longer real content.

Fix: widen RTM_RANKING!M/N to match the OFFER_CANONICAL convention, and
compute each row's height from its ACTUAL content (wrapped-line estimate +
explicit bullet newlines) instead of one flat height for all 722 rows --
short one-line shalls stay compact, the ~30 heavily-bulleted ones get the
room they need. Same treatment applied to RTM_LOOKUP's single dynamic
lookup row (widened + generously tall, since it can't be content-aware the
way a per-row table can), plus a note pointing at RTM_RANKING!N directly for
the rare item that still doesn't fully fit.
"""
import warnings, math
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

IN = "QPS_OFFER_Evaluation_FULL_v17.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v18.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)
CARLITO = "Carlito"

# ---- RTM_RANKING column widths: fix the whole J-N run together ----
# GBO flagged (on the v17 sheet, before this fix): J ("PDF page") looks
# oversized, while K/L/M/N (Section / Section title / Shall / Full verbatim)
# all look truncated. Confirmed: J was 68 wide for 2-3 digit page numbers
# (leftover from an unrelated earlier width pass, never meant for this
# column), K ("Section", e.g. "9.1") was only 5, and L ("Section title")
# was only 13 -- all three fought each other for the SAME cramped visual
# budget while J wasted space nobody needed. Rebalanced the whole run so
# width roughly matches actual content: short codes stay narrow, text
# fields get the room they need.
ws = wb["RTM_RANKING"]
ws.column_dimensions["J"].width = 10   # PDF page -- 2-3 digit numbers, was 68
ws.column_dimensions["K"].width = 9    # Section -- short dotted codes, was 5
ws.column_dimensions["L"].width = 24   # Section title -- was 13, wrapped word-by-word
ws.column_dimensions["M"].width = 46
ws.column_dimensions["N"].width = 60

def est_lines(text, chars_per_line):
    if not text:
        return 1
    text = str(text)
    total = 0
    for segment in text.split("\n"):
        total += max(1, math.ceil(len(segment) / chars_per_line))
    return total

n_widened = 0
for r in range(6, ws.max_row + 1):
    shall = ws.cell(r, 13).value
    verbatim = ws.cell(r, 14).value
    if not shall and not verbatim:
        continue
    lines_m = est_lines(shall, 58)
    lines_n = est_lines(verbatim, 76)
    lines = max(lines_m, lines_n, 2)
    ws.row_dimensions[r].height = max(30, min(lines * 13 + 8, 400))  # cap so one outlier can't blow out the sheet
    n_widened += 1
print(f"RTM_RANKING!M:N widened + {n_widened} row heights content-fitted")

# ---- RTM_LOOKUP row 19 (Shall statement) / row 20 (Full verbatim requirement) ----
rl = wb["RTM_LOOKUP"]
rl.column_dimensions["B"].width = 60
rl.row_dimensions[19].height = 110   # Shall statement -- comfortably fits the median case
rl.row_dimensions[20].height = 260   # Full verbatim requirement -- fits the p95 case (~800 chars / ~28 bullets)

# small note in row 30 (already blank, no insert needed -- inserting rows
# into a sheet full of absolute-range INDEX/MATCH formulas is avoidable risk
# for no benefit here) so the rare outlier (max observed: ~1950 chars / 28
# bullets) is still discoverable, not silently clipped
note = rl.cell(row=30, column=1,
    value="Note: if the verbatim text above still looks cut off (a handful of items run to 25+ bullets), the full text is always in RTM_RANKING!N for this RTM ID, or widen row 20 further.")
note.font = Font(name=CARLITO, italic=True, size=9.5, color="888888")
note.alignment = Alignment(wrap_text=True, vertical="top")
rl.merge_cells(start_row=30, start_column=1, end_row=30, end_column=3)
rl.row_dimensions[30].height = 26
print("RTM_LOOKUP verbatim row widened/heightened + outlier note added (row 30)")

# ---- RTM_CROSSWALK!J illegible-text bug: found and fixed ----
# GBO: "text no legble in RMT_crosswalk v13 colmn J in white". Root cause,
# confirmed directly from the saved conditional-formatting rules: the
# Direct/Supporting/Broad/Contextual FormulaRules on J5:J381 set the FONT
# to white (FFFFFF) but carry NO fill of their own (dxf fill is empty) --
# so the rule was relying on a coloured fill that was never actually set,
# leaving white text sitting directly on the sheet's pale blue/white row
# banding. Invisible by construction, not a rendering glitch. Fixed by
# changing the four rules' font colour to the same saturated hex already
# used for these labels' badges in the HTML Navigator (Direct=1E8449,
# Supporting=2874A6, Broad=B7791F, Contextual=7F8C8D) instead of white --
# reads clearly against both the white and EAF1F8 banded rows, no fill
# change needed, so the existing banding pattern is untouched.
from openpyxl.formatting.rule import FormulaRule

LINKTYPE_COLORS = {"Direct": "1E8449", "Supporting": "2874A6", "Broad": "B7791F", "Contextual": "7F8C8D"}
cw = wb["RTM_CROSSWALK"]
j_range = None
for rng in list(cw.conditional_formatting._cf_rules.keys()):
    if str(rng.sqref).startswith("J"):
        j_range = str(rng.sqref)
        break
if j_range:
    del cw.conditional_formatting[j_range]
    for label, hexcolor in LINKTYPE_COLORS.items():
        rule = FormulaRule(
            formula=[f'ISNUMBER(SEARCH("{label}",$J5))'],
            font=Font(name=CARLITO, bold=True, color=hexcolor),
            stopIfTrue=False,
        )
        cw.conditional_formatting.add(j_range, rule)
    print(f"RTM_CROSSWALK!{j_range} relation-type text recoloured (was white-on-white, now legible saturated colour)")
else:
    print("WARNING: could not find RTM_CROSSWALK column J conditional-formatting range to fix")

# ---- START_HERE: colour legend + stale version string ----
# GBO: "text and colour boxes seem not to mean anythign - where is clear
# colour legend or guide for all or subsystem specfic - In start here
# workbook". Confirmed: START_HERE uses two pale colour-coded inset boxes
# (pale green x2, pale yellow x1) with no legend anywhere explaining what
# green vs. yellow means -- unlike STATUS (COMPLIANCE_LEGEND), the weight
# dimensions (WEIGHTS_METHOD, this version) or tab colours (NAVIGATION_MAP/
# README, v13/v17), which all have a real legend. Added one here, plus
# pointers to those three so "is there a legend for X" always has one
# answer. Also fixed START_HERE!A1 still reading "v5" -- 13 versions stale.
sh = wb["START_HERE"]
if "v5" in str(sh.cell(1, 1).value or ""):
    sh.cell(1, 1, value="QPS OFFER evaluation workbook — v18 DMAIC-ready")
    print("START_HERE!A1 version string updated (was stale at v5)")

leg_r = 39
sh.cell(row=leg_r, column=1, value="How to read the colour on this page (and where the other legends live)").font = Font(name=CARLITO, size=13, bold=True, color="17365D")
sh.merge_cells(start_row=leg_r, start_column=1, end_row=leg_r, end_column=12)
leg_r += 1
LEGEND_ROWS = [
    ("E2F0D9", "Pale green box", "Supporting definition / explanation -- background you don't need to act on, just understand."),
    ("FFF2CC", "Pale yellow box", "Operational note or caution -- something to actually do or watch for (e.g. which sheet to use)."),
    ("D9EAF7", "Pale blue box", "Cross-reference / pointer to another sheet."),
]
for bg, label, meaning in LEGEND_ROWS:
    swatch = sh.cell(row=leg_r, column=1, value="")
    swatch.fill = PatternFill("solid", fgColor=bg)
    swatch.alignment = Alignment(vertical="top")
    lbl = sh.cell(row=leg_r, column=2, value=label)
    lbl.font = Font(name=CARLITO, bold=True)
    lbl.alignment = Alignment(vertical="top")
    m = sh.cell(row=leg_r, column=3, value=meaning)
    m.alignment = Alignment(wrap_text=True, vertical="top")
    sh.merge_cells(start_row=leg_r, start_column=3, end_row=leg_r, end_column=12)
    sh.row_dimensions[leg_r].height = 20
    leg_r += 1
leg_r += 1
pointer = sh.cell(row=leg_r, column=1, value=(
    "This page's colours are the ONLY workbook-wide, decorative kind. Everywhere else, colour carries specific "
    "meaning and has its own dedicated legend, not repeated here: STATUS colours (pink/orange/green review flags) "
    "-> COMPLIANCE_LEGEND. Weight-dimension colours (L/R/P/F/Q/LC/C headers + row heatmaps) -> WEIGHTS_METHOD. "
    "Sheet-tab colours (which of the 6 groups a tab belongs to) -> NAVIGATION_MAP or README. Cluster colours "
    "(C1-C8) -> CLUSTERS. RTM_CROSSWALK relation-type colours (Direct/Supporting/Broad/Contextual) -> the note on "
    "that sheet itself."))
pointer.font = Font(name=CARLITO, italic=True, size=10, color="555555")
pointer.alignment = Alignment(wrap_text=True, vertical="top")
sh.merge_cells(start_row=leg_r, start_column=1, end_row=leg_r+2, end_column=12)
for rr in range(leg_r, leg_r+3):
    sh.row_dimensions[rr].height = 24
print("START_HERE colour legend added")

wb.save(OUT)
print(f"saved {OUT}")
