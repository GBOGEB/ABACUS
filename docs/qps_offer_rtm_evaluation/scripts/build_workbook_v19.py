"""
build_workbook_v19.py -- four fixes from GBO's latest round, all found by
direct forensic inspection of the saved file (not guessed from the
screenshot alone -- see each section's comment for what was actually found).

1. RTM_RANKING!O:AG layout overhaul. GBO: "this layout is still weird with
   formating and spacing and zoop and sizes of coliumns". Root causes found:
     a) The navy/light-blue title bars (row 1 "RTM static BT ranking...",
        row 2 subtitle) are merged A1:R1 / A2:R2 -- but the sheet actually
        has 33 columns (through AG). Every column from S onward (Q/LC/C
        weight columns, Weighted S, BT Win%, BT lambda index, Primary
        dimension, Interpretation, and all 6 evidence/deliverable columns)
        sits UNDER NO HEADER BAND AT ALL -- plain white, which is exactly
        why the colour bars looked like they "don't align" with the table.
     b) A LEGACY dataBar conditional-format rule on S6:T727 (2 rules) was
        never removed when v17 added the current per-column ColorScaleRule
        heatmap on O6:U727 -- two different CF encodings stacked on the
        same cells produces the horizontal in-cell gradient-smear GBO's
        screenshot shows on column S ("Q" dimension). Confirmed directly
        from the saved rule objects: type='dataBar' at priority 6/7,
        sitting on the exact same range as the new ColorScaleRule.
     c) Column widths for the 7 single-digit (0-3) weight columns were
        wildly inconsistent -- O/P/Q had NO explicit width (Excel default
        ~8.43), R=12, S=11, T=None, U=22 (a single digit in a 22-wide
        column is most of GBO's "big gap" complaint). V/W/X (Weighted S /
        BT Win% / BT lambda index -- all ~5-7 character decimals) were 38 /
        None / 40 -- also inconsistent and oversized.
     d) AC:AF (Evidence/measurability, Explicit deliverable/proof,
        Applicable phase(s), Evidence basis/review -- real prose, up to 495
        characters) had NO explicit width at all (Excel default), and the
        v18 row-height fix only measured columns M/N when sizing each row
        -- so these 4 columns were wrapping into far more lines than the
        row height allowed, silently truncating worse than M/N ever did.
        Fixed by widening them AND folding them into the same
        content-aware row-height calculation as M/N.

2. RTM-side charts. GBO: "dashboard and graphs are only for offer - whet
   about RTM graphs or ranking specfic". Confirmed: DASHBOARD's 4 native
   charts (Top-15 bar, Tier distribution, Avg Weighted S by category, Rank
   profile) are ALL built from OFFER_RANKING/OFFER_CANONICAL. DASHBOARD_2
   is RTM-scoped but is a live numeric filter/counter panel with ZERO
   charts. Added 4 charts to DASHBOARD_2 mirroring DASHBOARD's set,
   built from RTM_RANKING/DOMAIN_SUMMARY (722 RTMs): Top-15 ranked RTMs by
   Weighted S, RTM Tier distribution, Average Weighted S by Domain (reusing
   DOMAIN_SUMMARY's existing table, no new derived data), and an RTM rank
   profile (Weighted S vs rank position, same chart type as the OFFER
   "spike" chart -- lets GBO see whether RTM shows the same gate-precedence
   pattern OFFER does).

3. TAXONOMY border inconsistency. GBO: "incomplete tale for taxonomy - cell
   boarders not fully the same". Confirmed: 3 real data tables on this
   sheet (OFFER evidence-type vocabulary rows 41-51; RTM Category-to-
   top-Primary-Dimension rows 58-68; OFFER Category-to-top-Primary-
   Dimension rows 71-76) have NO border at all on any cell, while the
   sheet's other 3 tables (Requirement Type rows 6-9, Lifecycle supergroups
   rows 14-18, Disposition rows 23-29) all have a full thin-border grid --
   these 3 were added in later phases and never got the same treatment.
   Fixed by applying the identical thin-border grid.

4. Nav-button affordance. GBO: "top 3 rows where colour is currently dark
   turquoise or teal - a border or is it only the text you click - the
   known where and what not clear... maybe make the cell a 3D or texture?"
   Confirmed: every sheet's row-3 HYPERLINK nav bar (fill 0F6B78) is one
   continuous solid-fill band with NO border between adjacent button
   cells -- reads as a single banner, not discrete clickable buttons, even
   though each cell IS individually clickable (whole-cell HYPERLINK).
   Fixed with a cheap raised-button illusion: light teal top/left border +
   dark teal bottom/right border on every such cell, applied wherever this
   exact fill colour + a HYPERLINK formula appears (scanned across all
   sheets, not just the 3 GBO happened to screenshot) -- so the whole
   workbook gets the same fix in one pass, not sheet-by-sheet.
"""
import warnings, math
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

IN = "QPS_OFFER_Evaluation_FULL_v18.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v19.xlsx"
CARLITO = "Carlito"

wb = openpyxl.load_workbook(IN, data_only=False)

# ============================================================ 1. RTM_RANKING layout
ws = wb["RTM_RANKING"]

# 1a. extend the title/subtitle bars to the sheet's real width (A:AG, 33 cols)
for rng in ["A1:R1", "A2:R2"]:
    if rng in [str(r) for r in ws.merged_cells.ranges]:
        ws.unmerge_cells(rng)
ws.merge_cells("A1:AG1")
ws.merge_cells("A2:AG2")
# re-apply fill across the newly-included cells -- merging only preserves the
# top-left cell's own format, the other cells need the same fill set directly
# so there's no seam where the old A1:R1 boundary used to be
title_fill = PatternFill("solid", fgColor="17365D")
subtitle_fill = PatternFill("solid", fgColor="D9EAF7")
for col_idx in range(1, 34):
    from openpyxl.utils import get_column_letter
    col = get_column_letter(col_idx)
    ws[f"{col}1"].fill = title_fill
    ws[f"{col}2"].fill = subtitle_fill
print("RTM_RANKING title/subtitle bars extended A1:AG1 / A2:AG2 (was A1:R1 / A2:R2)")

# 1b. remove the legacy dataBar CF that was stacking with the v17 ColorScaleRule
removed = 0
for rng in list(ws.conditional_formatting._cf_rules.keys()):
    if str(rng.sqref) == "S6:T727":
        del ws.conditional_formatting[str(rng.sqref)]
        removed += 1
print(f"RTM_RANKING legacy dataBar CF removed from S6:T727 ({removed} range(s))")

# 1c. normalise column widths
UNIFORM_DIM_WIDTH = 7.5
for col in ["O", "P", "Q", "R", "S", "T", "U"]:   # L R P F Q LC C
    ws.column_dimensions[col].width = UNIFORM_DIM_WIDTH
ws.column_dimensions["V"].width = 11   # Weighted S
ws.column_dimensions["W"].width = 11   # BT Win %
ws.column_dimensions["X"].width = 12   # BT lambda index
ws.column_dimensions["AC"].width = 50  # Evidence / measurability
ws.column_dimensions["AD"].width = 32  # Explicit deliverable / proof
ws.column_dimensions["AE"].width = 34  # Applicable phase(s)
ws.column_dimensions["AF"].width = 34  # Evidence basis / review
print("RTM_RANKING columns O:X normalised, AC:AF widened (were unset/default)")

# 1d. content-aware row height across ALL wrapped columns, not just M/N
def est_lines(text, chars_per_line):
    if not text:
        return 1
    text = str(text)
    total = 0
    for segment in text.split("\n"):
        total += max(1, math.ceil(len(segment) / chars_per_line))
    return total

WRAP_COLS = [("M", 58), ("N", 76), ("AC", 63), ("AD", 40), ("AE", 43), ("AF", 43)]
n_resized = 0
for r in range(6, ws.max_row + 1):
    any_val = False
    max_lines = 2
    for col, cpl in WRAP_COLS:
        v = ws[f"{col}{r}"].value
        if v:
            any_val = True
            max_lines = max(max_lines, est_lines(v, cpl))
    if not any_val:
        continue
    ws.row_dimensions[r].height = max(30, min(max_lines * 13 + 8, 400))
    n_resized += 1
print(f"RTM_RANKING row heights recalculated across M/N/AC/AD/AE/AF for {n_resized} rows")

# ============================================================ 2. TAXONOMY missing borders
# GBO: "incomplete tale for taxonomy - cell boarders not fully the same".
# Confirmed: 3 of the sheet's 6 data tables (added in later phases) never
# got the thin-border grid the other 3 have. Same border style copied
# exactly from the tables that already have it.
tax = wb["TAXONOMY"]
THIN = Side(style="thin", color="B7B7B7")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TAX_TABLES = [
    ("A41", "B51"),   # OFFER evidence-type vocabulary (header row 41, data to 51)
    ("A58", "D68"),   # RTM: Category -> top Primary Dimension
    ("A71", "D76"),   # OFFER: Category -> top Primary Dimension
]
n_bordered = 0
for start, end in TAX_TABLES:
    for row in tax[f"{start}:{end}"]:
        for cell in row:
            cell.border = GRID
            n_bordered += 1
print(f"TAXONOMY: thin-border grid applied to {len(TAX_TABLES)} previously-unbordered tables ({n_bordered} cells)")

# ============================================================ 3. QUALITY_CHECKS status colours + borders
# Confirmed: the sheet's own formulas produce "OPEN" and "CHECK" (see D6,
# D8, D9, D10-D15) but the Status-column CF vocabulary only covers
# OK/OK-Minor/NOK/MISS/Minor disc./Major disc./ERROR/TBD -- copied from a
# different sheet's disposition vocabulary and never extended to match
# THIS sheet's own two failure states. Result: a real "needs attention" row
# renders with no colour at all, indistinguishable at a glance from a
# not-yet-computed blank. Also had zero borders anywhere (matches the
# TAXONOMY gap, same fix).
qc = wb["QUALITY_CHECKS"]
from openpyxl.formatting.rule import FormulaRule
qc.conditional_formatting.add(
    "D6:D17",
    FormulaRule(formula=['D6="OPEN"'], fill=PatternFill("solid", fgColor="FFC000"), font=Font(name=CARLITO, bold=True, color="5C3A00")),
)
qc.conditional_formatting.add(
    "D6:D17",
    FormulaRule(formula=['D6="CHECK"'], fill=PatternFill("solid", fgColor="FFC000"), font=Font(name=CARLITO, bold=True, color="5C3A00")),
)
for row in qc["A5:D17"]:
    for cell in row:
        cell.border = GRID
print("QUALITY_CHECKS: OPEN/CHECK status colours added (were uncoloured), thin-border grid applied to check table")

# ============================================================ 4. nav-button 3D-bevel affordance
# GBO: row-3 teal HYPERLINK nav bar reads as one solid banner, not discrete
# clickable buttons. Cheap raised-button illusion: light top/left border +
# dark bottom/right border on every cell carrying this exact fill + a
# HYPERLINK formula, scanned across ALL sheets (not just the ones GBO
# happened to screenshot) so the fix is workbook-wide in one pass.
NAV_FILL_HEX = "0F6B78"
BEVEL_LIGHT = Side(style="medium", color="5FB8C4")
BEVEL_DARK = Side(style="medium", color="073F47")
BEVEL_BORDER = Border(top=BEVEL_LIGHT, left=BEVEL_LIGHT, bottom=BEVEL_DARK, right=BEVEL_DARK)
n_buttons = 0
n_sheets_touched = 0
for ws2 in wb.worksheets:
    touched_this_sheet = 0
    for row in ws2.iter_rows(min_row=1, max_row=6):
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("=HYPERLINK"):
                continue
            fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            if fg and fg[-6:].upper() == NAV_FILL_HEX:
                cell.border = BEVEL_BORDER
                n_buttons += 1
                touched_this_sheet += 1
    if touched_this_sheet:
        n_sheets_touched += 1
print(f"Nav-button bevel border applied to {n_buttons} button cells across {n_sheets_touched} sheets")

# ============================================================ 5. RTM-side charts + coverage panel on DASHBOARD_2
# GBO: "dashboard and graphs are only for offer - whet about RTM graphs or
# ranking specfic" + "fully cross walks - what if the rest? not covreed by
# OFFERS or not linked at all? - this view may be updated with additional
# as status dasuboard". Two related gaps, fixed together since both live
# on the RTM-side dashboard: (a) DASHBOARD_2 had live numeric counters but
# zero charts while DASHBOARD (OFFER-side) has 4; (b) RTM_RANKING!AG
# already carries a per-row Linked/Not-linked flag (293 linked / 429 not
# linked to any OFFER item -- confirmed by direct count) but that split
# was never surfaced anywhere as a summary view.
d2 = wb["DASHBOARD_2"]

d2.cell(row=16, column=1, value="RTM-side charts (mirrors DASHBOARD's OFFER-side charts) + crosswalk coverage").font = Font(name=CARLITO, size=13, bold=True, color="17365D")
d2.merge_cells("A16:I16")
helper_note = d2.cell(row=17, column=1, value="Chart data below is auto-computed (COUNTIF/COUNTIFS off RTM_RANKING) -- not for manual editing.")
helper_note.font = Font(name=CARLITO, italic=True, size=9.5, color="888888")
d2.merge_cells("A17:I17")

# ---- helper table 1: Tier distribution (for chart B) ----
d2.cell(row=18, column=1, value="Tier").font = Font(name=CARLITO, bold=True)
d2.cell(row=18, column=2, value="RTM count").font = Font(name=CARLITO, bold=True)
TIERS = ["T0 Gate", "T1 Primary", "T2 Secondary", "T3 Contextual"]
for i, t in enumerate(TIERS):
    r = 19 + i
    d2.cell(row=r, column=1, value=t)
    d2.cell(row=r, column=2, value=f'=COUNTIF(RTM_RANKING!$D$6:$D$727,"{t}")')

# ---- helper table 2: RTM<->OFFER crosswalk coverage (for chart E + callout) ----
# labels kept short ("Linked"/"Not linked", matching the tier-breakdown table's own
# headers below) rather than full sentences, so they fit column D's existing width
# (shared with the filter-helper cells in rows 5-6 above -- widening it here would
# have widened those too)
d2.cell(row=18, column=4, value="Coverage").font = Font(name=CARLITO, bold=True)
d2.cell(row=18, column=5, value="RTM count").font = Font(name=CARLITO, bold=True)
d2.cell(row=19, column=4, value="Linked")
d2.cell(row=19, column=5, value='=COUNTIFS(RTM_RANKING!$AG$6:$AG$727,"<>Not linked to an OFFER item",RTM_RANKING!$AG$6:$AG$727,"<>")')
d2.cell(row=20, column=4, value="Not linked")
d2.cell(row=20, column=5, value='=COUNTIF(RTM_RANKING!$AG$6:$AG$727,"Not linked to an OFFER item")')

d2.cell(row=22, column=4, value="Tier").font = Font(name=CARLITO, bold=True)
d2.cell(row=22, column=5, value="Linked").font = Font(name=CARLITO, bold=True)
d2.cell(row=22, column=6, value="Not linked").font = Font(name=CARLITO, bold=True)
for i, t in enumerate(TIERS):
    r = 23 + i
    d2.cell(row=r, column=4, value=t)
    d2.cell(row=r, column=5, value=f'=COUNTIFS(RTM_RANKING!$D$6:$D$727,"{t}",RTM_RANKING!$AG$6:$AG$727,"<>Not linked to an OFFER item",RTM_RANKING!$AG$6:$AG$727,"<>")')
    d2.cell(row=r, column=6, value=f'=COUNTIFS(RTM_RANKING!$D$6:$D$727,"{t}",RTM_RANKING!$AG$6:$AG$727,"Not linked to an OFFER item")')

callout = d2.cell(row=28, column=1, value=(
    "Coverage read: 293 of 722 RTMs (41%) have a direct RTM_CROSSWALK link to at least one OFFER item; 429 (59%) do not. "
    "That is expected, not a defect -- OFFER items are the bidder's own response document, so many RTMs (procedural/"
    "administrative requirements, or ones no OFFER section happens to address in prose) will never get a direct textual "
    "link even in a fully compliant bid. The number worth watching is the T0 Gate row below: 21 of 43 T0 Gate RTMs "
    "(49%) have NO direct OFFER link -- for a Gate requirement specifically, that means compliance for that item is "
    "resting on EVALUATION_WORKSPACE's manual review, not on a traceable OFFER-text citation. Worth a reviewer pass on "
    "just those 21 before sign-off, not a workbook defect to fix."
))
callout.font = Font(name=CARLITO, size=10.5, color="7A1F42")
callout.fill = PatternFill("solid", fgColor="FCE0EC")
callout.alignment = Alignment(wrap_text=True, vertical="top")
d2.merge_cells(start_row=28, start_column=1, end_row=31, end_column=9)
for r in range(28, 32):
    d2.row_dimensions[r].height = 22

# ---- chart A: Top 15 ranked RTMs -- Weighted S ----
chA = BarChart()
chA.type = "col"
chA.title = "Top 15 ranked RTMs — Weighted S"
chA.y_axis.title = "Weighted S"
data = Reference(wb["RTM_RANKING"], min_col=22, min_row=5, max_row=20)   # V = col 22
cats = Reference(wb["RTM_RANKING"], min_col=2, min_row=6, max_row=20)    # B = RTM ID
chA.add_data(data, titles_from_data=True)
chA.set_categories(cats)
chA.height, chA.width = 8, 15
d2.add_chart(chA, "A34")

# ---- chart B: RTM Tier distribution ----
chB = BarChart()
chB.type = "col"
chB.title = "RTM Tier distribution"
chB.y_axis.title = "RTM count"
dataB = Reference(d2, min_col=2, min_row=18, max_row=22)
catsB = Reference(d2, min_col=1, min_row=19, max_row=22)
chB.add_data(dataB, titles_from_data=True)
chB.set_categories(catsB)
chB.height, chB.width = 8, 15
d2.add_chart(chB, "I34")

# ---- chart C: Average Weighted S by Domain (reuses DOMAIN_SUMMARY's own table) ----
chC = BarChart()
chC.type = "bar"   # horizontal -- 36 domain labels need the room
chC.title = "Average Weighted S by Domain"
chC.y_axis.title = "Domain"
dataC = Reference(wb["DOMAIN_SUMMARY"], min_col=7, min_row=5, max_row=41)   # G = Average Weighted S
catsC = Reference(wb["DOMAIN_SUMMARY"], min_col=1, min_row=6, max_row=41)
chC.add_data(dataC, titles_from_data=True)
chC.set_categories(catsC)
chC.height, chC.width = 16, 15
d2.add_chart(chC, "A52")

# ---- chart D: RTM rank profile -- Weighted S vs rank (mirrors the OFFER "spike" chart) ----
chD = LineChart()
chD.title = "RTM rank profile: Weighted S vs rank position"
chD.y_axis.title = "Weighted S"
chD.x_axis.title = "Rank position"
dataD = Reference(wb["RTM_RANKING"], min_col=22, min_row=5, max_row=727)
chD.add_data(dataD, titles_from_data=True)
chD.height, chD.width = 8, 15
d2.add_chart(chD, "I52")

# ---- chart E: RTM<->OFFER crosswalk coverage by tier (linked vs not, stacked) ----
chE = BarChart()
chE.type = "col"
chE.grouping = "stacked"
chE.overlap = 100
chE.title = "RTM ↔ OFFER crosswalk coverage by Tier"
chE.y_axis.title = "RTM count"
dataE = Reference(d2, min_col=5, max_col=6, min_row=22, max_row=26)
catsE = Reference(d2, min_col=4, min_row=23, max_row=26)
chE.add_data(dataE, titles_from_data=True)
chE.set_categories(catsE)
chE.height, chE.width = 8, 15
d2.add_chart(chE, "A70")

print("DASHBOARD_2: 5 RTM-side charts added (Top-15/Tier/Domain-avg/rank-profile/coverage-by-tier) + coverage callout")

# ============================================================ 6. DOMAIN_SUMMARY stale AutoFilter (likely "Repaired" cause)
# GBO reported Excel opening QPS_OFFER_Evaluation_LITE_v18.xlsx with a
# "Repaired" title-bar warning -- a standing open question from earlier in
# the project that had never been confirmed or root-caused. Systematic
# check this round (every sheet's <dimension> vs <autoFilter> extent,
# every relationship target, every Content_Types override -- all via the
# raw XML, since real-Excel-strictness issues can silently survive an
# openpyxl round-trip and even a LibreOffice re-save without ever showing
# up as a "formula error"): relationships and content types are clean, but
# DOMAIN_SUMMARY's AutoFilter is stale -- <autoFilter ref="A5:L27"/> while
# the sheet's real data now runs to row 41 (36 domains; the filter was
# last set when there were ~22 and never re-applied after more domains
# were added). This is a genuine, confirmed defect independent of whether
# it's the exact trigger Excel's repair dialog is reacting to -- 14
# domains' worth of rows currently sit outside the AutoFilter/sort range
# entirely. Fixed by re-applying the filter to the sheet's actual extent;
# openpyxl regenerates the matching _xlnm._FilterDatabase defined name
# from this automatically on save, so both copies of the range move
# together this time.
ds = wb["DOMAIN_SUMMARY"]
old_af = str(ds.auto_filter.ref)
ds.auto_filter.ref = f"A5:L{ds.max_row}"
print(f"DOMAIN_SUMMARY AutoFilter corrected: was {old_af}, now A5:L{ds.max_row} (matches actual {ds.max_row-5} data rows)")

wb.save(OUT)
print(f"--- saved: {OUT} ---")
