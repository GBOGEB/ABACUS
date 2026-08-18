"""
build_workbook_v9.py -- Phase 9, two GBO asks from the same message thread:

  1. "make the 8 clusters also have some colour representation... softer
     tones or lights... not to clash" -- gives C1-C8 a soft, low-saturation
     palette (blue/teal/lavender/mauve/tan family -- deliberately NOT
     red/orange/yellow/green/pink, since those hues are already spoken for
     by the STATUS legend and the Tier legend; see the new "How to read the
     colours" block on COMPLIANCE_LEGEND). Applied to CLUSTERS!A (exact
     match) and, via search-based conditional formatting since the cell
     text is a full descriptive string, to RTM_RANKING!Cluster and
     RTM_REVIEW_QUEUE!Cluster.

  2. "XLE key columns should be better used with some formatting to train
     the user in... what to press or change, COLOURS safe to edit" -- the
     STATUS dropdown cells already get a colour once a value is picked
     (existing conditional formatting), but the OTHER manual-entry columns
     on EVALUATION_WORKSPACE (Technical Depth, Source/Location, Reviewer
     Comment, Board/Negotiation Topic, Owner, Closure/Decision Note) had NO
     visual cue at all before a value was entered -- nothing told a
     first-time user those cells were meant to be typed into. Gives them a
     soft cream "safe to edit here" base fill (distinct from the cluster
     and status palettes) so the invitation to edit is visible before any
     value is chosen, not just after.

Both are documented in a new "How to read the colours" reference block
added to COMPLIANCE_LEGEND, together with a "used in" trace (exact
sheet!range) for every colour system in the workbook -- so the legend says
where each colour actually shows up, not just what it means in the
abstract. Scope note: this pass covers EVALUATION_WORKSPACE (the primary
reviewer-input sheet). Other manual-entry sheets (RTM_REVIEW_QUEUE
Comments, DMAIC_AUDIT, AUDIT_NOTES, NEGOTIATION_AGENDA) are flagged in the
same block as not yet covered, rather than silently left inconsistent.
"""
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule

IN = "QPS_OFFER_Evaluation_FULL_v8.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v9.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)

# ---- soft, non-clashing C1-C8 palette (blue/teal/lavender/mauve/tan family,
# deliberately outside the red/orange/yellow/green/pink hues already used by
# the STATUS and Tier legends) --------------------------------------------
CLUSTER_COLORS = {
    "C1": "D6E9F8",  # Performance -- sky blue
    "C2": "D6F0EA",  # Process Design -- mint/teal
    "C3": "E2E6ED",  # Mechanical & Equipment -- steel/blue-grey
    "C4": "E6DFF6",  # Software & Control -- lavender
    "C5": "F0E0EE",  # Infrastructure & Integration -- mauve
    "C6": "DCEAE6",  # Reliability & Maintenance -- pale teal-grey
    "C7": "D9E4F0",  # Quality, Testing & Risk -- periwinkle
    "C8": "EDE6DC",  # Commercial & Execution -- warm taupe
}
EDIT_FILL = PatternFill("solid", fgColor="FFFCF0")  # soft cream -- "safe to type here"

# ============================================================== CLUSTERS sheet
ws = wb["CLUSTERS"]
for r in range(6, ws.max_row + 1):
    cid = ws.cell(row=r, column=1).value
    if cid in CLUSTER_COLORS:
        fill = PatternFill("solid", fgColor=CLUSTER_COLORS[cid])
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill

# ==================================================== RTM_RANKING / QUEUE CF
def add_cluster_cf(ws, col_letter, first_row, last_row):
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    for cid, hexcolor in CLUSTER_COLORS.items():
        dxf_fill = PatternFill("solid", fgColor=hexcolor)
        # matches "C1 —" (Direct) and "Inferred: C1 —" (rule-inferred) alike,
        # but NOT the two-digit cluster ids ("C1" would also match inside
        # e.g. nothing here since ids only run C1-C8) -- SEARCH() is safe.
        formula = f'ISNUMBER(SEARCH("{cid} —",{col_letter}{first_row}))'
        rule = FormulaRule(formula=[formula], fill=dxf_fill, stopIfTrue=False)
        ws.conditional_formatting.add(rng, rule)

ws = wb["RTM_RANKING"]
add_cluster_cf(ws, "I", 6, ws.max_row)   # Cluster column
ws2 = wb["RTM_REVIEW_QUEUE"]
add_cluster_cf(ws2, "G", 6, ws2.max_row)  # Cluster column

# ============================================================ editable cells
ws3 = wb["EVALUATION_WORKSPACE"]
EDIT_COLS = ["R", "S", "U", "X", "Y", "AA", "AD", "AE", "AF"]  # Q/W excluded: already status-coloured by existing CF
for col in EDIT_COLS:
    for r in range(6, 56):
        cell = ws3[f"{col}{r}"]
        if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = EDIT_FILL

# =================================================== COMPLIANCE_LEGEND notes
leg = wb["COMPLIANCE_LEGEND"]
r = leg.max_row + 3
leg.cell(row=r, column=1, value="How to read the colours in this workbook").font = Font(name="Carlito", size=13, bold=True, color="17365D")
r += 1
rows_to_add = [
    ("STATUS (this table, above)", "Green/pink/orange/purple/yellow, per-value as shown above.",
     "EVALUATION_WORKSPACE!Q6:Q55, W6:W55 (live, driven by the dropdown) · QUALITY_CHECKS!D6:D17 (health-check OK/CHECK/OPEN)"),
    ("Tier (T0 Gate / T1 / T2 / T3)", "Pale red / orange / yellow / green -- a SEPARATE colour system from STATUS, do not read them against each other.",
     "EVALUATION_WORKSPACE!B6:B55 and every Tier column on RTM_RANKING, OFFER_RANKING, RTM_REVIEW_QUEUE"),
    ("Cluster (C1-C8)", "New (Phase 9): soft blue/teal/lavender/mauve/tan tones, one per cluster, deliberately kept out of the STATUS and Tier hue families so none of the three systems gets misread against another.",
     "CLUSTERS!A6:A13 (exact) · RTM_RANKING!I6:I727 and RTM_REVIEW_QUEUE!G6:G294 (conditional, matches the cluster id inside the longer cell text, incl. \"Inferred:\" rows)"),
    ("Soft cream fill = safe to type/select here", "New (Phase 9): manual-entry cells that had no visual cue before a value was entered now carry a light cream base fill, so the invitation to edit is visible up front, not just after a STATUS is picked. White/no-fill cells elsewhere are formulas or reference text -- not meant to be edited directly.",
     "EVALUATION_WORKSPACE!R,S,U,X,Y,AA,AD,AE,AF (rows 6-55). NOT yet extended to RTM_REVIEW_QUEUE!Comments, DMAIC_AUDIT, AUDIT_NOTES, or NEGOTIATION_AGENDA -- flagged as a follow-up, not done silently."),
]
for label, meaning, usedin in rows_to_add:
    leg.cell(row=r, column=1, value=label).font = Font(name="Carlito", bold=True, size=10.5)
    leg.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    c2 = leg.cell(row=r, column=2, value=meaning)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    leg.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    c4 = leg.cell(row=r, column=4, value=usedin)
    c4.font = Font(name="Carlito", italic=True, size=9.5, color="444444")
    c4.alignment = Alignment(wrap_text=True, vertical="top")
    leg.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    leg.row_dimensions[r].height = 56
    r += 1

# ======================================= Category vs Primary Dimension note
# GBO asked whether Category and Primary Dimension "correlate or group" and
# whether that's captured in TAXONOMY. Answer, computed directly from the
# live data (not asserted): they are NOT duplicates of each other -- Category
# is a content classification (what KIND of requirement/deliverable this is,
# same vocabulary as the Requirement Type subcategories above), while Primary
# Dimension is which of the seven frozen-weight BT scoring axes (defined in
# WEIGHTS_METHOD, not redefined here) this item scored highest on. Most
# Category values spread across 3+ Primary Dimensions with no single one
# above ~50% -- genuinely different axes. The one notable exception, worth
# flagging because it's counter-intuitive: "Quality Assurance" RTMs (91%) and
# "Compliance / Quality" OFFER items (83%) score highest on Safety / Legal,
# not Quality / Verifiability, despite the category name.
from collections import Counter, defaultdict

def crosstab(ws, cat_col, pd_col, id_col=2):
    ct = defaultdict(Counter)
    for r in range(6, ws.max_row + 1):
        rid = ws.cell(row=r, column=id_col).value
        if not rid:
            continue
        cat = ws.cell(row=r, column=cat_col).value or "(blank)"
        pdv = ws.cell(row=r, column=pd_col).value or "(blank)"
        ct[cat][pdv] += 1
    return ct

rtm_ct = crosstab(wb["RTM_RANKING"], 7, 25)
offer_ct = crosstab(wb["OFFER_RANKING"], 6, 18)

tax2 = wb["TAXONOMY"]
r = tax2.max_row + 3
tax2.cell(row=r, column=1, value="Category vs. Primary Dimension -- are they the same thing?").font = Font(name="Carlito", size=13, bold=True, color="17365D")
r += 1
note3 = (
    "No -- confirmed by cross-tabulating the live data, not assumed. Category (RTM_RANKING/OFFER_RANKING) is a "
    "CONTENT classification -- what kind of requirement this is, using the same vocabulary as the Requirement Type "
    "subcategories above. Primary Dimension is a SCORING classification -- which of the seven frozen-weight BT axes "
    "(Safety/Legal, Reliability, Performance, Functional, Quality/Verifiability, Lifecycle, Cost -- defined once in "
    "WEIGHTS_METHOD, not repeated here) this item scored highest on. For most Category values the split across 3+ "
    "Primary Dimensions with no majority confirms these are genuinely different axes, not a duplicate label. The one "
    "exception worth knowing when reading the ranking: \"Quality Assurance\" RTMs (91%) and \"Compliance / Quality\" "
    "OFFER items (83%) score highest on Safety / Legal, not Quality / Verifiability, despite the category name -- "
    "don't assume a \"Quality\"-named category means the Quality/Verifiability dimension drove its score."
)
tax2.cell(row=r, column=1, value=note3).font = Font(name="Carlito", italic=True, size=10.5, color="444444")
tax2.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
tax2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
tax2.row_dimensions[r].height = 95
r += 2

def write_crosstab_table(ws, r, title, ct):
    ws.cell(row=r, column=1, value=title).font = Font(name="Carlito", bold=True, size=11, color="17365D")
    r += 1
    headers = ["Category", "n", "Top Primary Dimension", "Share"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(name="Carlito", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    r += 1
    for cat, dist in sorted(ct.items(), key=lambda kv: -sum(kv[1].values())):
        total = sum(dist.values())
        top_pd, top_n = dist.most_common(1)[0]
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=total)
        ws.cell(row=r, column=3, value=top_pd)
        ws.cell(row=r, column=4, value=f"{top_n/total*100:.0f}%")
        r += 1
    return r + 1

r = write_crosstab_table(tax2, r, "RTM: Category -> top Primary Dimension (by share of that Category's RTMs)", rtm_ct)
r = write_crosstab_table(tax2, r, "OFFER: Category -> top Primary Dimension (by share of that Category's OFFER items)", offer_ct)

# ============================================================ floating legend
# GBO asked for a "floating colour legend" on the core evaluation sheet so
# the STATUS/taxonomy meaning is visible without flipping to
# COMPLIANCE_LEGEND. A shape/textbox can't actually stay pinned on screen
# while scrolling in Excel (only frozen panes can) -- so this uses Excel's
# real floating-UI primitive instead: the data-validation input message,
# which pops up right next to the cell the moment a reviewer clicks it, no
# extra sheet real estate, dismisses itself. Added to the STATUS, Technical
# Depth and Negotiation dropdowns (both bidders) -- the three columns a
# reviewer actually clicks into.
#
# Note on "ACR" (COMPLIANCE_LEGEND!F10, "Minor disc." -> "Who resolves: ACR
# level"): this acronym is not defined anywhere else in the workbook. Rather
# than invent a meaning for it, the tooltip below flags it as undefined and
# points back to COMPLIANCE_LEGEND -- ask GBO what ACR should expand to
# before it's stated as fact anywhere else.
ws4 = wb["EVALUATION_WORKSPACE"]
STATUS_TOOLTIP = (
    "OK=fulfils (green) | OK/Minor=very low risk, can close (light green) | NOK=does not fulfil (pink) | "
    "MISS=info missing (peach) | Minor disc.=resolvable without management, escalates to \"ACR level\" "
    "-- undefined acronym, see COMPLIANCE_LEGEND (light green) | Major disc.=resolvable WITH management "
    "(orange) | ERROR=misinterpretation, clarify with supplier (lilac) | TBD=not yet evaluated, default "
    "(yellow). Full definitions + examples: COMPLIANCE_LEGEND tab."
)
DEPTH_TOOLTIP = (
    "Technical Depth multiplier applied to this row's Weighted S: 0.4 = shallow/cursory check, "
    "0.7 = moderate depth, 1.0 = full depth review. See WEIGHTS_METHOD for how this feeds the Review Index."
)
NEGOT_TOOLTIP = (
    "Negotiation item = flag for the negotiation agenda | RFI - info missing = formal request needed | "
    "Not available = bidder did not provide it | Clarification only = ambiguity, not a gap. "
    "Feeds NEGOTIATION_AGENDA and the QUALITY_CHECKS negotiation-flag counts."
)
for dv in ws4.data_validations.dataValidation:
    rng = str(dv.sqref)
    if "Q6" in rng or "W6" in rng:
        dv.showInputMessage = True
        dv.promptTitle = "STATUS legend (floating)"
        dv.prompt = STATUS_TOOLTIP
    elif "R6" in rng or "X6" in rng:
        dv.showInputMessage = True
        dv.promptTitle = "Technical Depth legend"
        dv.prompt = DEPTH_TOOLTIP
    elif "T6" in rng or "Z6" in rng:
        dv.showInputMessage = True
        dv.promptTitle = "Negotiation flag legend"
        dv.prompt = NEGOT_TOOLTIP

wb.save(OUT)
print(f"saved {OUT}")
