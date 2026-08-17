"""
xlsx_copy_helpers.py -- cross-workbook worksheet copy for openpyxl.

openpyxl's Workbook.copy_worksheet() only works within a single workbook.
This implements a full-fidelity cross-workbook copy: cell values/formulas,
per-cell styles (font/fill/border/alignment/number_format), merged cell
ranges, conditional formatting rules, column widths, row heights, and
freeze panes. Images/charts are NOT copied (none of the sheets in this
project's scope carry any -- verified before use).
"""
import copy
from openpyxl.utils import get_column_letter


def copy_sheet_cross_workbook(src_ws, dst_wb, new_title=None, index=None):
    title = new_title or src_ws.title
    if index is not None:
        dst_ws = dst_wb.create_sheet(title=title, index=index)
    else:
        dst_ws = dst_wb.create_sheet(title=title)

    # cells: value/formula + style
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)

    # merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # column widths / hidden state
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_letter]
        dst_dim.width = dim.width
        dst_dim.hidden = dim.hidden

    # row heights
    for row_idx, dim in src_ws.row_dimensions.items():
        dst_dim = dst_ws.row_dimensions[row_idx]
        dst_dim.height = dim.height
        dst_dim.hidden = dim.hidden

    # freeze panes
    dst_ws.freeze_panes = src_ws.freeze_panes

    # conditional formatting
    for cf_range in src_ws.conditional_formatting:
        for rule in cf_range.rules:
            dst_ws.conditional_formatting.add(str(cf_range.sqref), copy.copy(rule))

    # sheet view / tab color / zoom
    dst_ws.sheet_properties.tabColor = copy.copy(src_ws.sheet_properties.tabColor)
    if src_ws.sheet_view is not None:
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
        dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale

    return dst_ws
