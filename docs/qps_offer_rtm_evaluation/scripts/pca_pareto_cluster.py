"""
pca_pareto_cluster.py -- GBO's task #58 ask: "yes data driven BUT already
knowledge and process: rather try the PCA to group the relevant pareto?"

Distinguishes two different PCA-adjacent questions that were previously
conflated risk:
  (1) The EXISTING PCA (PC1=30.6% var) -- describes how correlated the 7
      scoring DIMENSIONS are with each other (a variable-space question).
  (2) THIS analysis -- clusters the 722 RTM ITEMS in 7-dim PC-space and
      checks whether natural item groupings align with (or contradict) the
      flat Sum-Weighted-S ranking's already-debunked rank-36/top-5% cutoff.

Method:
  - Load the 722 RTMs' raw 0-3 scores across L/R/P/F/Q/LC/C (cols O-U).
  - Standardize (z-score) each dimension -- required before PCA/KMeans since
    the 7 dims aren't on comparable variance scales even though all are 0-3
    (weight-independent scores, but distributions differ).
  - PCA to 2 components for visualization + reuse the loadings for context.
  - KMeans on the STANDARDIZED RAW SCORES (not PC-projection) for k=2..6,
    silhouette score per k to pick a principled k rather than assuming one.
  - For the chosen k, per-cluster: mean Weighted-S, mean rank, size, dimension
    profile (which of L/R/P/F/Q/LC/C dominates) -- to see whether any cluster
    is a coherent "these are the truly critical ones" group distinct from an
    arbitrary top-N slice.
  - Cross-tab: does the existing top-36 (5%) cut fall inside one cluster, or
    is it split across several (i.e. is the admin cutoff cutting through a
    natural group, or does it already respect one)?

Honest framing carried into the output: this is exploratory, not a
replacement recommendation -- it's an answer to "does PCA-based item
clustering reveal a data-driven cut point," which is a different and
answerable question from "is the top-5% cut a natural break" (already
answered NO in the prior Pareto-gap analysis).
"""
import warnings
warnings.filterwarnings("ignore")
import openpyxl
import numpy as np
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score

IN = "QPS_OFFER_Evaluation_FULL_v23.xlsx"
DIMS = ["L", "R", "P", "F", "Q", "LC", "C"]
WEIGHTS = {"L": 0.20, "R": 0.22, "P": 0.20, "F": 0.16, "Q": 0.12, "LC": 0.07, "C": 0.03}

wb = openpyxl.load_workbook(IN, data_only=True)
ws = wb["RTM_RANKING"]

HEADER_ROW = 5
DATA_START = 6
DATA_END = ws.max_row  # 727 -> 722 rows of data (6..727)

hdr = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
col = {name: i + 1 for i, name in enumerate(hdr) if name}

rows = []
for r in range(DATA_START, DATA_END + 1):
    rtm_id = ws.cell(row=r, column=col["RTM ID"]).value
    if not rtm_id:
        continue
    scores = [ws.cell(row=r, column=col[d]).value for d in DIMS]
    if any(s is None for s in scores):
        continue
    rank = ws.cell(row=r, column=col["Rank"]).value
    tier = ws.cell(row=r, column=col["Tier"]).value
    domain = ws.cell(row=r, column=col["Domain"]).value
    cluster_lbl = ws.cell(row=r, column=col["Cluster"]).value
    wS = sum(WEIGHTS[d] * s for d, s in zip(DIMS, scores))
    rows.append({"rtm": rtm_id, "rank": rank, "tier": tier, "domain": domain,
                 "existing_cluster": cluster_lbl, "scores": scores, "wS": wS})

print(f"Loaded {len(rows)} RTMs with complete 7-dim scores")

X = np.array([r["scores"] for r in rows], dtype=float)
scaler = StandardScaler()
Xz = scaler.fit_transform(X)

pca = PCA(n_components=7)
pcs = pca.fit_transform(Xz)
print("\nPCA explained variance ratio (all 7 PCs):")
for i, v in enumerate(pca.explained_variance_ratio_, start=1):
    print(f"  PC{i}: {v*100:.1f}%")
print(f"  PC1+PC2 cumulative: {(pca.explained_variance_ratio_[0]+pca.explained_variance_ratio_[1])*100:.1f}%")

print("\nPC1 loadings (which dims drive PC1):")
for d, l in zip(DIMS, pca.components_[0]):
    print(f"  {d}: {l:+.3f}")
print("PC2 loadings:")
for d, l in zip(DIMS, pca.components_[1]):
    print(f"  {d}: {l:+.3f}")

print("\n--- KMeans silhouette scan, k=2..6 (on standardized raw 7-dim scores) ---")
best_k, best_sil = None, -1
sil_scores = {}
for k in range(2, 7):
    km = KMeans(n_clusters=k, n_init=20, random_state=0)
    labels = km.fit_predict(Xz)
    sil = silhouette_score(Xz, labels)
    sil_scores[k] = sil
    print(f"  k={k}: silhouette={sil:.4f}")
    if sil > best_sil:
        best_sil, best_k = sil, k

print(f"\nBest k by silhouette: {best_k} (silhouette={best_sil:.4f})")

km = KMeans(n_clusters=best_k, n_init=20, random_state=0)
labels = km.fit_predict(Xz)
for i, r in enumerate(rows):
    r["pc_cluster"] = int(labels[i])

print(f"\n--- Cluster profiles (k={best_k}) ---")
for c in range(best_k):
    members = [r for r in rows if r["pc_cluster"] == c]
    n = len(members)
    mean_scores = np.mean([m["scores"] for m in members], axis=0)
    mean_wS = np.mean([m["wS"] for m in members])
    mean_rank = np.mean([m["rank"] for m in members])
    min_rank = min(m["rank"] for m in members)
    max_rank = max(m["rank"] for m in members)
    dom_dim = DIMS[int(np.argmax(mean_scores))]
    print(f"Cluster {c}: n={n} ({n/len(rows)*100:.1f}%), mean wS={mean_wS:.3f}, "
          f"mean rank={mean_rank:.0f}, rank range=[{min_rank}-{max_rank}], "
          f"dominant dim={dom_dim}")
    print(f"  mean scores: " + ", ".join(f"{d}={s:.2f}" for d, s in zip(DIMS, mean_scores)))

print("\n--- Where does the existing top-36 (5%) admin cutoff land across PC-clusters? ---")
top36 = sorted(rows, key=lambda r: -r["wS"])[:36]
from collections import Counter
cnt = Counter(r["pc_cluster"] for r in top36)
for c in sorted(cnt):
    print(f"  Cluster {c}: {cnt[c]} of the top-36 items")

print("\n--- Does any PC-cluster align with a Weighted-S rank break (a genuine 'critical tier')? ---")
# Sort clusters by mean wS descending, check if there's a clean rank-contiguous group
sorted_clusters = sorted(range(best_k), key=lambda c: -np.mean([m["wS"] for m in rows if m["pc_cluster"] == c]))
top_cluster = sorted_clusters[0]
top_members = [r for r in rows if r["pc_cluster"] == top_cluster]
top_members_sorted = sorted(top_members, key=lambda r: r["rank"])
ranks_in_top_cluster = sorted(m["rank"] for m in top_members)
print(f"Highest-wS cluster = Cluster {top_cluster}, n={len(top_members)} "
      f"({len(top_members)/len(rows)*100:.1f}% of all 722)")
print(f"  Rank span: {min(ranks_in_top_cluster)} to {max(ranks_in_top_cluster)}")
# how contiguous is it -- what % of ranks 1..N (N=len(top_members)) actually fall in this cluster?
N = len(top_members)
top_by_rank = set(r["rtm"] for r in sorted(rows, key=lambda r: r["rank"])[:N])
cluster_set = set(m["rtm"] for m in top_members)
overlap = len(top_by_rank & cluster_set)
print(f"  Overlap with 'top-{N} by flat rank': {overlap}/{N} ({overlap/N*100:.0f}%) -- "
      f"if this is near 100%, the PC-cluster just reproduces the flat ranking with a fancier name; "
      f"if much lower, the cluster is finding a genuinely different (non-linear) grouping.")

import json
out = {
    "n_rtms": len(rows),
    "pca_explained_variance": [float(v) for v in pca.explained_variance_ratio_],
    "pc1_loadings": {d: float(l) for d, l in zip(DIMS, pca.components_[0])},
    "pc2_loadings": {d: float(l) for d, l in zip(DIMS, pca.components_[1])},
    "silhouette_by_k": {str(k): float(v) for k, v in sil_scores.items()},
    "best_k": best_k,
    "cluster_profiles": [],
    "top36_cluster_distribution": {str(c): cnt[c] for c in cnt},
    "top_cluster_id": int(top_cluster),
    "top_cluster_size": N,
    "top_cluster_rank_span": [int(min(ranks_in_top_cluster)), int(max(ranks_in_top_cluster))],
    "top_cluster_vs_flat_rank_overlap_pct": round(overlap/N*100, 1),
}
for c in range(best_k):
    members = [r for r in rows if r["pc_cluster"] == c]
    mean_scores = np.mean([m["scores"] for m in members], axis=0)
    out["cluster_profiles"].append({
        "cluster": c, "n": len(members),
        "mean_wS": float(np.mean([m["wS"] for m in members])),
        "mean_rank": float(np.mean([m["rank"] for m in members])),
        "rank_range": [int(min(m["rank"] for m in members)), int(max(m["rank"] for m in members))],
        "mean_scores": {d: float(s) for d, s in zip(DIMS, mean_scores)},
    })
with open("/tmp/pca_cluster_results.json", "w") as f:
    json.dump(out, f, indent=2)
print("\nwrote /tmp/pca_cluster_results.json")
