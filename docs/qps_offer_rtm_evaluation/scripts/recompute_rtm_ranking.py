# -*- coding: utf-8 -*-
"""Recompute RTM_RANKING (Rank/Tier/Weighted S/BT Win%/BT lambda index) after a
dimension-score correction, on both the FULL and LITE workbooks.

This is NOT part of the original v5..v23 build_workbook_vNN.py chain -- it was
added 2026-08-18 to let a dimension-score correction be applied *consistently*
(all 4 derived columns recomputed together, not one cell hand-patched) without
needing the original authoring session's build pipeline. The formulas were
independently reverse-derived and validated against all 722 existing rows
before this script was trusted to write anything:
  - Weighted S = 100 * sum(w_d * x_d) / 3                      -- 0/722 mismatches
  - Rank = stable sort by (gate=='Yes' first, then S desc)      -- 0/722 mismatches
  - Tier = positional bands: gate rows = T0; next 156 = T1;
           next 244 = T2; remainder = T3 (band sizes read off
           the v23 workbook's own existing T1/T2/T3 counts)
  - BT Win% = 100*(wins+0.5*ties)/(N-1), gate-aware pairwise      -- 0/722 mismatches
  - BT lambda index = regularised Zermelo/MM MLE (pseudo-count
           0.10/pair, 320 iterations, geometric-mean normalised,
           rescaled so max lambda = 100), per SESSION_SSOT.yaml's
           own documented method                                -- Pearson r = 0.999999999999 vs stored

Every run appends a dated entry to a RECOMPUTE_LOG sheet in both workbooks and
to RTM_RANKING_RECOMPUTE_LOG.json alongside them (run number, run date, what
changed, before/after, who/why) -- this is intentionally the same
"append, never overwrite" convention as the project's own METRIC_HISTORY.json /
Version_Log, not a new invention.

Gate ('Yes'/'No') is treated as a curated INPUT, never derived -- 2 of 722
existing rows have L=3 but Gate=No (RTM-465, RTM-650), confirming gate status
is a disclosed hand-curated field, not a pure function of the L dimension.
This script never changes Gate; it only recomputes what's downstream of the
dimension scores + the existing gate flags.

Usage (edit OVERRIDES below, then run):
    python recompute_rtm_ranking.py
"""
import json
import math
import os
from datetime import date

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # docs/qps_offer_rtm_evaluation/
FULL_PATH = os.path.join(BASE_DIR, 'current', 'QPS_OFFER_Evaluation_FULL_v23.xlsx')
LITE_PATH = os.path.join(BASE_DIR, 'current', 'QPS_OFFER_Evaluation_LITE_v23.xlsx')
LOG_PATH = os.path.join(BASE_DIR, 'current', 'RTM_RANKING_RECOMPUTE_LOG.json')

WEIGHTS = [0.20, 0.22, 0.20, 0.16, 0.12, 0.07, 0.03]  # L R P F Q LC C
DIM_COLS = list(range(15, 22))  # L..C
RANK_COL, GATE_COL, TIER_COL, S_COL, WIN_COL, LAMBDA_COL = 1, 3, 4, 22, 23, 24
ID_COL = 2
FIRST_DATA_ROW = 6

# T1/T2/T3 band sizes read off the existing v23 sheet (see docstring) -- edit
# here if the project's own tier-banding convention ever changes.
T1_SIZE, T2_SIZE = 156, 244

# --------------------------------------------------------------------- edit this per run
RUN_NOTE = (
    "RTM-320 (Network Infrastructure, SS4.6.4.4): Performance dimension corrected "
    "0 -> 1. The requirement explicitly states 'All links within the aggregation "
    "network ... shall support 1 Gbit/s bandwidth' -- a concrete capacity/"
    "throughput figure the Performance dimension (heat loads/capacity/transients) "
    "is defined to capture, which the existing score missed. Reliability(3)/"
    "Functional(3) unchanged -- both remain well supported by the redundancy-"
    "manager, failover, and fault-recovery language. Requested by GBO "
    "(system owner) after independent re-read of the full RTM text against the "
    "7-dimension rubric; part of a full re-judgment of all 7 previously-flagged "
    "rows (RTM-328/482/603/718/599/018/320) -- see "
    "README_ABACUS_IMPORT.md for the other 6, which held up unchanged on review."
)
OVERRIDES = {
    'RTM-320': {17: 1},  # column 17 = P (Performance); 0 -> 1
}
TRIGGERED_BY = "GBO (system owner / data owner / responsible), via Claude Code"
# ---------------------------------------------------------------------------------------


def weighted_s(dims):
    return 100 * sum(w * x for w, x in zip(WEIGHTS, dims)) / 3


def gate_key(gate, s):
    return (0 if gate == 'Yes' else 1, -s)


def bt_lambda_fit(rows, iterations=320, pseudo_count=0.10, tol=1e-12):
    n = len(rows)
    keys = [gate_key(r['gate'], r['S']) for r in rows]
    lam = [1.0] * n
    for _ in range(iterations):
        new_lam = [0.0] * n
        for i in range(n):
            num = 0.0
            for j in range(n):
                if i == j:
                    continue
                if keys[i] < keys[j]:
                    wij = 1.0
                elif keys[i] > keys[j]:
                    wij = 0.0
                else:
                    wij = 0.5
                num += wij + pseudo_count
            new_lam[i] = num
        for i in range(n):
            den = 0.0
            for j in range(n):
                if i == j:
                    continue
                nij = 1.0 + 2 * pseudo_count
                den += nij / (lam[i] + lam[j])
            new_lam[i] = new_lam[i] / den if den > 0 else lam[i]
        gm = math.exp(sum(math.log(x) for x in new_lam) / n)
        new_lam = [x / gm for x in new_lam]
        diff = max(abs(a - b) for a, b in zip(lam, new_lam))
        lam = new_lam
        if diff < tol:
            break
    maxlam = max(lam)
    return [100 * x / maxlam for x in lam]


def recompute_sheet(ws, overrides):
    """Applies dimension overrides, recomputes Rank/Tier/S/Win%/lambda, and
    PHYSICALLY re-sorts the data rows to match (row order IS the documented
    sort order on this sheet -- 'Sorted by gate precedence and frozen-weight
    composite' -- rank was literally row-5 for all 722 rows pre-edit, verified
    before this script was trusted). Comments (7, all on the row-5 header) and
    all conditional formatting on this sheet are value/expression-driven, not
    tied to a fixed row position -- confirmed safe to reorder before this
    function was written. Returns (dim_change_records, rank_tier_change_records)."""
    max_col = ws.max_column
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        rid = ws.cell(row=r, column=ID_COL).value
        if rid is None:
            continue
        values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        dims = [values[c - 1] for c in DIM_COLS]
        gate = values[GATE_COL - 1]
        rows.append(dict(orig_row=r, rid=rid, dims=dims, gate=gate, values=values,
                          S_old=values[S_COL - 1], rank_old=values[RANK_COL - 1],
                          tier_old=values[TIER_COL - 1]))

    changes = []
    for rid, dim_overrides in overrides.items():
        for row in rows:
            if row['rid'] != rid:
                continue
            for col, new_val in dim_overrides.items():
                old_val = row['values'][col - 1]
                if old_val == new_val:
                    continue
                dim_idx = col - DIM_COLS[0]
                row['dims'][dim_idx] = new_val
                row['values'][col - 1] = new_val
                changes.append(dict(rtm_id=rid, field=ws.cell(row=5, column=col).value,
                                     old_value=old_val, new_value=new_val))

    for row in rows:
        row['S'] = weighted_s(row['dims'])

    ordered = sorted(rows, key=lambda r: gate_key(r['gate'], r['S']))

    gate_count = sum(1 for r in rows if r['gate'] == 'Yes')
    for pos, row in enumerate(ordered, start=1):
        row['rank_new'] = pos
        if row['gate'] == 'Yes':
            row['tier_new'] = 'T0 Gate'
        else:
            non_gate_pos = pos - gate_count
            if non_gate_pos <= T1_SIZE:
                row['tier_new'] = 'T1 Primary'
            elif non_gate_pos <= T1_SIZE + T2_SIZE:
                row['tier_new'] = 'T2 Secondary'
            else:
                row['tier_new'] = 'T3 Contextual'

    win_pct = {}
    n = len(rows)
    for a in rows:
        ka = gate_key(a['gate'], a['S'])
        wins = sum(1 for b in rows if b['rid'] != a['rid'] and gate_key(b['gate'], b['S']) > ka)
        ties = sum(1 for b in rows if b['rid'] != a['rid'] and gate_key(b['gate'], b['S']) == ka)
        win_pct[a['rid']] = 100 * (wins + 0.5 * ties) / (n - 1)

    lam_values = bt_lambda_fit(rows)
    lam_by_rid = {r['rid']: l for r, l in zip(rows, lam_values)}

    rank_tier_changes = []
    for row in rows:
        row['values'][S_COL - 1] = round(row['S'], 6)
        row['values'][RANK_COL - 1] = row['rank_new']
        row['values'][TIER_COL - 1] = row['tier_new']
        row['values'][WIN_COL - 1] = round(win_pct[row['rid']], 4)
        row['values'][LAMBDA_COL - 1] = round(lam_by_rid[row['rid']], 4)

        if row['rank_old'] != row['rank_new'] or row['tier_old'] != row['tier_new']:
            rank_tier_changes.append(dict(
                rtm_id=row['rid'], old_rank=row['rank_old'], new_rank=row['rank_new'],
                old_tier=row['tier_old'], new_tier=row['tier_new'],
                old_S=row['S_old'], new_S=round(row['S'], 6)))

    # physically rewrite rows 6..727 in the new order
    for pos, row in enumerate(ordered, start=1):
        target_row = FIRST_DATA_ROW - 1 + pos
        for col_idx, val in enumerate(row['values'], start=1):
            ws.cell(row=target_row, column=col_idx).value = val

    return changes, rank_tier_changes


def append_recompute_log_sheet(wb, run_number, run_date, dim_changes, rank_tier_changes, note):
    if 'RECOMPUTE_LOG' in wb.sheetnames:
        ws = wb['RECOMPUTE_LOG']
    else:
        ws = wb.create_sheet('RECOMPUTE_LOG')
        ws.append(['RECOMPUTE_LOG -- audit trail for recompute_rtm_ranking.py runs (not part of the original v5..v23 build chain)'])
        ws.append(['Run', 'Date', 'Field', 'RTM ID', 'Old value', 'New value', 'Note / rank+tier side-effects'])

    side_effects = '; '.join(
        f"{c['rtm_id']}: rank {c['old_rank']}->{c['new_rank']}, tier {c['old_tier']}->{c['new_tier']}"
        for c in rank_tier_changes
    ) if rank_tier_changes else '(no other row crossed a rank/tier boundary)'

    first_row_of_run = True
    for c in dim_changes:
        ws.append([run_number, run_date, c['field'], c['rtm_id'], c['old_value'], c['new_value'],
                   note if first_row_of_run else ''])
        first_row_of_run = False
    ws.append([run_number, run_date, 'RANK/TIER side-effects', '(all 722 rows recomputed)', '', '', side_effects])


def run():
    run_date = date.today().isoformat()

    if os.path.exists(LOG_PATH):
        with open(LOG_PATH, 'r', encoding='utf-8') as f:
            log = json.load(f)
    else:
        log = {'runs': []}
    run_number = len(log['runs']) + 1

    all_rank_tier_changes = {}
    for label, path in [('FULL', FULL_PATH), ('LITE', LITE_PATH)]:
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb['RTM_RANKING']
        dim_changes, rank_tier_changes = recompute_sheet(ws, OVERRIDES)
        append_recompute_log_sheet(wb, run_number, run_date, dim_changes, rank_tier_changes, RUN_NOTE)
        wb.calculation.fullCalcOnLoad = True
        wb.save(path)
        all_rank_tier_changes[label] = rank_tier_changes
        print(f'{label}: {len(dim_changes)} dimension cell(s) changed, '
              f'{len(rank_tier_changes)} row(s) crossed a rank/tier boundary')
        for c in rank_tier_changes[:20]:
            print('  ', c)

    log['runs'].append(dict(
        run_number=run_number,
        run_date=run_date,
        triggered_by=TRIGGERED_BY,
        overrides=OVERRIDES,
        note=RUN_NOTE,
        rank_tier_side_effects=all_rank_tier_changes['FULL'],
    ))
    with open(LOG_PATH, 'w', encoding='utf-8') as f:
        json.dump(log, f, indent=2)
    print(f'Run #{run_number} ({run_date}) logged to {LOG_PATH}')


if __name__ == '__main__':
    run()
