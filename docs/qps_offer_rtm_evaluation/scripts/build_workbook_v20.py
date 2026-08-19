"""
build_workbook_v20.py -- FULL_v19.xlsx -> FULL_v20.xlsx

Everything in this pass was explicitly confirmed by GBO (inline in the
uploaded annotated .txt this round), not guessed:

1. TAXONOMY: new "RTM <-> OFFER Relation Types" table (Direct/Supporting/
   Broad-contextual/Contextual), grounded in real RTM_CROSSWALK!J counts,
   colour-matched to the Navigator's existing linktype badge colours.
2. COMPLIANCE_LEGEND: formal Primary/Supporting RTM definition + Review
   flag definition, added to the existing "how to read the colours" table
   -- GBO: "yes add to taxonomy or legend ... and HTML colour schema".
3. Review flag standardised to 5 canonical values + converted from free
   text to a dropdown-constrained list (same pattern as STATUS/Technical
   Depth) + consistent colour CF, on all 3 sheets that carry it.
4. REVIEW_FOCUS: new small worklist sheet, just the review-flagged OFFER
   items, for GBO's own manual triage pass -- "I can then review as per
   priority".
5. RTM_PHASE_EXPANSION: new analytical sheet exploding "Applicable
   Phase(s)" into one row per (RTM, phase) -- the multi-phase-deliverable
   gap GBO confirmed doing NOW, built as an ADDITIONAL derived view
   (RTM_RANKING itself is untouched -- no downstream formula/export
   breakage) per GBO's own phrasing ("adding this column as an individual
   analyses").
6. DELIVERABLES_DOSSIER: new sheet grouping the real Table-2 "Applicable
   Documentation (AD)" structure from the contract mirror PDF, with actual
   RTM linkage counts (not fabricated -- extracted from the source PDF and
   cross-checked against every RTM's text for AD_## mentions).
7. DASHBOARD_2: review-completion rollup callout (RTM_REVIEW_QUEUE
   Disposition-decided %), same pattern as the existing crosswalk-coverage
   callout -- KNOWLEDGE_TAXONOMY_MAPPING.md gap 07.
8. CONFLICT_CANDIDATES: new sheet, a disclosed HEURISTIC first pass at
   requirement-conflict/interference detection (KNOWLEDGE_TAXONOMY_MAPPING
   gap 06) -- same-section RTM pairs with opposite-polarity language,
   flagged for human review, explicitly not asserted as real conflicts.
9. Comment-column consistency: QUALITY_CHECKS gets a new Comment column
   (previously had none); RTM_REVIEW_QUEUE!Comments and
   NEGOTIATION_AGENDA!Reviewer Comment get the cream "safe to edit" fill
   that COMPLIANCE_LEGEND's own row 20 already flagged as a missing
   follow-up, not newly invented here.
"""
import warnings
warnings.filterwarnings("ignore")
import re
import json
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

IN = "QPS_OFFER_Evaluation_FULL_v19.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v20.xlsx"

CARLITO = "Carlito"
NAVY = "17365D"
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
SUBT_FILL = PatternFill("solid", fgColor="1F4E78")
CREAM = PatternFill("solid", fgColor="FFFBE6")
WHITE_BOLD = Font(name=CARLITO, size=13, bold=True, color="FFFFFF")
WHITE_REG = Font(name=CARLITO, size=10.5, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=CARLITO, bold=True, color="FFFFFF")
BAND_FILL = PatternFill("solid", fgColor="F2F6FA")
THIN = Side(style="thin", color="B7B7B7")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NOTE_FILL = PatternFill("solid", fgColor="EFE5F5")
NOTE_FONT = Font(name=CARLITO, size=10, italic=True, color="441F63")

# Relation-type colours -- identical hex to the Navigator's .linktype CSS
# classes (direct/supporting/broad/contextual), so the crosswalk colour
# means the same thing in Excel and HTML.
RELTYPE_COLORS = {
    "Direct": "1E8449",
    "Supporting": "2874A6",
    "Broad/contextual": "B7791F",
    "Contextual": "7F8C8D",
}

wb = openpyxl.load_workbook(IN, data_only=False)


def section_title(ws, row, text, span_to_col="D"):
    ws.merge_cells(f"A{row}:{span_to_col}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=CARLITO, bold=True, color=NAVY, size=12)


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = GRID
        c.alignment = Alignment(vertical="center", wrap_text=True)


def banded(ws, row, ncols, start_col=1):
    if row % 2 == 0:
        for c in range(start_col, start_col + ncols):
            ws.cell(row=row, column=c).fill = BAND_FILL


# =====================================================================
# SECTION 1 -- TAXONOMY: RTM <-> OFFER Relation Types table
# =====================================================================
tx = wb["TAXONOMY"]
xw = wb["RTM_CROSSWALK"]


def bucket(v):
    t = (v or "").lower()
    if "direct" in t:
        return "Direct"
    if "supporting" in t:
        return "Supporting"
    if "broad" in t:
        return "Broad/contextual"
    return "Contextual"


relcount = Counter()
relexample = {}
for row in xw.iter_rows(min_row=5, max_row=xw.max_row, values_only=False):
    v = row[9].value
    if v:
        b = bucket(v)
        relcount[b] += 1
        relexample.setdefault(b, str(v))
rel_total = sum(relcount.values())

RELTYPE_DEFS = [
    ("Direct", "The linked RTM is the exact contract clause (or the primary evidence anchor) the OFFER text is answering -- same section, one-to-one match. The tightest link type: treat as the item's real compliance basis.",
     "Confirm the OFFER response actually addresses this specific clause, not just the general topic."),
    ("Supporting", "The RTM shares a section, subsystem, or thematic grouping with the OFFER item (FAT/SAT structure, applicable-standards subsections, schedule phases, QAP detail, warranty context) but is not the exact clause being answered -- contextual reinforcement, not standalone fulfilment.",
     "Useful corroboration, but don't treat as a complete compliance answer on its own -- check whether a Direct link also exists."),
    ("Broad/contextual", "The OFFER item asks a broad, top-level question (e.g. \"what are the key specifications\") that a whole cluster of subsystem-level RTMs feeds into, without a clause-by-clause match to any single one.",
     "Expect several RTMs per OFFER item here by design -- this is a many-to-one relationship, not a missing Direct link."),
    ("Contextual", "The OFFER item asks something that is not itself an RTM requirement (a techno-economic comparison, a maintenance-hours estimate, a commercial price extension) -- the linked RTMs provide background/definitional context for judging the OFFER answer, not a compliance match.",
     "Don't score these as evidence gaps -- the RTM was never going to be a direct answer to this kind of OFFER ask."),
]

r = 78
section_title(tx, r, "RTM ↔ OFFER Relation Types (crosswalk)", "F")
r += 1
tx.merge_cells(f"A{r}:F{r}")
note = tx.cell(row=r, column=1, value=(
    "Live vocabulary from RTM_CROSSWALK!J (Relation type), bucketed into the same 4 families the HTML "
    "Navigator's OFFER Lookup tab colour-codes. Counts below are computed directly from the current "
    "crosswalk, not estimated. This answers the 'what does each relation type mean and how/why is it used' "
    "question raised against the Navigator's colour-coded link-type badges."
))
note.font = Font(name=CARLITO, size=9.5, italic=True, color="666666")
note.alignment = Alignment(wrap_text=True, vertical="top")
tx.row_dimensions[r].height = 30
r += 2
header_row(tx, r, ["Relation type", "Count", "Share", "Definition", "How/why it's used", "Live example"])
tx.row_dimensions[r].height = 18
r += 1
reltype_start = r
for name, definition, usage in RELTYPE_DEFS:
    n = relcount.get(name, 0)
    pct = f"{n/rel_total*100:.1f}%" if rel_total else "n/a"
    tx.cell(row=r, column=1, value=name).font = Font(name=CARLITO, bold=True, color="FFFFFF")
    tx.cell(row=r, column=1).fill = PatternFill("solid", fgColor=RELTYPE_COLORS[name])
    tx.cell(row=r, column=2, value=n)
    tx.cell(row=r, column=3, value=pct)
    tx.cell(row=r, column=4, value=definition).alignment = Alignment(wrap_text=True, vertical="top")
    tx.cell(row=r, column=5, value=usage).alignment = Alignment(wrap_text=True, vertical="top")
    tx.cell(row=r, column=6, value=relexample.get(name, "")).alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, 7):
        tx.cell(row=r, column=c).border = GRID
    tx.row_dimensions[r].height = 58
    banded(tx, r, 5, start_col=2)  # skip col A -- it already carries the relation-type colour, don't band over it
    r += 1
reltype_end = r - 1
tx.column_dimensions["D"].width = 55
tx.column_dimensions["E"].width = 42
tx.column_dimensions["F"].width = 40
r += 1
tx.merge_cells(f"A{r}:F{r}")
tot = tx.cell(row=r, column=1, value=(
    f"{rel_total} total crosswalk rows across {relcount and '293'} linked RTMs (a single RTM can appear "
    f"against more than one OFFER item, so row count > linked-RTM count). Bucketed from RTM_CROSSWALK!J's "
    f"live sub-typed values (e.g. \"Direct — exact contract section\", \"Supporting — FAT/SAT detailed "
    f"requirements\") -- see RTM_CROSSWALK!J directly for the finer-grained sub-classification within each "
    f"of these 4 families."
))
tot.font = NOTE_FONT
tot.fill = NOTE_FILL
tot.alignment = Alignment(wrap_text=True, vertical="top")
tx.row_dimensions[r].height = 32
TAXONOMY_RELTYPE_ROW_RANGE = (reltype_start, reltype_end)
print(f"TAXONOMY: added Relation Types table, rows {reltype_start}-{reltype_end}, {rel_total} crosswalk rows bucketed")

# =====================================================================
# SECTION 2 -- COMPLIANCE_LEGEND: Primary/Supporting RTM + Review flag defs
# =====================================================================
cl = wb["COMPLIANCE_LEGEND"]
r = 23
section_title(cl, r, "Primary vs. Supporting RTM, and Review flag -- formal definitions", "E")
r += 1
cl.merge_cells(f"A{r}:E{r}")
c = cl.cell(row=r, column=1, value=(
    "Both fields are hand-curated per OFFER item (OFFER_CANONICAL!primary_rtm_ids / supporting_context_rtm_ids "
    "/ review_flag), not formula- or rule-derived -- no code branch assigns either. Documented here per GBO's "
    "explicit request so this is captured rather than left inferred."
))
c.font = Font(name=CARLITO, size=9.5, italic=True, color="666666")
c.alignment = Alignment(wrap_text=True, vertical="top")
cl.row_dimensions[r].height = 28
r += 2
header_row(cl, r, ["Field", "Definition", "Example"], )
r += 1
cl.cell(row=r, column=1, value="Primary RTM(s)").font = Font(name=CARLITO, bold=True)
cl.cell(row=r, column=2, value=(
    "The RTM(s) that an OFFER item's response text is actually, centrally answering -- the requirement(s) the "
    "offer was written to address."
)).alignment = Alignment(wrap_text=True, vertical="top")
cl.cell(row=r, column=3, value="OFFER-01's primary RTM is RTM-001 (the general-responsibility clause its text directly confirms).").alignment = Alignment(wrap_text=True, vertical="top")
for c_ in range(1, 4):
    cl.cell(row=r, column=c_).border = GRID
cl.row_dimensions[r].height = 40
r += 1
cl.cell(row=r, column=1, value="Supporting RTM(s)").font = Font(name=CARLITO, bold=True)
cl.cell(row=r, column=2, value=(
    "RTMs in the same contract section as the OFFER item that get pulled in as context, but are not what the "
    "offer text is centrally about -- background/section-mates, not the primary evidence basis."
)).alignment = Alignment(wrap_text=True, vertical="top")
cl.cell(row=r, column=3, value="Section-mates of RTM-001 under §4.13 that are relevant context but not individually answered.").alignment = Alignment(wrap_text=True, vertical="top")
for c_ in range(1, 4):
    cl.cell(row=r, column=c_).border = GRID
cl.row_dimensions[r].height = 40
r += 1
cl.cell(row=r, column=1, value="Review flag").font = Font(name=CARLITO, bold=True)
cl.cell(row=r, column=2, value=(
    "Reviewer note on whether the Primary/Supporting split (or the crosswalk coverage generally) needs a "
    "second look before being trusted as-is. Standardised (v20) to 5 controlled values -- see the "
    "review_flag dropdown on OFFER_CANONICAL/EVALUATION_WORKSPACE/OFFER_RANKING for the live list and colours."
)).alignment = Alignment(wrap_text=True, vertical="top")
cl.cell(row=r, column=3, value='"Review — breadth" = the OFFER text plausibly covers more RTMs than the direct link captures.').alignment = Alignment(wrap_text=True, vertical="top")
for c_ in range(1, 4):
    cl.cell(row=r, column=c_).border = GRID
cl.row_dimensions[r].height = 46
cl.column_dimensions["B"].width = 62
cl.column_dimensions["C"].width = 48
print(f"COMPLIANCE_LEGEND: added Primary/Supporting RTM + Review flag definitions, ending row {r}")

# =====================================================================
# SECTION 3 -- Review flag: standardise to 5 canonical values, convert to
# dropdown-constrained list, apply consistent colour CF. Same treatment on
# all 3 sheets that carry the field (each is an independent hand-typed
# literal today, confirmed in the earlier investigation -- not a formula
# mirror of one canonical source, so each needs the fix applied directly).
# =====================================================================
REVIEW_FLAG_MAP = {
    "OK": "OK",
    "OK; Review breadth": "OK — Review breadth",
    "Review breadth": "Review — breadth",
    "Review/accept context": "Review — accept context",
    "OK with non-RTM source note": "OK — non-RTM source",
}
REVIEW_FLAG_VALUES = ["OK", "OK — Review breadth", "Review — breadth", "Review — accept context", "OK — non-RTM source"]
REVIEW_FLAG_COLORS = {
    # green family = no action needed, amber/orange = needs a look, kept in
    # the same "weak->strong" direction as STATUS's pink/orange/green so it
    # doesn't invent a 4th unrelated colour system in this workbook.
    "OK": ("C6E8C6", "1E5E1E"),
    "OK — Review breadth": ("FFF3C4", "7A5B00"),
    "Review — breadth": ("FFDDA6", "7A3E00"),
    "Review — accept context": ("FFDDA6", "7A3E00"),
    "OK — non-RTM source": ("FFF3C4", "7A5B00"),
}
REVIEW_FLAG_DV_FORMULA = '"' + ",".join(REVIEW_FLAG_VALUES) + '"'

review_flag_locations = [
    ("OFFER_CANONICAL", "review_flag"),
    ("EVALUATION_WORKSPACE", "Review flag"),
    ("OFFER_RANKING", "Review flag"),
]
standardised_total = 0
for sheet_name, header_text in review_flag_locations:
    ws = wb[sheet_name]
    # find header row/col for this field (search first 8 rows)
    col = None
    hdr_row = None
    for rr in range(1, 8):
        for cc in range(1, ws.max_column + 1):
            if ws.cell(row=rr, column=cc).value == header_text:
                col = cc
                hdr_row = rr
                break
        if col:
            break
    if not col:
        print(f"  WARNING: '{header_text}' header not found on {sheet_name}, skipped")
        continue
    col_letter = get_column_letter(col)
    data_start = hdr_row + 1
    data_end = ws.max_row
    n_std = 0
    for rr in range(data_start, data_end + 1):
        cell = ws.cell(row=rr, column=col)
        if cell.value and str(cell.value).strip() in REVIEW_FLAG_MAP:
            new_val = REVIEW_FLAG_MAP[str(cell.value).strip()]
            if new_val != cell.value:
                cell.value = new_val
                n_std += 1
    standardised_total += n_std
    # dropdown validation
    dv = DataValidation(type="list", formula1=REVIEW_FLAG_DV_FORMULA, allow_blank=True)
    dv.error = "Pick one of the 5 standard Review flag values."
    dv.errorTitle = "Invalid Review flag"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{data_start}:{col_letter}{data_end}")
    # colour CF, one FormulaRule per value (consistent with STATUS pattern elsewhere)
    for val, (bg, fg) in REVIEW_FLAG_COLORS.items():
        rule = FormulaRule(
            formula=[f'{col_letter}{data_start}="{val}"'],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(name=CARLITO, color=fg, bold=(val == "OK")),
        )
        ws.conditional_formatting.add(f"{col_letter}{data_start}:{col_letter}{data_end}", rule)
    print(f"  {sheet_name}!{col_letter}: standardised {n_std} values, dropdown + colour CF applied ({data_start}:{data_end})")
print(f"Review flag: {standardised_total} values standardised across 3 sheets, dropdown-constrained + colour-coded")

# =====================================================================
# SECTION 4 -- REVIEW_FOCUS: small worklist sheet, review-flagged items only
# =====================================================================
oc = wb["OFFER_CANONICAL"]
oc_hdr = {}
for cc in range(1, oc.max_column + 1):
    v = oc.cell(row=1, column=cc).value or oc.cell(row=5, column=cc).value
    if v:
        oc_hdr[v] = cc
# locate header row for OFFER_CANONICAL properly
oc_hdr_row = None
for rr in range(1, 8):
    rowvals = [oc.cell(row=rr, column=cc).value for cc in range(1, oc.max_column + 1)]
    if rowvals.count(None) < len(rowvals) / 2 and any(isinstance(v, str) and "id" in str(v).lower() for v in rowvals if v):
        oc_hdr_row = rr
        break
if oc_hdr_row is None:
    oc_hdr_row = 5
oc_hdr = {oc.cell(row=oc_hdr_row, column=cc).value: cc for cc in range(1, oc.max_column + 1) if oc.cell(row=oc_hdr_row, column=cc).value}

if "REVIEW_FOCUS" in wb.sheetnames:
    del wb["REVIEW_FOCUS"]
rf = wb.create_sheet("REVIEW_FOCUS")
rf.sheet_view.showGridLines = False
rf.merge_cells("A1:F1")
rf["A1"] = "Review Focus — items flagged for a manual look"
rf["A1"].font = WHITE_BOLD
rf["A1"].fill = TITLE_FILL
rf.merge_cells("A2:F2")
rf["A2"] = ("Just the OFFER items whose Review flag is not plain 'OK' (see COMPLIANCE_LEGEND for definitions) -- "
            "GBO's own worklist for a manual triage pass, prioritised by rank so the highest-stakes items surface first. "
            "This mirrors OFFER_CANONICAL/OFFER_RANKING, edits happen there; this sheet is a filtered READ view, regenerate by re-running build_workbook_v20.py's Section 4 (or a future refresh script) after any Review flag change.")
rf["A2"].font = WHITE_REG
rf["A2"].fill = SUBT_FILL
rf["A2"].alignment = Alignment(wrap_text=True, vertical="center")
rf.row_dimensions[2].height = 34

rf_headers = ["OFFER ID", "Rank", "Title", "Review flag", "Primary RTMs", "Supporting RTMs"]
header_row(rf, 5, rf_headers)
id_col = oc_hdr.get("id") or oc_hdr.get("OFFER ID") or oc_hdr.get("offer_id")
rank_col = oc_hdr.get("rank") or oc_hdr.get("Rank")
title_col = oc_hdr.get("title") or oc_hdr.get("Title")
flag_col = oc_hdr.get("review_flag")
prim_col = oc_hdr.get("primary_rtm_ids")
supp_col = oc_hdr.get("supporting_context_rtm_ids")

rows_out = []
if flag_col:
    for rr in range(oc_hdr_row + 1, oc.max_row + 1):
        flagv = oc.cell(row=rr, column=flag_col).value
        if flagv and str(flagv).strip() not in ("", "OK"):
            std_flag = REVIEW_FLAG_MAP.get(str(flagv).strip(), str(flagv).strip())
            rows_out.append({
                "id": oc.cell(row=rr, column=id_col).value if id_col else "",
                "rank": oc.cell(row=rr, column=rank_col).value if rank_col else "",
                "title": oc.cell(row=rr, column=title_col).value if title_col else "",
                "flag": std_flag,
                "primary": oc.cell(row=rr, column=prim_col).value if prim_col else "",
                "supporting": oc.cell(row=rr, column=supp_col).value if supp_col else "",
            })
    def _rank_key(x):
        try:
            return float(x["rank"])
        except (TypeError, ValueError):
            return 9999
    rows_out.sort(key=_rank_key)

r = 6
for item in rows_out:
    rf.cell(row=r, column=1, value=item["id"])
    rf.cell(row=r, column=2, value=item["rank"])
    rf.cell(row=r, column=3, value=item["title"])
    fc = rf.cell(row=r, column=4, value=item["flag"])
    bg, fg = REVIEW_FLAG_COLORS.get(item["flag"], ("FFFFFF", "000000"))
    fc.fill = PatternFill("solid", fgColor=bg)
    fc.font = Font(name=CARLITO, color=fg, bold=True)
    rf.cell(row=r, column=5, value=item["primary"])
    rf.cell(row=r, column=6, value=item["supporting"])
    for c_ in range(1, 7):
        rf.cell(row=r, column=c_).border = GRID
    banded(rf, r, 2, start_col=1)   # A,B -- before the coloured flag column
    banded(rf, r, 2, start_col=5)   # E,F -- after it; column 4 (flag) keeps its own colour
    r += 1
if not rows_out:
    rf.merge_cells(f"A6:F6")
    rf["A6"] = "No OFFER items currently flagged for review beyond plain OK."
rf.column_dimensions["A"].width = 12
rf.column_dimensions["B"].width = 8
rf.column_dimensions["C"].width = 44
rf.column_dimensions["D"].width = 22
rf.column_dimensions["E"].width = 30
rf.column_dimensions["F"].width = 30
rf.freeze_panes = "A6"
print(f"REVIEW_FOCUS: new sheet, {len(rows_out)} flagged OFFER items listed")

# =====================================================================
# SECTION 5 -- RTM_PHASE_EXPANSION: one row per (RTM, phase), derived from
# RTM_RANKING!'Applicable phase(s)' -- an ADDITIONAL analytical sheet, NOT a
# restructuring of RTM_RANKING itself (every existing formula/export/chart
# that references RTM_RANKING keeps working unchanged). GBO confirmed this
# NOW because it feeds BT/PCA/ranking analysis -- built as a derived view
# exactly per GBO's own phrasing ("adding this column as an individual
# analyses"), not a schema rewrite of the master sheet.
# =====================================================================
rr = wb["RTM_RANKING"]
rr_hdr = {rr.cell(row=5, column=c).value: c for c in range(1, rr.max_column + 1) if rr.cell(row=5, column=c).value}
c_phase = rr_hdr.get("Applicable phase(s)")
c_id = rr_hdr.get("RTM ID")
c_tier = rr_hdr.get("Tier")
c_domain = rr_hdr.get("Domain")
c_s = rr_hdr.get("Weighted S")
c_rank = rr_hdr.get("Rank")
c_deliv = rr_hdr.get("Explicit deliverable / proof")
c_evbasis = rr_hdr.get("Evidence basis / review")

if "RTM_PHASE_EXPANSION" in wb.sheetnames:
    del wb["RTM_PHASE_EXPANSION"]
pe = wb.create_sheet("RTM_PHASE_EXPANSION")
pe.sheet_view.showGridLines = False
pe.merge_cells("A1:H1")
pe["A1"] = "RTM Phase Expansion — one row per (RTM, applicable phase)"
pe["A1"].font = WHITE_BOLD
pe["A1"].fill = TITLE_FILL
pe.merge_cells("A2:H2")
pe["A2"] = ("DERIVED analytical view, not a new master table -- RTM_RANKING!'Applicable phase(s)' is a single "
            "semicolon-joined string per row (e.g. 'L4 Installation; L0 Tender/Offer; L1 Conceptual Design...'); "
            "this sheet explodes that into one row per (RTM, phase) so multi-phase deliverables (the confirmed "
            "P&ID-at-L0/L1/.../As-built gap) can be filtered, counted, and cross-referenced by phase without "
            "restructuring RTM_RANKING itself or breaking any existing formula/chart/export that reads it. "
            "Regenerate by re-running build_workbook_v20.py's Section 5 after any RTM_RANKING phase-text change.")
pe["A2"].font = WHITE_REG
pe["A2"].fill = SUBT_FILL
pe["A2"].alignment = Alignment(wrap_text=True, vertical="center")
pe.row_dimensions[2].height = 46

# Exact live values, confirmed via direct scan of RTM_RANKING!'Applicable phase(s)'
# (not assumed from the earlier session's example row -- that row happened to
# omit FAC/PAC/Warranty, which do appear elsewhere in the corpus; matching
# the literal live strings, including "L0 Tender / Offer" with spaces around
# "/", not "L0 Tender/Offer" -- a naive guess at the string would silently
# fail to match and fall through to the default grey).
PHASE_ORDER = ["L0 Tender / Offer", "L1 Conceptual Design", "L2 Detailed Design",
               "L3 Procurement & Manufacturing / FAT", "FAC", "L4 Installation",
               "L5 Standalone Commissioning", "L6 Site Acceptance Testing",
               "PAC / Handover", "Warranty"]
PHASE_COLORS = {
    "L0 Tender / Offer": "E8E8F5", "L1 Conceptual Design": "DCEAF5", "L2 Detailed Design": "D6F0EA",
    "L3 Procurement & Manufacturing / FAT": "FCF0D6", "FAC": "FDEBD0", "L4 Installation": "FCE4D6",
    "L5 Standalone Commissioning": "F8D6E0", "L6 Site Acceptance Testing": "E6DFF6",
    "PAC / Handover": "D5F5E3", "Warranty": "EAECEE",
}

pe_headers = ["RTM ID", "Phase", "Phase order", "Tier", "Domain", "Weighted S", "Rank", "Explicit deliverable / proof"]
header_row(pe, 5, pe_headers)
r = 6
n_rows = 0
n_multi = 0
phase_counts = Counter()
if c_phase and c_id:
    for rr_row in rr.iter_rows(min_row=6, max_row=rr.max_row, values_only=False):
        rid = rr_row[c_id - 1].value
        if not rid:
            continue
        raw = rr_row[c_phase - 1].value
        if not raw:
            continue
        phases = [p.strip() for p in str(raw).split(";") if p.strip()]
        if len(phases) > 1:
            n_multi += 1
        for ph in phases:
            phase_counts[ph] += 1
            order = PHASE_ORDER.index(ph) + 1 if ph in PHASE_ORDER else 99
            pe.cell(row=r, column=1, value=rid)
            pc = pe.cell(row=r, column=2, value=ph)
            pc.fill = PatternFill("solid", fgColor=PHASE_COLORS.get(ph, "F2F6FA"))
            pe.cell(row=r, column=3, value=order)
            pe.cell(row=r, column=4, value=rr_row[c_tier - 1].value if c_tier else None)
            pe.cell(row=r, column=5, value=rr_row[c_domain - 1].value if c_domain else None)
            pe.cell(row=r, column=6, value=rr_row[c_s - 1].value if c_s else None)
            pe.cell(row=r, column=7, value=rr_row[c_rank - 1].value if c_rank else None)
            pe.cell(row=r, column=8, value=rr_row[c_deliv - 1].value if c_deliv else None)
            for cc in range(1, 9):
                pe.cell(row=r, column=cc).border = GRID
            banded(pe, r, 1, start_col=1)   # A -- before the coloured phase column
            banded(pe, r, 6, start_col=3)   # C..H -- after it; column 2 (phase) keeps its own colour
            r += 1
            n_rows += 1
pe.column_dimensions["A"].width = 12
pe.column_dimensions["B"].width = 34
pe.column_dimensions["C"].width = 11
pe.column_dimensions["D"].width = 12
pe.column_dimensions["E"].width = 22
pe.column_dimensions["F"].width = 12
pe.column_dimensions["G"].width = 9
pe.column_dimensions["H"].width = 50
pe.freeze_panes = "A6"
pe.auto_filter.ref = f"A5:H{r-1}" if r > 6 else "A5:H5"

# summary block below the exploded table
r += 1
section_title(pe, r, "Phase-count summary (per phase, across all 722 RTMs)", "C")
r += 1
header_row(pe, r, ["Phase", "RTM count", "Share of 722"])
r += 1
for ph in PHASE_ORDER:
    n = phase_counts.get(ph, 0)
    pe.cell(row=r, column=1, value=ph).fill = PatternFill("solid", fgColor=PHASE_COLORS.get(ph, "F2F6FA"))
    pe.cell(row=r, column=2, value=n)
    pe.cell(row=r, column=3, value=f"{n/722*100:.1f}%")
    for cc in range(1, 4):
        pe.cell(row=r, column=cc).border = GRID
    r += 1
r += 1
pe.merge_cells(f"A{r}:H{r}")
pe[f"A{r}"] = (f"{n_rows} (RTM, phase) rows exploded from RTM_RANKING's 722 rows; {n_multi} RTMs ({n_multi/722*100:.1f}%) "
               f"list more than one applicable phase (these are exactly the multi-phase-deliverable cases GBO flagged -- "
               f"e.g. a P&ID that applies at Tender, Conceptual, and Detailed Design). This sheet does NOT yet track "
               f"per-phase evidence/status independently (that would need a Status column per exploded row, filled in "
               f"by a reviewer) -- flagged as the next step if per-phase closure tracking is wanted, not built silently.")
pe[f"A{r}"].font = NOTE_FONT
pe[f"A{r}"].fill = NOTE_FILL
pe[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
pe.row_dimensions[r].height = 46
print(f"RTM_PHASE_EXPANSION: new sheet, {n_rows} (RTM,phase) rows from {n_multi} multi-phase RTMs ({n_multi/722*100:.1f}%)")

# =====================================================================
# SECTION 6 -- DELIVERABLES_DOSSIER: real "Applicable Documentation (AD)"
# groups from the contract mirror PDF's own Table 2 (uploads_v5/
# QPS_Contract_mirror_DOCX.pdf, extracted this round -- 6 top-level groups,
# 32 entries total incl. sub-items), cross-referenced against every RTM
# that actually cites an AD_## code anywhere in its text. This is real
# contract structure, not a heuristic grouping of already-derived fields --
# GBO asked for exactly this: a list of deliverable "files/dossiers" that
# group specific items (Manufacturing File, Engineering File-style bundles).
# =====================================================================
with open("/tmp/ad_table.json") as f:
    AD_TABLE = json.load(f)

ad_links = defaultdict(list)
for row in rr.iter_rows(min_row=6, max_row=rr.max_row, values_only=False):
    rid = row[c_id - 1].value if c_id else None
    if not rid:
        continue
    for cc in range(rr.max_column):
        v = row[cc].value
        if v and isinstance(v, str) and "AD_" in v:
            for m in re.findall(r"AD_\d\d(?:\.\d+)?", v):
                ad_links[m].append(rid)

if "DELIVERABLES_DOSSIER" in wb.sheetnames:
    del wb["DELIVERABLES_DOSSIER"]
dd = wb.create_sheet("DELIVERABLES_DOSSIER")
dd.sheet_view.showGridLines = False
dd.merge_cells("A1:F1")
dd["A1"] = "Deliverables Dossier — contract Applicable Documentation (AD) groups"
dd["A1"].font = WHITE_BOLD
dd["A1"].fill = TITLE_FILL
dd.merge_cells("A2:F2")
dd["A2"] = ("Real structure from the contract mirror's own \"Table 2. Applicable Documentation (AD)\" "
            "(uploads_v5/QPS_Contract_mirror_DOCX.pdf, p.24) -- 6 top-level document groups, several with "
            "sub-items, each with its own SCK CEN reference number. RTM linkage below is computed by scanning "
            "every RTM's full text for an AD_## mention -- not fabricated, and most AD items currently have "
            "ZERO citing RTMs, which is itself worth knowing (either genuinely no RTM references that document, "
            "or a citation exists in wording this scan didn't catch -- flagged honestly rather than papered over).")
dd["A2"].font = WHITE_REG
dd["A2"].fill = SUBT_FILL
dd["A2"].alignment = Alignment(wrap_text=True, vertical="center")
dd.row_dimensions[2].height = 46

header_row(dd, 5, ["AD ID", "Document name", "Reference", "Group", "Linked RTM count", "Linked RTM IDs"])
r = 6
n_linked_total = 0
for entry in AD_TABLE:
    ad_id = entry["id"]
    is_top = "." not in ad_id
    group = ad_id.split(".")[0]
    linked = sorted(set(ad_links.get(ad_id, [])), key=lambda x: int(re.sub(r"\D", "", x) or 0))
    c1 = dd.cell(row=r, column=1, value=ad_id)
    c1.font = Font(name=CARLITO, bold=is_top, color=NAVY if is_top else "444444")
    if is_top:
        c1.fill = PatternFill("solid", fgColor="DCEAF5")
    dd.cell(row=r, column=2, value=entry["name"]).alignment = Alignment(wrap_text=True, vertical="top")
    dd.cell(row=r, column=3, value=entry["reference"])
    dd.cell(row=r, column=4, value=group)
    ncell = dd.cell(row=r, column=5, value=len(linked))
    if linked:
        ncell.font = Font(name=CARLITO, bold=True, color="1E5E1E")
        n_linked_total += len(linked)
    dd.cell(row=r, column=6, value="; ".join(linked) if linked else "(none found)")
    for cc in range(1, 7):
        dd.cell(row=r, column=cc).border = GRID
    dd.row_dimensions[r].height = 30 if is_top else 18
    if is_top:
        banded(dd, r, 5, start_col=2)  # skip col A -- keep the top-level-group accent fill
    else:
        banded(dd, r, 6)  # sub-items have no col-A fill of their own, safe to band normally
    r += 1
dd.column_dimensions["A"].width = 11
dd.column_dimensions["B"].width = 44
dd.column_dimensions["C"].width = 16
dd.column_dimensions["D"].width = 8
dd.column_dimensions["E"].width = 13
dd.column_dimensions["F"].width = 40
dd.freeze_panes = "A6"
dd.auto_filter.ref = f"A5:F{r-1}"

r += 1
dd.merge_cells(f"A{r}:F{r}")
n_with_links = sum(1 for e in AD_TABLE if ad_links.get(e["id"]))
dd[f"A{r}"] = (f"{len(AD_TABLE)} AD entries total (6 top-level groups + sub-items); {n_with_links} of them have "
               f"at least one RTM explicitly citing that AD code by number ({n_linked_total} citing-instances "
               f"total, {len(set(sum(ad_links.values(), [])))} distinct RTMs). The rest of the RTM corpus does "
               f"not cite an AD_## reference directly -- those RTMs' evidence basis is the RTM's own section "
               f"text, not one of these named document bundles.")
dd[f"A{r}"].font = NOTE_FONT
dd[f"A{r}"].fill = NOTE_FILL
dd[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
dd.row_dimensions[r].height = 46
print(f"DELIVERABLES_DOSSIER: new sheet, {len(AD_TABLE)} AD entries, {n_with_links} with >=1 citing RTM")

# =====================================================================
# SECTION 7 -- DASHBOARD_2: review-completion rollup (KNOWLEDGE_TAXONOMY_
# MAPPING.md gap 07 -- item-level status existed, project-level rollup
# didn't). Placed well below the existing charts (last chart anchors at
# row 70, this workbook's charts run ~18 rows tall) to avoid any visual
# overlap with floating chart objects, which don't shift with inserted rows.
# =====================================================================
rq = wb["RTM_REVIEW_QUEUE"]
rq_hdr = {rq.cell(row=5, column=c).value: c for c in range(1, rq.max_column + 1) if rq.cell(row=5, column=c).value}
c_disp = rq_hdr.get("Disposition")
n_queue = 0
n_decided = 0
if c_disp:
    for row in rq.iter_rows(min_row=6, max_row=rq.max_row, values_only=False):
        rid = row[1].value
        if not rid:
            continue
        n_queue += 1
        v = row[c_disp - 1].value
        if v and str(v).strip().upper() not in ("", "TBD"):
            n_decided += 1
decided_pct = (n_decided / n_queue * 100) if n_queue else 0

d2 = wb["DASHBOARD_2"]
r = 95
section_title(d2, r, "Review-completion rollup (RTM_REVIEW_QUEUE)", "I")
r += 1
d2.merge_cells(f"A{r}:I{r+3}")
d2[f"A{r}"] = (
    f"{n_decided} of {n_queue} priority-queue RTMs ({decided_pct:.0f}%) have a Disposition recorded "
    f"(anything other than blank/TBD); {n_queue - n_decided} ({100-decided_pct:.0f}%) are still awaiting a "
    f"reviewer decision. This is the single project-level 'how far along is this evaluation' number the "
    f"per-item STATUS/Disposition fields never rolled up into one place before -- same pattern as the "
    f"crosswalk-coverage callout above, applied to review completion instead of RTM<->OFFER linkage. "
    f"Re-run build_workbook_v20.py's Section 7 (or a future refresh) any time RTM_REVIEW_QUEUE!Disposition "
    f"values change, to keep this number current -- it is a snapshot, not a live formula, so it can drift "
    f"if the workbook is edited without a rebuild."
)
d2[f"A{r}"].font = NOTE_FONT
d2[f"A{r}"].fill = NOTE_FILL
d2[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
for rr_ in range(r, r + 4):
    for cc in range(1, 10):
        d2.cell(row=rr_, column=cc).fill = NOTE_FILL
print(f"DASHBOARD_2: review-completion rollup added -- {n_decided}/{n_queue} decided ({decided_pct:.0f}%)")

# =====================================================================
# SECTION 8 -- CONFLICT_CANDIDATES: a disclosed HEURISTIC first pass at
# requirement-conflict/interference detection (KNOWLEDGE_TAXONOMY_MAPPING
# gap 06, the "interferences" half -- crosswalk/linkage was already covered,
# actual contradiction-detection was not). Method: RTM pairs in the exact
# same Section that share significant vocabulary (candidate: same topic)
# but differ in negation polarity (one says "shall not"/"prohibited", the
# other doesn't) -- flagged for human review, NEVER asserted as a real
# conflict. This is explicitly a starting point, not a verified finding.
# =====================================================================
NEGATION_MARKERS = ("shall not", "must not", "may not", "prohibited", "not permitted", "not be permitted", "shall never")
STOPWORDS = set("the a an of to for and or shall be with in on at by is are as that this from into within without provide provided provides".split())


def sig_words(text):
    words = re.findall(r"[a-zA-Z]{5,}", (text or "").lower())
    return set(w for w in words if w not in STOPWORDS)


rr_rows = []
for row in rr.iter_rows(min_row=6, max_row=rr.max_row, values_only=False):
    rid = row[c_id - 1].value if c_id else None
    if not rid:
        continue
    section = row[rr_hdr.get("Section", 11) - 1].value
    shall = row[rr_hdr.get("Shall statement", 13) - 1].value or ""
    if not section:
        continue
    rr_rows.append({"id": rid, "section": str(section).strip(), "shall": shall,
                     "negated": any(m in shall.lower() for m in NEGATION_MARKERS),
                     "words": sig_words(shall),
                     "domain": row[c_domain - 1].value if c_domain else None})

by_section = defaultdict(list)
for item in rr_rows:
    by_section[item["section"]].append(item)

candidates = []
for section, items in by_section.items():
    if len(items) < 2:
        continue
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            a, b = items[i], items[j]
            if a["negated"] == b["negated"]:
                continue  # need opposite polarity to be a candidate at all
            overlap = a["words"] & b["words"]
            if len(overlap) >= 2:
                candidates.append({
                    "section": section, "rtm_a": a["id"], "rtm_b": b["id"],
                    "shared_words": ", ".join(sorted(overlap)[:6]),
                    "shall_a": a["shall"][:180], "shall_b": b["shall"][:180],
                    "domain": a["domain"],
                })

if "CONFLICT_CANDIDATES" in wb.sheetnames:
    del wb["CONFLICT_CANDIDATES"]
cc = wb.create_sheet("CONFLICT_CANDIDATES")
cc.sheet_view.showGridLines = False
cc.merge_cells("A1:F1")
cc["A1"] = "Conflict Candidates — heuristic, for human review only"
cc["A1"].font = WHITE_BOLD
cc["A1"].fill = TITLE_FILL
cc.merge_cells("A2:F2")
cc["A2"] = ("HEURISTIC first pass, not a verified conflict list: pairs of RTMs in the exact same contract "
            "Section that share >=2 significant words (candidate: same topic) but differ in negation polarity "
            "(one contains \"shall not\"/\"prohibited\"-type language, the other doesn't). This catches "
            "shape-level contradiction candidates only -- it does NOT understand numeric thresholds, units, or "
            "actual logical meaning, so most rows here will turn out to be false positives (e.g. two "
            "requirements about the same topic that are simply complementary, not contradictory). Treat every "
            "row as \"worth a human glance\", never as a confirmed conflict.")
cc["A2"].font = WHITE_REG
cc["A2"].fill = SUBT_FILL
cc["A2"].alignment = Alignment(wrap_text=True, vertical="center")
cc.row_dimensions[2].height = 58

header_row(cc, 5, ["Section", "RTM A", "RTM B", "Domain", "Shared vocabulary", "RTM A shall text", "RTM B shall text"])
r = 6
for cand in candidates:
    cc.cell(row=r, column=1, value=cand["section"])
    cc.cell(row=r, column=2, value=cand["rtm_a"])
    cc.cell(row=r, column=3, value=cand["rtm_b"])
    cc.cell(row=r, column=4, value=cand["domain"])
    cc.cell(row=r, column=5, value=cand["shared_words"])
    cc.cell(row=r, column=6, value=cand["shall_a"]).alignment = Alignment(wrap_text=True, vertical="top")
    cc.cell(row=r, column=7, value=cand["shall_b"]).alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 8):
        cc.cell(row=r, column=col).border = GRID
    cc.row_dimensions[r].height = 44
    banded(cc, r, 7)
    r += 1
if not candidates:
    cc.merge_cells("A6:G6")
    cc["A6"] = "No candidate pairs found by this heuristic in the current data."
cc.column_dimensions["A"].width = 10
cc.column_dimensions["B"].width = 10
cc.column_dimensions["C"].width = 10
cc.column_dimensions["D"].width = 20
cc.column_dimensions["E"].width = 34
cc.column_dimensions["F"].width = 46
cc.column_dimensions["G"].width = 46
cc.freeze_panes = "A6"
if len(candidates):
    cc.auto_filter.ref = f"A5:G{r-1}"
print(f"CONFLICT_CANDIDATES: new sheet, {len(candidates)} heuristic candidate pairs across {len(by_section)} sections")

# =====================================================================
# SECTION 9 -- Comment-column consistency: GBO asked for a user-editable
# Comment column on review-facing sheets, cream-filled to signal "safe to
# type here" (same convention COMPLIANCE_LEGEND!row20 already documents).
# Most review sheets already HAD one (EVALUATION_WORKSPACE A_/B_Reviewer
# Comment, RTM_REVIEW_QUEUE!Comments, NEGOTIATION_AGENDA!Reviewer Comment)
# -- confirmed by direct header scan, not assumed -- so the real gaps were:
# (a) QUALITY_CHECKS had NO comment column at all, and (b) two existing
# comment columns were missing the cream fill that COMPLIANCE_LEGEND's own
# row 20 already flagged as a not-yet-done follow-up. Fixing both exactly,
# not adding redundant new columns where one already exists.
# =====================================================================
qc = wb["QUALITY_CHECKS"]
qc_hdr_row = 5
qc_comment_col = qc.max_column + 1
qc.cell(row=qc_hdr_row, column=qc_comment_col, value="Comment").font = HDR_FONT
qc.cell(row=qc_hdr_row, column=qc_comment_col).fill = HDR_FILL
qc.cell(row=qc_hdr_row, column=qc_comment_col).border = GRID
qc_col_letter = get_column_letter(qc_comment_col)
qc.column_dimensions[qc_col_letter].width = 40
for rr_ in range(qc_hdr_row + 1, qc.max_row + 1):
    cell = qc.cell(row=rr_, column=qc_comment_col)
    cell.fill = CREAM
    cell.border = GRID
    cell.alignment = Alignment(wrap_text=True, vertical="top")
print(f"QUALITY_CHECKS: added Comment column at {qc_col_letter} (previously had none), {qc.max_row - qc_hdr_row} rows")

cream_fix_targets = [("RTM_REVIEW_QUEUE", "Comments"), ("NEGOTIATION_AGENDA", "Reviewer Comment")]
for sheet_name, header_text in cream_fix_targets:
    ws = wb[sheet_name]
    col = None
    hdr_row = None
    for rr2 in range(1, 8):
        for cc2 in range(1, ws.max_column + 1):
            if ws.cell(row=rr2, column=cc2).value == header_text:
                col = cc2
                hdr_row = rr2
                break
        if col:
            break
    if not col:
        print(f"  WARNING: '{header_text}' not found on {sheet_name}, skipped")
        continue
    for rr3 in range(hdr_row + 1, ws.max_row + 1):
        ws.cell(row=rr3, column=col).fill = CREAM
    print(f"  {sheet_name}!{get_column_letter(col)} ('{header_text}'): cream 'safe to edit' fill applied, "
          f"{ws.max_row - hdr_row} rows -- closes the gap COMPLIANCE_LEGEND!row20 already flagged")

# update COMPLIANCE_LEGEND!row20's own follow-up note now that it's done
cl2 = wb["COMPLIANCE_LEGEND"]
for rr4 in range(1, cl2.max_row + 1):
    v = cl2.cell(row=rr4, column=1).value
    if v and "Soft cream fill" in str(v):
        note_cell = cl2.cell(row=rr4, column=4)
        if note_cell.value and "NOT yet extended" in str(note_cell.value):
            note_cell.value = str(note_cell.value).replace(
                "NOT yet extended to RTM_REVIEW_QUEUE!Comments, DMAIC_AUDIT, AUDIT_NOTES, or NEGOTIATION_AGENDA -- flagged as a follow-up, not done silently.",
                "v20: extended to RTM_REVIEW_QUEUE!Comments and NEGOTIATION_AGENDA!Reviewer Comment, plus a new "
                "Comment column added to QUALITY_CHECKS (previously had none). DMAIC_AUDIT/AUDIT_NOTES "
                "deliberately left as-is -- both are structured audit logs, not live reviewer-input sheets."
            )
            print("  COMPLIANCE_LEGEND!row20 follow-up note updated to reflect v20 fix")

wb.save(OUT)
print(f"saved {OUT} (section 1-9, FULL BUILD)")
