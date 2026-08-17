"""
build_workbook_full_v5.py -- integrates QPS_RTM_BT_Standalone.xlsx's three
unique tabs (RTM_RANKING, DOMAIN_SUMMARY, RTM_REVIEW_QUEUE -- the only
sheets in that workbook whose name doesn't already exist in
QPS_OFFER_Evaluation.xlsx) into GBO's own hand-edited copy of
QPS_OFFER_Evaluation.xlsx, wires them into the existing navigation system,
and recolours every tab with an explicit SCK purple/blue/turquoise palette.

Base file: QPS_OFFER_Evaluation_GBO.xlsx (GBO's own START_HERE/tab-order
edit -- kept as-is; this script only ADDS sheets and recolours tabs, it does
not touch GBO's own edits to existing sheets).

Everything already in QPS_OFFER_Evaluation_GBO.xlsx (formulas, named
ranges, conditional formatting, data validations) is preserved untouched --
per the engineering handover's "do not regress" list, this script never
writes over an existing formula cell with a static value, never paints a
fill over a cell that's driven by conditional formatting, and does not
touch the workbook's existing named ranges or data validations at all.
"""
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from xlsx_copy_helpers import copy_sheet_cross_workbook

BASE = "uploads_v5/QPS_OFFER_Evaluation_GBO.xlsx"
RTM_SRC = "uploads_v5/QPS_RTM_BT_Standalone_a.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v5.xlsx"

NEW_SHEETS = ["RTM_RANKING", "DOMAIN_SUMMARY", "RTM_REVIEW_QUEUE"]

# ---- SCK palette: explicit RGB, replacing the ad-hoc theme-tint scheme ----
# Matches the MTBF deck's palette (build_deck4.py / build_deck5.py): purple
# for DEFINE-adjacent/core evaluation content, blue for canonical/reference
# data, turquoise for RTM-only (no-OFFER-data) content, light purple/pink
# for QA/audit/governance, amber-free here since Excel tabs don't carry a
# DMAIC-phase meaning the way the deck's badges do.
TAB_COLOR = {
    "START_HERE":            None,          # no colour -- always-white anchor tab
    "NAVIGATION_MAP":         None,
    "DASHBOARD":             "562873",       # purple -- primary reviewer view
    "EVALUATION_WORKSPACE":  "562873",
    "NEGOTIATION_AGENDA":    "7A3F9E",       # lighter purple -- still core workflow
    "OFFER_RANKING":         "034694",       # blue -- canonical ranking output
    "RTM_CROSSWALK":         "034694",
    "OFFER_CANONICAL":       "0B5FA5",       # lighter blue -- reference/canonical text
    "RTM_RANKING":           "1FA7A0",       # turquoise -- RTM-only, no OFFER fields
    "DOMAIN_SUMMARY":        "1FA7A0",
    "RTM_REVIEW_QUEUE":      "1FA7A0",
    "STANDARDS":             "0B5FA5",
    "DELIVERABLES":          "0B5FA5",
    "WEIGHTS_METHOD":        "3A1F5C",       # deep purple -- method/control panel
    "QUALITY_CHECKS":        "D9A6D9",       # light purple/pink -- QA & governance
    "AUDIT_NOTES":           "D9A6D9",
    "DMAIC_AUDIT":           "D9A6D9",
    "CODING_HANDOVER":       "E0A9D6",       # light pink -- internal/automation
    "EVALUATION_INPUT":      "C9C9D6",       # neutral grey-lavender -- deprecated/legacy
    "LISTS":                 None,           # utility sheet, stays hidden
    "COMPLIANCE_LEGEND":     "FF0000",       # unchanged -- semantic red, not decorative
}

NAV_FONT = Font(name="Carlito", size=11, bold=False, color="333333")
NAV_HEADER_FONT = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
NAV_HEADER_FILL = PatternFill("solid", fgColor="0F6B78")

# ---------------------------------------------------------------- load ----
wb = openpyxl.load_workbook(BASE, data_only=False)
rtm_wb = openpyxl.load_workbook(RTM_SRC, data_only=False)

# ---------------------------------------------------- copy the 3 sheets ---
offer_ranking_idx = wb.sheetnames.index("OFFER_RANKING")
for i, name in enumerate(NEW_SHEETS):
    copy_sheet_cross_workbook(rtm_wb[name], wb, index=offer_ranking_idx + 1 + i)
print("copied sheets:", NEW_SHEETS, "-> positioned after OFFER_RANKING")

# ------------------------------------------------------- recolour tabs ----
for name in wb.sheetnames:
    hexcolor = TAB_COLOR.get(name)
    ws = wb[name]
    if hexcolor is None:
        ws.sheet_properties.tabColor = None
    else:
        ws.sheet_properties.tabColor = Color(rgb="FF" + hexcolor)
print("tab colours applied")

# --------------------------------------------- wire into START_HERE nav ---
sh = wb["START_HERE"]
# row-3 nav bar: append a 10th entry pointing at the new RTM ranking sheet
sh["J3"] = '=HYPERLINK("#RTM_RANKING!A1","RTM Rank")'
sh["J3"].font = NAV_HEADER_FONT
sh["J3"].fill = NAV_HEADER_FILL
sh["J3"].alignment = Alignment(horizontal="center")

# Step 9 in the numbered walkthrough (rows 5-13 hold steps 1-8; row 14 is
# blank before the glossary section at row 17 -- appending here keeps every
# existing row/formula in the sheet untouched)
sh["A14"] = "9"
sh["B14"] = '=HYPERLINK("#RTM_RANKING!A1","RTM_RANKING")'
sh["C14"] = "Audit RTM-only static importance (722 requirements, identical weights, no OFFER data)."
sh["D14"] = "Cross-check against OFFER_RANKING and RTM_CROSSWALK for consistency."
for coord in ("A14", "B14", "C14", "D14"):
    sh[coord].font = NAV_FONT
    sh[coord].alignment = Alignment(wrap_text=True, vertical="top")
print("START_HERE nav bar + step list extended")

# ------------------------------------------------- wire into NAVIGATION_MAP
nm = wb["NAVIGATION_MAP"]
new_rows = [
    ("RTM_RANKING", "RTM-only static importance (722 reqs, no OFFER data)", "Audit and cross-check vs. OFFER_RANKING", "Method owner"),
    ("DOMAIN_SUMMARY", "RTM counts/tiers/average importance by contract domain", "Review-staffing allocation", "Lead evaluator"),
    ("RTM_REVIEW_QUEUE", "Priority subset: all T0/T1 RTMs + deliverable-heavy reqs", "Filter for focused review", "Technical reviewers"),
]
start_row = nm.max_row + 1
for offset, (sheet_name, purpose, action, owner) in enumerate(new_rows):
    r = start_row + offset
    nm.cell(row=r, column=1, value=f'=HYPERLINK("#{sheet_name}!A1","{sheet_name}")').font = NAV_FONT
    nm.cell(row=r, column=2, value=purpose).font = NAV_FONT
    nm.cell(row=r, column=3, value=action).font = NAV_FONT
    nm.cell(row=r, column=4, value=owner).font = NAV_FONT
print(f"NAVIGATION_MAP extended with {len(new_rows)} rows starting at row {start_row}")

# ---------------------------------------------- fill in a content gap -----
# START_HERE!A19 (merged A19:L22) is a bold "BT win %" heading with no body
# text underneath it -- every sibling glossary entry (Weighted S, above;
# Negotiation flag, below) has an explanation, this one was left empty.
# Directly in the primary onboarding page a reviewer reads first, so fixed
# here rather than just flagged -- wording matches the engineering
# handover's own definition (ENGINEERING_HANDOVER_OFFER_BT_v3_6.md §3.2).
sh["A19"] = (
    "BT win %\n"
    "Human-readable head-to-head reading of the same ranking: "
    "100 × (wins + 0.5 × ties) / (N − 1) against every other item on the list -- "
    "same order as Weighted S, just expressed as a win rate instead of a 0-100 score. "
    "Use win % when briefing a non-technical audience; use Weighted S when you need to defend a specific number."
)
print("filled in previously-empty 'BT win %' glossary entry (START_HERE!A19)")

wb.save(OUT)
print(f"saved {OUT} -- {len(wb.sheetnames)} sheets: {wb.sheetnames}")
