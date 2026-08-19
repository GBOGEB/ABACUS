# -*- coding: utf-8 -*-
"""
4th weight-sensitivity scenario, requested as a follow-on to
NEXT_ITERATION_BACKLOG.md Section 19 (Base / Equal / Cost=70%-flat already
built). WEIGHTS_METHOD row 22 itself names this exact variant as "not
computed here... ask if you want that variant too":

    "keep the other 6 dimensions' CURRENT relative proportions and rescale
    them to fit the remaining 30%"

Scenario definition: Cost = 70%. The other 6 dimensions keep their existing
frozen relative RATIOS to each other (L:R:P:F:Q:LC = .20:.22:.20:.16:.12:.07,
summing to 0.97), rescaled so together they fill the remaining 30% instead
of summing to 0.97. weight_new(d) = base_weight(d) * (0.30 / 0.97) for the
6 non-Cost dims.

Method matches Section 19 exactly: recomputed from OFFER_RANKING!H:N's real
0-3 relevance scores (50 items, live from the workbook) -- nothing invented,
same Weighted S formula (100 x sum(weight x relevance/3)) as WEIGHTS_METHOD
row 6.

Usage:
    python compute_weight_scenario4.py <FULL_vNN.xlsx> <output.json>
"""
import sys
import json
import numpy as np
import openpyxl

BASE_WEIGHTS = {"L": 0.20, "R": 0.22, "P": 0.20, "F": 0.16, "Q": 0.12, "LC": 0.07, "C": 0.03}
DIM_ORDER = ["L", "R", "P", "F", "Q", "LC", "C"]
COL_OFFER_ID = 2
COL_TITLE = 3
COL_DIMS_START = 8  # H = L, through N = C (7 columns)
HEADER_ROW = 5
DATA_START_ROW = 6


def scenario4_weights():
    non_cost = {d: w for d, w in BASE_WEIGHTS.items() if d != "C"}
    non_cost_sum = sum(non_cost.values())  # 0.97
    scale = 0.30 / non_cost_sum
    scaled = {d: round(w * scale, 5) for d, w in non_cost.items()}
    scaled["C"] = 0.70
    return scaled


def weighted_s(scores, weights):
    return 100.0 * sum((scores[d] / 3.0) * weights[d] for d in DIM_ORDER)


def spearman(a, b):
    a, b = np.array(a, dtype=float), np.array(b, dtype=float)
    ra = a.argsort().argsort().astype(float)
    rb = b.argsort().argsort().astype(float)
    return float(np.corrcoef(ra, rb)[0, 1])


def main():
    if len(sys.argv) != 3:
        print("usage: python compute_weight_scenario4.py <FULL_vNN.xlsx> <output.json>")
        sys.exit(1)
    xlsx_path, out_path = sys.argv[1], sys.argv[2]

    weights4 = scenario4_weights()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["OFFER_RANKING"]
    items = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        offer_id = ws.cell(row=r, column=COL_OFFER_ID).value
        if offer_id is None:
            continue
        title = ws.cell(row=r, column=COL_TITLE).value
        base_rank = ws.cell(row=r, column=1).value
        dim_vals = {d: ws.cell(row=r, column=COL_DIMS_START + i).value for i, d in enumerate(DIM_ORDER)}
        if any(v is None for v in dim_vals.values()):
            continue
        base_s = ws.cell(row=r, column=15).value  # O = Weighted S
        s4 = weighted_s(dim_vals, weights4)
        items.append({
            "offerId": offer_id, "title": title, "baseRank": base_rank,
            "baseWeightedS": base_s, "scenario4WeightedS": round(s4, 4),
        })

    # Rank by scenario4 S descending (ties broken by base rank, matching
    # the project's existing tie-break convention of stable-sort-by-base-order)
    items_sorted = sorted(items, key=lambda it: (-it["scenario4WeightedS"], it["baseRank"]))
    for i, it in enumerate(items_sorted, start=1):
        it["scenario4Rank"] = i
        it["rankDelta"] = it["baseRank"] - i  # positive = moved up (more important)

    base_ranks = [it["baseRank"] for it in items_sorted]
    s4_ranks = [it["scenario4Rank"] for it in items_sorted]
    rho = spearman(base_ranks, s4_ranks)

    result = {
        "source_workbook": xlsx_path,
        "scenario_name": "Cost=70%, remaining 30% proportional to existing frozen ratios",
        "weights": weights4,
        "n_items": len(items_sorted),
        "spearman_base_vs_scenario4": round(rho, 4),
        "items": sorted(items_sorted, key=lambda it: it["scenario4Rank"]),
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)

    top_movers = sorted(items_sorted, key=lambda it: -it["rankDelta"])[:5]
    print(f"wrote {out_path}: n={len(items_sorted)}, spearman(base,scenario4)={rho:.4f}")
    print("top 5 movers UP:", [(it["offerId"], it["rankDelta"]) for it in top_movers])


if __name__ == "__main__":
    main()
