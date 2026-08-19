# -*- coding: utf-8 -*-
"""Adds a cluster-level aggregate importance view to the CLUSTERS sheet.

NOT part of the original v5..v24 build_workbook_vNN.py chain -- added
2026-08-19 (ABACUS-side) per GBO's explicit scoping decision on the
standing "should the BT method rank at cluster level too?" backlog
question (SESSION_SSOT.yaml status.still_pending_this_session). GBO chose:
(a) both "rank clusters by aggregate Weighted S" AND "DMAIC-targeting /
concentration lens" framing -- not mutually exclusive; (b) show BOTH Sum
and Average Weighted S, matching the Sum-vs-Average pattern this workbook
already uses on DASHBOARD_2 for domains.

Does NOT touch the existing item-level Tier/Rank/Gate system in any way --
this is a separate, clearly-dated, disclosed aggregate view, exactly as
CLUSTERS!A2's own existing text already draws that distinction for the
original per-item design.

Run against FULL_v24.xlsx and LITE_v24.xlsx (CLUSTERS sheet is identical in
both).
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# Resolved relative to this script's own location, never cwd -- matches the
# project's own path-resolution control rule (DMAIC_BUGFIX_LOG.md rule #7).
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILES = [
    os.path.join(BASE_DIR, "current", "QPS_OFFER_Evaluation_FULL_v24.xlsx"),
    os.path.join(BASE_DIR, "current", "QPS_OFFER_Evaluation_LITE_v24.xlsx"),
]

HEADER_FILL = PatternFill("solid", fgColor="FF2F4858")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=12)
NOTE_FONT = Font(italic=True, color="FF555555")
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def compute_cluster_aggregates(wb):
    ws_offer = wb["OFFER_RANKING"]
    s_by_id = {}
    for r in range(6, ws_offer.max_row + 1):
        oid = ws_offer.cell(row=r, column=2).value
        if oid is None:
            continue
        s_by_id[oid] = ws_offer.cell(row=r, column=15).value

    ws_clusters = wb["CLUSTERS"]
    rows = []
    for r in range(6, 14):
        cid = ws_clusters.cell(row=r, column=1).value
        name = ws_clusters.cell(row=r, column=2).value
        members_str = ws_clusters.cell(row=r, column=3).value
        if not cid:
            continue
        members = [m.strip() for m in members_str.split(";")]
        scores = [s_by_id[m] for m in members if m in s_by_id]
        total = sum(scores)
        avg = total / len(scores) if scores else 0
        rows.append(dict(cid=cid, name=name, count=len(members),
                          total=round(total, 2), avg=round(avg, 2)))

    by_sum = sorted(rows, key=lambda x: -x["total"])
    for i, row in enumerate(by_sum, start=1):
        row["rank_sum"] = i
    by_avg = sorted(rows, key=lambda x: -x["avg"])
    for i, row in enumerate(by_avg, start=1):
        row["rank_avg"] = i

    return sorted(rows, key=lambda x: x["cid"])


def write_section(ws, rows):
    start = ws.max_row + 3

    ws.cell(row=start, column=1).value = (
        "Cluster-level aggregate importance (added 2026-08-19, ABACUS-side)"
    )
    ws.cell(row=start, column=1).font = TITLE_FONT

    ws.cell(row=start + 1, column=1).value = (
        "Answers a standing backlog question (does the BT method rank at cluster level?) "
        "GBO scoped explicitly: yes, as an aggregate Weighted S view (not a new weighting "
        "scheme), AND as a DMAIC-targeting/concentration lens -- which clusters carry the "
        "most total contract weight vs. which clusters' typical item matters most, "
        "independent of size. Does NOT change or override the item-level Tier/Rank/Gate "
        "system above -- clusters still do not gate or override individual item ranking; "
        "this is a separate, disclosed view for review-staffing and improvement-focus "
        "decisions. Sum = total Weighted S across cluster members (favors larger clusters, "
        "answers 'which cluster carries the most weight'). Average = per-item mean (size-"
        "independent, answers 'which cluster's typical item matters most') -- same "
        "Sum-vs-Average distinction already used on DASHBOARD_2 for domains."
    )
    ws.cell(row=start + 1, column=1).font = NOTE_FONT
    ws.merge_cells(start_row=start + 1, start_column=1, end_row=start + 1, end_column=8)
    ws.row_dimensions[start + 1].height = 60

    header_row = start + 3
    headers = ["Cluster ID", "Cluster Name", "Item count", "Sum Weighted S",
               "Rank (Sum)", "Avg Weighted S", "Rank (Avg)", "Read"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i)
        c.value = h
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER

    for offset, row in enumerate(rows, start=1):
        r = header_row + offset
        divergence = row["rank_sum"] - row["rank_avg"]
        if divergence >= 3:
            read = "Small but concentrated -- higher priority per-item than its total volume suggests"
        elif divergence <= -3:
            read = "Large by volume, lower average -- importance driven by item count, not concentration"
        else:
            read = "Sum and average agree -- no notable volume/concentration divergence"
        vals = [row["cid"], row["name"], row["count"], row["total"],
                row["rank_sum"], row["avg"], row["rank_avg"], read]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i)
            c.value = v
            c.border = BORDER

    # color scales on the two numeric importance columns
    sum_col_range = f"D{header_row + 1}:D{header_row + len(rows)}"
    avg_col_range = f"F{header_row + 1}:F{header_row + len(rows)}"
    for rng in (sum_col_range, avg_col_range):
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(start_type="min", start_color="FFF8D7DA",
                            end_type="max", end_color="FFD4EDDA"),
        )

    widths = {1: 12, 2: 26, 3: 11, 4: 15, 5: 11, 6: 15, 7: 11, 8: 52}
    for col, w in widths.items():
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width is None or ws.column_dimensions[letter].width < w:
            ws.column_dimensions[letter].width = w

    return header_row + len(rows)


def main():
    for path in FILES:
        wb = openpyxl.load_workbook(path, data_only=False)
        rows = compute_cluster_aggregates(wb)
        print(f"=== {path} ===")
        for row in sorted(rows, key=lambda x: x["rank_sum"]):
            print(f"  {row['cid']} {row['name']:<30} sum={row['total']:>7} (#{row['rank_sum']})  "
                  f"avg={row['avg']:>6} (#{row['rank_avg']})")

        ws = wb["CLUSTERS"]
        last_row = write_section(ws, rows)
        print(f"  written through row {last_row}")

        wb.calculation.fullCalcOnLoad = True
        wb.save(path)
        print(f"  saved {path}")


if __name__ == "__main__":
    main()
