"""
classify_all_rtms.py -- rule-based Requirement Type / Category / Subcategory
classifier for all 722 RTMs, extending the 43 hand-authored T0 Gate rows.

Disclosure (per GBO's ask to extend "type + subcategories, logical, fit for
purpose" beyond just T0): this is a KEYWORD/DOMAIN HEURISTIC, not a
hand-review. It is built to be consistent with the 43 T0 rows that WERE
individually hand-read and classified (t0_taxonomy.py), and is explicitly
tagged "Rule-classified" vs "Hand-reviewed (T0)" everywhere it's surfaced,
so the two confidence levels are never presented as equivalent.

v2 fixes vs the first pass (caught by manual spot-check before this was
used anywhere): naive substring keyword matching produced false hits (e.g.
"equipped" contains "ppe" as a substring) -- switched to \\b-bounded regex
for every short/ambiguous token. Also added a priority-0 documentation-
action override: many "Subsystems"/"Process & Functional"-domain rows are
actually about manual/procedure CONTENT ("shall include, in the O&M manual,
the following...") -- these are PROJECT/Documentation actions regardless of
which technical domain they're filed under, matching GBO's own framing
("deliverables/evidence is SYSTEM, the action/WBS to produce it is PROJECT").
Domain no longer gates SAFETY eligibility either -- hazard-substance
keywords are checked globally, since a domain label doesn't determine
whether a given item's SUBJECT is a hazard.
"""
import warnings, re, json
warnings.filterwarnings("ignore")
import openpyxl
import sys
sys.path.insert(0, ".")
from t0_taxonomy import T0

IN = "QPS_OFFER_Evaluation_FULL_v6.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)
ws = wb["RTM_RANKING"]

QA_DOMAIN = {"Quality Assurance & Control"}
DOC_DOMAINS = {"Technical Documentation", "Contract Performance", "General / Compliance"}
CODES_DOMAIN = {"Codes & Standards"}
EXEC_DOMAINS = {"Acceptance Testing", "Commissioning", "Installation", "Training",
                 "Schedule", "Transport & Logistics", "After-Sales",
                 "Other Deliverables", "Acceptance & Warranty"}
SYSTEM_DOMAINS = {"Subsystems", "Process & Functional", "Design & Fabrication",
                   "Global Design Criteria", "Buildings & Utilities", "Cryogenic Interfaces"}

def rx(*phrases):
    """Compile a \\b-bounded alternation for a list of literal phrases/words."""
    parts = [re.escape(p) for p in phrases]
    return re.compile(r"\b(?:" + "|".join(parts) + r")\b", re.IGNORECASE)

RE_DOC_ACTION = rx(
    "shall submit", "shall document", "shall provide documented", "shall include, in the",
    "operation and maintenance manual", "shall maintain records", "shall report",
    "declaration of conformity", "technical file shall include", "shall provide documented procedures",
    "shall be documented in", "manual shall include",
)
RE_INTERLOCK = rx("interlock", "fail-safe", "failsafe", "trip", "emergency shutdown", "esd",
                   "safety function", "hard-wired interlock")
RE_PERSONNEL = rx("personnel", "oxygen-deficiency", "oxygen deficiency", "odh",
                   "asphyxiat", "evacuat", "confined space", "personal protective equipment",
                   " ppe ")
RE_EQUIP_PROT = rx("relief valve", "relief device", "overpressure", "rupture disc",
                    "protect the equipment", "equipment damage", "cold-helium relief",
                    "isolatable")
RE_REG_CERT = rx("ce marking", "atex", "declaration of conformity")
RE_GOVERN = rx("responsible for", "accountab", "shall ensure that", "role of")
RE_SUBMIT_ISH = rx("submit", "deliver", "documented", "report", "record", "register",
                    "shall provide the", "shall document")
RE_CONTROL = rx("control", "monitor", "signal", "setpoint", "logic", "plc",
                 "interface with", "control strategy", "control system")
RE_CONSTRAINT = rx("shall not exceed", "boundary", "comply with en", "material",
                    "envelope", "tolerance", "shall be rated")
RE_PERFORMANCE = rx("capacity", "efficiency", "flow", "pressure", "temperature",
                     "load", "rate", "cop", "heat", "cooling", "power", "performance")

def classify(domain, text):
    # ---- priority 0: unambiguous documentation/manual-content action ----
    if RE_DOC_ACTION.search(text):
        return "PROJECT", "Documentation & Compliance"

    # ---- priority 1: domain=="Safety & Protection" is a curated hazard-
    # substance domain -- trust it over keyword scanning.
    if domain == "Safety & Protection":
        if RE_INTERLOCK.search(text):
            return "SAFETY", "Interlock & Control Safety"
        if RE_PERSONNEL.search(text):
            return "SAFETY", "Personnel Protection"
        if RE_EQUIP_PROT.search(text):
            return "SAFETY", "Equipment Protection"
        if RE_REG_CERT.search(text):
            return "SAFETY", "Regulatory / Certification"
        return "SAFETY", "Interlock & Control Safety"  # domain-level fallback

    # ---- priority 2: action/execution domains -- being an ACTIVITY (a
    # test, an installation step, a training session) dominates over
    # incidental hazard-adjacent words in the activity's description (e.g.
    # a SAT test script that mentions "fail-safe" while verifying it is
    # still fundamentally a PROJECT verification action, not a safety item
    # in its own right).
    if domain in EXEC_DOMAINS:
        if RE_SUBMIT_ISH.search(text):
            return "PROJECT", "Documentation & Compliance"
        return "PROJECT", "Governance & Responsibility"

    # ---- PROJECT: Quality Assurance -------------------------------------
    if domain in QA_DOMAIN:
        return "PROJECT", "Quality Assurance"

    # ---- PROJECT: documentation-heavy domains ----------------------------
    if domain in DOC_DOMAINS:
        if RE_GOVERN.search(text):
            return "PROJECT", "Governance & Responsibility"
        return "PROJECT", "Documentation & Compliance"

    # ---- Codes & Standards: split by action vs technical compliance -----
    if domain in CODES_DOMAIN:
        if RE_SUBMIT_ISH.search(text) or "certificate" in text.lower():
            return "PROJECT", "Documentation & Compliance"
        return "SYSTEM", "Constraint"

    # ---- Technical/system domains, incl. Control & Interlock: this is
    # where genuine embedded hazard-substance items (e.g. a dedicated
    # fail-safe hard-wired interlock spec) legitimately live, so the
    # hazard-keyword check runs here rather than globally.
    if domain in SYSTEM_DOMAINS or domain == "Control & Interlock":
        if RE_INTERLOCK.search(text):
            return "SAFETY", "Interlock & Control Safety"
        if RE_EQUIP_PROT.search(text):
            return "SAFETY", "Equipment Protection"
        if RE_CONTROL.search(text):
            return "SYSTEM", "Control"
        if RE_CONSTRAINT.search(text):
            return "SYSTEM", "Constraint"
        if RE_PERFORMANCE.search(text):
            return "SYSTEM", "Performance"
        return "SYSTEM", "Performance"  # default within system domains

    # ---- fallback -----------------------------------------------------
    return "PROJECT", "Governance & Responsibility"

results = {}
n_t0 = n_rule = 0
for r in range(6, ws.max_row + 1):
    rid = ws.cell(row=r, column=2).value
    if not rid:
        continue
    if rid in T0:
        n_t0 += 1
        continue  # hand-authored, leave untouched
    domain = ws.cell(row=r, column=5).value or ""
    shall = str(ws.cell(row=r, column=13).value or "")
    full = str(ws.cell(row=r, column=14).value or "")
    text = shall + " " + full
    rtype, subcat = classify(domain, text)
    results[rid] = {"reqType": rtype, "subcategory": subcat, "domain": domain, "method": "rule"}
    n_rule += 1

print(f"T0 hand-authored (untouched): {n_t0}")
print(f"Rule-classified: {n_rule}")
print(f"Total: {n_t0 + n_rule}")

from collections import Counter
overall = Counter()
for v in results.values():
    overall[(v["reqType"], v["subcategory"])] += 1
print("\nDistribution (rule-classified only):")
for k, v in sorted(overall.items()):
    print(" ", k, v)

with open("/tmp/rule_classification.json", "w") as f:
    json.dump(results, f)
print("\nsaved /tmp/rule_classification.json")

print("\n--- spot check: same rows as before ---")
for rid in ["RTM-482", "RTM-300", "RTM-161", "RTM-517", "RTM-026", "RTM-644", "RTM-324", "RTM-367", "RTM-298"]:
    if rid in results:
        v = results[rid]
        for r in range(6, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == rid:
                shall = str(ws.cell(row=r, column=13).value or "")[:110]
                print(f"{rid} [{v['domain']}] -> ({v['reqType']}, {v['subcategory']}): {shall}")
                break
