"""
export_nav_data.py -- exports the JSON data blob the HTML navigator splices
in (rtmRanking, offerRanking, reviewQueue, clusters, domainSummary,
domainClusterBreakdown, taxonomy) directly from the FULL workbook, so the
navigator can never drift from the workbook's own SSOT. Run after any
build_workbook_v*.py that touches RTM_RANKING / OFFER_RANKING /
RTM_REVIEW_QUEUE / CLUSTERS / DOMAIN_SUMMARY / TAXONOMY.
"""
import warnings, json, sys
warnings.filterwarnings("ignore")
import openpyxl

IN = sys.argv[1] if len(sys.argv) > 1 else "QPS_OFFER_Evaluation_FULL_v8.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/tmp/nav_data_v8.json"

wb = openpyxl.load_workbook(IN, data_only=False)

def cellv(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v if v is not None else ""

# ---------------------------------------------------------------- RTM_RANKING
ws = wb["RTM_RANKING"]
rtmRanking = []
for r in range(6, ws.max_row + 1):
    rid = cellv(ws, r, 2)
    if not rid:
        continue
    rtmRanking.append({
        "rank": cellv(ws, r, 1), "id": rid, "gate": cellv(ws, r, 3), "tier": cellv(ws, r, 4),
        "domain": cellv(ws, r, 5), "reqType": cellv(ws, r, 6), "category": cellv(ws, r, 7),
        "subcategory": cellv(ws, r, 8), "cluster": cellv(ws, r, 9), "pdfPage": cellv(ws, r, 10),
        "section": cellv(ws, r, 11), "sectionTitle": cellv(ws, r, 12), "shall": cellv(ws, r, 13),
        # col 14 "Full verbatim requirement" -- the complete shall statement
        # INCLUDING every bulleted/numbered sub-item, straight from the
        # source PDF (col 13 "shall" is a condensed one-liner by contrast).
        # This existed in the workbook already; it was never added to this
        # export, so the Navigator's RTM Lookup silently only ever showed
        # the condensed version. GBO asked for the "full verbatim text /
        # snippet" this round -- this closes that gap using data that was
        # already there, not newly authored.
        "fullVerbatim": cellv(ws, r, 14),
        "weightedS": cellv(ws, r, 22), "btWin": cellv(ws, r, 23), "primaryDim": cellv(ws, r, 25),
        "codeStd": cellv(ws, r, 27), "evidence": cellv(ws, r, 29), "deliverable": cellv(ws, r, 30),
        "phases": cellv(ws, r, 31), "linkedOffer": cellv(ws, r, 33),
    })

# --------------------------------------------------------------- OFFER_RANKING
ws = wb["OFFER_RANKING"]
offerRanking = []
for r in range(6, ws.max_row + 1):
    oid = cellv(ws, r, 2)
    if not oid:
        continue
    offerRanking.append({
        "rank": cellv(ws, r, 1), "id": oid, "title": cellv(ws, r, 3), "gate": cellv(ws, r, 4),
        "tier": cellv(ws, r, 5), "category": cellv(ws, r, 6), "section": cellv(ws, r, 7),
        "weightedS": cellv(ws, r, 15), "btWin": cellv(ws, r, 16), "primaryDim": cellv(ws, r, 18),
        "interpretation": cellv(ws, r, 19), "codeStd": cellv(ws, r, 20),
        "phases": cellv(ws, r, 24),
    })

# ------------------------------------------------------------- RTM_REVIEW_QUEUE
ws = wb["RTM_REVIEW_QUEUE"]
reviewQueue = []
for r in range(6, ws.max_row + 1):
    rid = cellv(ws, r, 2)
    if not rid:
        continue
    reviewQueue.append({
        "rank": cellv(ws, r, 1), "id": rid, "reqType": cellv(ws, r, 3), "category": cellv(ws, r, 4),
        "subcategory": cellv(ws, r, 5), "domain": cellv(ws, r, 6), "cluster": cellv(ws, r, 7),
        "tier": cellv(ws, r, 8), "weightedS": cellv(ws, r, 9), "btWin": cellv(ws, r, 10),
        "section": cellv(ws, r, 11), "shall": cellv(ws, r, 15), "codeStd": cellv(ws, r, 16),
        "deliverable": cellv(ws, r, 17), "evidence": cellv(ws, r, 18), "phases": cellv(ws, r, 19),
        "basis": cellv(ws, r, 20), "disposition": cellv(ws, r, 21),
    })

# ------------------------------------------------------------------- CLUSTERS
ws = wb["CLUSTERS"]
clusters = []
for r in range(6, ws.max_row + 1):
    cid = cellv(ws, r, 1)
    if not cid:
        continue
    clusters.append({"id": cid, "name": cellv(ws, r, 2), "items": cellv(ws, r, 3), "count": cellv(ws, r, 4)})

# -------------------------------------------------------------- DOMAIN_SUMMARY
# NOTE: this sheet has a SECOND table below the per-domain one (a "Cluster
# breakdown" table, stale/direct-crosswalk-only) -- stop at the first blank
# row so that second table's rows don't get read in as if they were domains.
ws = wb["DOMAIN_SUMMARY"]
domainSummary = []
for r in range(6, ws.max_row + 1):
    dom = cellv(ws, r, 1)
    if not dom:
        break
    domainSummary.append({
        "domain": dom, "rtmCount": cellv(ws, r, 2), "gateCount": cellv(ws, r, 3),
        "t1": cellv(ws, r, 4), "t2": cellv(ws, r, 5), "t3": cellv(ws, r, 6),
        "avgS": cellv(ws, r, 7), "maxS": cellv(ws, r, 8), "topRtm": cellv(ws, r, 9),
        "topStatement": cellv(ws, r, 10), "stds": cellv(ws, r, 11), "delivs": cellv(ws, r, 12),
    })

# --------------------------------------------------------- domainClusterBreakdown
# (kept for backward compat with older template code paths; the navigator's
# own buildDomains()/buildDomainsNav() now compute this client-side from
# rtmRanking instead, so it always reflects direct+inferred combined.)
from collections import Counter
cnt = Counter()
CLUSTER_NAMES = {
    "C1": "Performance", "C2": "Process Design", "C3": "Mechanical & Equipment",
    "C4": "Software & Control", "C5": "Infrastructure & Integration",
    "C6": "Reliability & Maintenance", "C7": "Quality, Testing & Risk",
    "C8": "Commercial & Execution",
}
import re
def cluster_key(c):
    if not c:
        return "Not linked"
    m = re.search(r"C[1-8]", c)
    return m.group(0) if m else "Not linked"
for row in rtmRanking:
    k = cluster_key(row.get("cluster"))
    cnt[k] += 1
domainClusterBreakdown = []
for k in ["C1","C2","C3","C4","C5","C6","C7","C8","Not linked"]:
    label = "Not linked" if k == "Not linked" else f"{k} — {CLUSTER_NAMES[k]}"
    domainClusterBreakdown.append({"cluster": label, "count": cnt.get(k, 0)})

# -------------------------------------------------------------------- TAXONOMY
ws = wb["TAXONOMY"]
taxonomy = {
    "requirementTypes": [
        {"type": cellv(ws, r, 1), "definition": cellv(ws, r, 2), "subcategories": cellv(ws, r, 3), "example": cellv(ws, r, 4)}
        for r in (7, 8, 9)
    ],
    "requirementTypeNote": cellv(ws, 11, 1),
    "supergroups": [
        {"name": cellv(ws, r, 1), "phases": cellv(ws, r, 2), "evidence": cellv(ws, r, 3)}
        for r in (15, 16, 17, 18)
    ],
    "supergroupNote": cellv(ws, 20, 1),
    "disposition": [
        {"disposition": cellv(ws, r, 1), "definition": cellv(ws, r, 2)}
        for r in (24, 25, 26, 27, 28, 29)
    ],
    "dispositionNote": cellv(ws, 31, 1),
    # v20 addition: RTM<->OFFER Relation Types table (TAXONOMY rows 82-85,
    # note row 87) -- answers "what does each relation-type badge mean and
    # how/why is it used", grounded in live RTM_CROSSWALK!J counts, colours
    # matching the Navigator's own .linktype CSS classes 1:1.
    "relationTypes": [
        {"type": cellv(ws, r, 1), "count": cellv(ws, r, 2), "share": cellv(ws, r, 3),
         "definition": cellv(ws, r, 4), "usage": cellv(ws, r, 5), "example": cellv(ws, r, 6)}
        for r in (82, 83, 84, 85)
    ] if "TAXONOMY" in wb.sheetnames else [],
    "relationTypesNote": cellv(ws, 87, 1),
}

# ---------------------------------------------------------- COMPLIANCE_LEGEND
# v22 addition: GBO asked that the Taxonomy tab LEAD with the STATUS legend
# and its real colours, not bury it as a text callout after the OFFER
# disposition table. Pulls the actual 8-row STATUS table (Status/Definition/
# Example/Deviation/Next steps/Who resolves) plus each row's real fill hex
# straight from COMPLIANCE_LEGEND!A6:F13, so the Navigator's swatches are
# never invented/approximated -- they are the literal same colours reviewers
# see in EVALUATION_WORKSPACE's STATUS dropdown cells.
statusLegend = []
if "COMPLIANCE_LEGEND" in wb.sheetnames:
    cl = wb["COMPLIANCE_LEGEND"]
    for r in range(6, 14):
        status = cellv(cl, r, 1)
        if not status:
            continue
        fill = cl.cell(row=r, column=1).fill
        hexcolor = None
        try:
            rgb = fill.fgColor.rgb
            if rgb and isinstance(rgb, str) and rgb not in ("00000000",):
                hexcolor = "#" + rgb[-6:]
        except Exception:
            pass
        statusLegend.append({
            "status": status, "definition": cellv(cl, r, 2), "example": cellv(cl, r, 3),
            "deviation": cellv(cl, r, 4), "nextSteps": cellv(cl, r, 5), "whoResolves": cellv(cl, r, 6),
            "color": hexcolor or "#EEEEEE",
        })
# row 17 col D = "EVALUATION_WORKSPACE!Q6:Q55, W6:W55 (live, driven by the
# dropdown) ... " -- where this exact colour system is actually used/entered
# in the workbook, not a generic blurb.
statusLegendNote = cellv(wb["COMPLIANCE_LEGEND"], 17, 4) if "COMPLIANCE_LEGEND" in wb.sheetnames else ""

# ------------------------------------------------------------ DELIVERABLES_DOSSIER
# v20 addition: real contract "Table 2. Applicable Documentation (AD)"
# structure (32 entries, 6 top-level groups + sub-items), cross-referenced
# against every RTM's full text for an AD_## mention. Data rows 6-37, note
# row 39.
deliverablesDossier = []
if "DELIVERABLES_DOSSIER" in wb.sheetnames:
    dd = wb["DELIVERABLES_DOSSIER"]
    for r in range(6, dd.max_row + 1):
        adid = cellv(dd, r, 1)
        if not adid or not str(adid).startswith("AD_"):
            continue
        linked = cellv(dd, r, 6)
        deliverablesDossier.append({
            "id": adid, "name": cellv(dd, r, 2), "reference": cellv(dd, r, 3),
            "group": cellv(dd, r, 4), "linkedCount": cellv(dd, r, 5),
            "linkedRtmIds": [] if linked in ("", "(none found)") else [x.strip() for x in str(linked).split(";")],
            "isTopLevel": "." not in str(adid),
        })
deliverablesDossierNote = cellv(wb["DELIVERABLES_DOSSIER"], dd.max_row, 1) if "DELIVERABLES_DOSSIER" in wb.sheetnames else ""

DATA = {
    "rtmRanking": rtmRanking, "offerRanking": offerRanking, "reviewQueue": reviewQueue,
    "clusters": clusters, "domainSummary": domainSummary,
    "domainClusterBreakdown": domainClusterBreakdown, "taxonomy": taxonomy,
    "deliverablesDossier": deliverablesDossier, "deliverablesDossierNote": deliverablesDossierNote,
    "statusLegend": statusLegend, "statusLegendNote": statusLegendNote,
}
json.dump(DATA, open(OUT, "w", encoding="utf-8"))
print(f"wrote {OUT}: {len(rtmRanking)} rtmRanking, {len(offerRanking)} offerRanking, "
      f"{len(reviewQueue)} reviewQueue, {len(clusters)} clusters, {len(domainSummary)} domainSummary, "
      f"{len(taxonomy.get('relationTypes', []))} relationTypes, {len(deliverablesDossier)} deliverablesDossier, "
      f"{len(statusLegend)} statusLegend")
