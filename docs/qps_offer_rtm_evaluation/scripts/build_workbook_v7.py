"""
build_workbook_v7.py -- Phase 7: extends Requirement Type/Category coverage
from the 43 hand-reviewed T0 rows to all 722 RTMs (rule-based, disclosed),
and assigns an inferred cluster to the 429 RTMs with no direct crosswalk
link (also disclosed, also confidence-tagged). Applied to both RTM_RANKING
and RTM_REVIEW_QUEUE. Adds a methodology note to TAXONOMY.
"""
import warnings, json
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, Alignment

IN = "QPS_OFFER_Evaluation_FULL_v6.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v7.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)

rule_class = json.load(open("/tmp/rule_classification.json"))
inferred_cl = json.load(open("/tmp/inferred_clusters.json"))

RULE_FONT = Font(name="Carlito", italic=True, color="5B7FA6")   # muted blue -- "rule-classified"
RULE_FONT_BOLD = Font(name="Carlito", italic=True, bold=True, color="5B7FA6")
INFER_FONT = Font(name="Carlito", italic=True, color="B7791F")  # amber -- "inferred cluster"
DASH_FONT = Font(name="Carlito", italic=True, color="9A9AA5")
TAX_ALIGN = Alignment(vertical="top", wrap_text=True)

# ============================================================ RTM_RANKING
ws = wb["RTM_RANKING"]
n_rule = n_infer = 0
for r in range(6, ws.max_row + 1):
    rid = ws.cell(row=r, column=2).value
    if not rid:
        continue
    # ---- Requirement Type / Category / Subcategory (cols 6,7,8) --------
    if rid in rule_class:
        v = rule_class[rid]
        c6 = ws.cell(row=r, column=6, value=v["reqType"])
        c6.font = RULE_FONT_BOLD
        c6.alignment = TAX_ALIGN
        c7 = ws.cell(row=r, column=7, value=v["subcategory"])
        c7.font = RULE_FONT
        c7.alignment = TAX_ALIGN
        c8 = ws.cell(row=r, column=8, value="—")
        c8.font = DASH_FONT
        c8.alignment = TAX_ALIGN
        n_rule += 1
    # ---- Cluster (col 9) -- only touch rows still saying "Not linked" --
    cur_cluster = ws.cell(row=r, column=9).value
    if cur_cluster == "Not linked to an OFFER item" and rid in inferred_cl:
        iv = inferred_cl[rid]
        c9 = ws.cell(row=r, column=9,
                      value=f"Inferred: {iv['cluster']} — {iv['clusterName']} ({iv['confidence']} conf.)")
        c9.font = INFER_FONT
        c9.alignment = TAX_ALIGN
        n_infer += 1
print(f"RTM_RANKING: {n_rule} rows rule-classified, {n_infer} rows cluster-inferred")

# ======================================================= RTM_REVIEW_QUEUE
ws2 = wb["RTM_REVIEW_QUEUE"]
n_rule2 = n_infer2 = 0
for r in range(6, ws2.max_row + 1):
    rid = ws2.cell(row=r, column=2).value
    if not rid:
        continue
    if rid in rule_class:
        v = rule_class[rid]
        c3 = ws2.cell(row=r, column=3, value=v["reqType"])
        c3.font = RULE_FONT_BOLD
        c4 = ws2.cell(row=r, column=4, value=v["subcategory"])
        c4.font = RULE_FONT
        c5 = ws2.cell(row=r, column=5, value="—")
        c5.font = DASH_FONT
        n_rule2 += 1
    cur_cluster = ws2.cell(row=r, column=7).value
    if cur_cluster == "Not linked to an OFFER item" and rid in inferred_cl:
        iv = inferred_cl[rid]
        c7 = ws2.cell(row=r, column=7,
                       value=f"Inferred: {iv['cluster']} — {iv['clusterName']} ({iv['confidence']} conf.)")
        c7.font = INFER_FONT
print(f"RTM_REVIEW_QUEUE: {n_rule2} rows rule-classified")

# ================================================================ TAXONOMY
ws3 = wb["TAXONOMY"]
# find the end of the existing content to append a methodology note
last_row = ws3.max_row
r = last_row + 3
ws3.cell(row=r, column=1, value="Classification coverage & confidence").font = Font(name="Carlito", size=13, bold=True, color="17365D")
r += 1
note = (
    "43 T0 Gate RTMs above were individually hand-read and classified (shown in bold black in RTM_RANKING/"
    "RTM_REVIEW_QUEUE). The remaining 679 RTMs' Requirement Type and Category (shown in italic blue) are "
    "RULE-CLASSIFIED -- a keyword/domain heuristic (classify_all_rtms.py), not a hand review. Their Subcategory "
    "is left as \"—\" (not sub-typed this round) rather than guessed. Treat italic-blue rows as a starting point "
    "for review, not a verified classification."
)
ws3.cell(row=r, column=1, value=note).font = Font(name="Carlito", italic=True, size=11, color="444444")
ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
ws3.row_dimensions[r].height = 60
r += 2
note2 = (
    "Cluster: 293 RTMs have a Direct/Supporting/Broad/Contextual crosswalk link to an OFFER item (shown in "
    "plain black, e.g. \"C7 — Quality, Testing & Risk\"). The other 429 had no link at all. 353 of those got an "
    "INFERRED cluster (shown in italic amber, e.g. \"Inferred: C4 — Software & Control (medium conf.)\") from "
    "their Domain's cluster affinity among the 293 known links (plus a keyword tie-break for the ambiguous "
    "Subsystems domain). 76 RTMs (Technical Documentation, Safety & Protection, Cryogenic Interfaces, "
    "Commissioning, Installation, Transport & Logistics) are left \"Not linked to an OFFER item\" because their "
    "whole domain has zero ground-truth links to infer from -- guessing without any grounding data would cross "
    "into fabrication, so those are left honestly unclassified instead."
)
ws3.cell(row=r, column=1, value=note2).font = Font(name="Carlito", italic=True, size=11, color="444444")
ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
ws3.row_dimensions[r].height = 90

wb.save(OUT)
print(f"saved {OUT}")
