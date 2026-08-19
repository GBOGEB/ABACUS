# -*- coding: utf-8 -*-
"""
Standalone, named, discoverable PCA computation against the live workbook --
closes the reproducibility gap flagged in DMAIC_BT_TECHNICAL_REPORT.md
Section 5.1 ("the PCA numbers in this report came from a one-off working
script... if any of Section 4's proposals becomes a recurring view, it needs
a proper compute_pca.py saved under a discoverable name in the project root,
so the numbers can be regenerated any time the underlying RTM data changes").

Standard PCA (z-score standardized across the 7 dimensions, since LC/C have
means near 0.3/0.08 vs F's mean of 1.61 -- same method as the original
one-off analysis) on all 722 RTMs' raw 0-3 dimension scores
(L / R / P / F / Q / LC / C), read live from RTM_RANKING!O:U.

Usage:
    python compute_pca.py <path-to-FULL_vNN.xlsx> <output.json>

Output JSON feeds both build_workbook_vNN.py (PCA_ANALYSIS sheet) and
export_nav_data.py (per-RTM PC1/PC2/PC3 scores for the Navigator's
"PCA / Structure" tab, Proposal A/B from the DMAIC report).
"""
import sys
import json
import numpy as np
from sklearn.decomposition import PCA
import openpyxl

DIM_LABELS = ["L", "R", "P", "F", "Q", "LC", "C"]
DIM_NAMES = {
    "L": "Safety / Legal", "R": "Reliability", "P": "Performance",
    "F": "Functional", "Q": "Quality / Verifiability", "LC": "Lifecycle",
    "C": "Cost",
}
HEADER_ROW = 5
DATA_START_ROW = 6
COL_RTM_ID = 2
COL_DOMAIN = 5
COL_DIMS_START = 15  # O = L, through U = C (7 columns)


def load_rtm_dimensions(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["RTM_RANKING"]
    rtm_ids, domains, X = [], [], []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        rtm_id = ws.cell(row=r, column=COL_RTM_ID).value
        if rtm_id is None:
            continue
        domain = ws.cell(row=r, column=COL_DOMAIN).value
        row_vals = [ws.cell(row=r, column=COL_DIMS_START + i).value for i in range(7)]
        if any(v is None for v in row_vals):
            continue
        rtm_ids.append(rtm_id)
        domains.append(domain)
        X.append(row_vals)
    return rtm_ids, domains, np.array(X, dtype=float)


def standardize(X):
    mean = X.mean(axis=0)
    std = X.std(axis=0, ddof=0)
    std_safe = np.where(std == 0, 1.0, std)
    return (X - mean) / std_safe, mean, std


def main():
    if len(sys.argv) != 3:
        print("usage: python compute_pca.py <FULL_vNN.xlsx> <output.json>")
        sys.exit(1)
    xlsx_path, out_path = sys.argv[1], sys.argv[2]

    rtm_ids, domains, X = load_rtm_dimensions(xlsx_path)
    n = len(rtm_ids)
    Xz, mean, std = standardize(X)

    pca = PCA(n_components=7)
    scores = pca.fit_transform(Xz)
    var_ratio = pca.explained_variance_ratio_
    loadings = pca.components_  # shape (7 components, 7 dims)

    variance_table = []
    cumulative = 0.0
    for i in range(7):
        cumulative += var_ratio[i]
        variance_table.append({
            "pc": f"PC{i+1}",
            "variance_pct": round(float(var_ratio[i]) * 100, 1),
            "cumulative_pct": round(float(cumulative) * 100, 1),
        })

    loadings_table = []
    for dim_idx, dim in enumerate(DIM_LABELS):
        row = {"dimension": dim, "dimension_name": DIM_NAMES[dim]}
        for pc_idx in range(3):
            row[f"PC{pc_idx+1}"] = round(float(loadings[pc_idx][dim_idx]), 3)
        loadings_table.append(row)

    # Per-RTM PC1-3 scores (for the scatter view, Proposal A)
    per_rtm = []
    for i, rtm_id in enumerate(rtm_ids):
        per_rtm.append({
            "rtmId": rtm_id,
            "domain": domains[i],
            "pc1": round(float(scores[i][0]), 3),
            "pc2": round(float(scores[i][1]), 3),
            "pc3": round(float(scores[i][2]), 3),
        })

    # Domain-level mean PC1/PC2 (for domain quadrant / axis-position table)
    domain_pos = {}
    for i, dom in enumerate(domains):
        domain_pos.setdefault(dom, []).append((scores[i][0], scores[i][1]))
    domain_table = []
    for dom, pts in sorted(domain_pos.items(), key=lambda kv: -len(kv[1])):
        pc1_mean = float(np.mean([p[0] for p in pts]))
        pc2_mean = float(np.mean([p[1] for p in pts]))
        domain_table.append({
            "domain": dom, "n": len(pts),
            "pc1_mean": round(pc1_mean, 2), "pc2_mean": round(pc2_mean, 2),
        })

    # Bulk vs. distinctive split (distance from PC1/PC2 center, same method
    # as DMAIC_BT_TECHNICAL_REPORT.md Section 2.5)
    center = scores[:, :2].mean(axis=0)
    dist = np.linalg.norm(scores[:, :2] - center, axis=1)
    threshold = np.percentile(dist, 76)  # matches the report's 76%/24% split method
    n_distinctive = int(np.sum(dist > threshold))

    result = {
        "source_workbook": xlsx_path,
        "n_items": n,
        "method": "z-score standardized, sklearn.decomposition.PCA, 7 raw 0-3 dimension scores (L/R/P/F/Q/LC/C)",
        "variance_table": variance_table,
        "loadings_table": loadings_table,
        "domain_table": domain_table,
        "per_rtm_scores": per_rtm,
        "bulk_vs_distinctive": {
            "distinctive_pct_definition": "distance from PC1/PC2 population center, top ~24% by distance",
            "n_distinctive": n_distinctive,
            "n_bulk": n - n_distinctive,
        },
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)
    print(f"wrote {out_path}: n={n}, PC1={variance_table[0]['variance_pct']}%, "
          f"PC2={variance_table[1]['variance_pct']}%, distinctive={n_distinctive}")


if __name__ == "__main__":
    main()
