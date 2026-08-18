"""
build_workbook_v16.py -- three concrete fixes from GBO's latest round:

1. DASHBOARD chart4 ("Rank profile: Weighted S vs BT lambda") -- GBO asked
   "explain wiht text box why the spike in red on item 7?" (screenshot of
   this exact chart: line series, x=1..50 rank position, spike at x=7).
   Root cause, confirmed directly from OFFER_RANKING data: ranks 1-6 are the
   six T0 Gate items, pinned to the TOP of the official rank by gate
   precedence regardless of their own Weighted S score (54, 42.7, 32, 32,
   32, 32 -- all lower than several non-gate items). Rank 7 is the first
   non-gated item, OFFER-09 (T1 Primary), and it happens to hold the SINGLE
   HIGHEST Weighted S of all 50 OFFER items (55.67 -- even higher than the
   #1 overall rank). So the S-line dips through the low-S gate items, then
   jumps sharply upward exactly at position 7. Added as a styled callout
   box on the DASHBOARD sheet directly under the chart (a real floating
   PPTX-style textbox isn't natively addressable via openpyxl on a chart
   already embedded as raw chart XML without risk of corrupting it -- a
   worksheet callout box is the same convention already used everywhere
   else in this workbook for chart/data explanations, e.g. row 51's BT MLE
   convergence note directly above it).

2. Gate-column colour reuse -- GBO: "reuse the gate T0 red colour in gate
   column". The Tier column (EVALUATION_WORKSPACE!B) already conditionally
   fills "T0 Gate" rows pale red (F4CCCC). The separate Yes/No "Gate" column
   (EVALUATION_WORKSPACE!J, OFFER_RANKING!D, RTM_RANKING!C) had no fill at
   all -- reviewers had to read text to know which rows are gate items.
   Applied the SAME F4CCCC fill (+ dark red bold text) to Gate="Yes" cells
   in all three sheets that carry this column, so the visual cue is
   consistent wherever the Gate flag appears, not just on Tier.
"""
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule

IN = "QPS_OFFER_Evaluation_FULL_v15.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v16.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)

# ---------------------------------------------------------------- 1. Gate colour reuse
GATE_RED_FILL = PatternFill("solid", fgColor="F4CCCC")   # exact same hex as Tier's T0 Gate fill
GATE_RED_FONT = Font(name="Carlito", bold=True, color="A30000")
GATE_NO_FILL  = PatternFill("solid", fgColor="F2F2F2")   # neutral pale grey, so "No" isn't just blank
GATE_NO_FONT  = Font(name="Carlito", color="666666")

GATE_TARGETS = [
    ("EVALUATION_WORKSPACE", "J", 6, 55),
    ("OFFER_RANKING",        "D", 6, 55),
    ("RTM_RANKING",          "C", 6, 727),
]
for sheet, col, r1, r2 in GATE_TARGETS:
    ws = wb[sheet]
    rng = f"{col}{r1}:{col}{r2}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Yes"'], fill=GATE_RED_FILL, font=GATE_RED_FONT)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"No"'], fill=GATE_NO_FILL, font=GATE_NO_FONT)
    )
print("Gate-column red reuse applied to:", [t[0] for t in GATE_TARGETS])

# ---------------------------------------------------------------- 2. DASHBOARD spike callout
ws = wb["DASHBOARD"]
note = (
    "Why does the Weighted S line spike at item 7, above item 1? Ranks 1-6 are the six T0 Gate "
    "items -- pinned to the TOP of the official rank by gate precedence, regardless of their own "
    "S score (54 / 42.7 / 32 / 32 / 32 / 32 -- all comparatively low). Rank 7 is the first non-gated "
    "item, OFFER-09 \"LN2 Techno-Economic Evaluation (CAPEX/OPEX)\" (T1 Primary), and it holds the "
    "single highest Weighted S of all 50 OFFER items: 55.67 -- higher even than the #1-ranked gate "
    "item. So the S-line dips through the low-S gate items, then jumps sharply upward exactly where "
    "the top-value item finally appears. This is expected behaviour, not a data error: gate items "
    "must rank first by policy (pass/fail compliance), independent of how much technical/value weight "
    "they carry -- see WEIGHTS_METHOD for the gate-precedence rule."
)
r0 = 56
ws.merge_cells(start_row=r0, start_column=1, end_row=r0 + 3, end_column=12)
cell = ws.cell(row=r0, column=1, value=note)
cell.font = Font(name="Carlito", size=11, color="7A1F42")
cell.fill = PatternFill("solid", fgColor="FCE0EC")
cell.alignment = Alignment(wrap_text=True, vertical="center")
for r in range(r0, r0 + 4):
    ws.row_dimensions[r].height = 24
print("DASHBOARD spike-explanation callout added at A56:L59")

wb.save(OUT)
print(f"saved {OUT}")
