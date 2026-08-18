"""
build_workbook_v6.py -- Phase 6: RTM taxonomy, cluster grouping, RTM_LOOKUP
search page, OFFER disposition columns, Z_comment formula, and filterable
dashboard, layered onto QPS_OFFER_Evaluation_FULL_v5.xlsx.

Per GBO's requests in this round:
  - Non-generic evidence/deliverable text for T0 Gate RTMs (see t0_taxonomy.py)
  - OFFER Z_comment: auto-drafting formula, not fabricated (no bidder data yet)
  - Domain summary filters + C1-C8 cluster grouping (thematic, not ranking)
  - Requirement Type (SYSTEM/PROJECT/SAFETY) + Category/Subcategory taxonomy
  - RTM_REVIEW_QUEUE reordered into a more usable column sequence
  - New OFFER disposition columns (Deviation/Clarification/NF/Partial/
    Conditional/Unconditional Accept, Comments, cross-link/check)
  - "Insert an RTM number, see all artefacts" lookup page
"""
import warnings
warnings.filterwarnings("ignore")
import json
import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from t0_taxonomy import T0

IN = "QPS_OFFER_Evaluation_FULL_v5.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v6.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)

# ------------------------------------------------------------- style kit ---
TITLE_FILL = PatternFill("solid", fgColor="17365D")
TITLE_FONT = Font(name="Carlito", size=18, bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
SUB_FONT = Font(name="Carlito", size=11, color="203040")
NAV_FILL = PatternFill("solid", fgColor="0F6B78")
NAV_FONT = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Carlito", size=11, color="222222")
BODY_FONT_WRAP = Alignment(wrap_text=True, vertical="top")
BAND_FILL = PatternFill("solid", fgColor="F2F6FA")
THIN = Side(style="thin", color="C9D6E3")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INPUT_FILL = PatternFill("solid", fgColor="FFF6C9")
INPUT_FONT = Font(name="Carlito", size=14, bold=True, color="7A5B00")

NAV_LINKS = [
    ("START_HERE", "Start"), ("DASHBOARD", "Dashboard"), ("RTM_LOOKUP", "RTM Lookup"),
    ("DASHBOARD_2", "Filter view"), ("RTM_REVIEW_QUEUE", "Queue"),
    ("RTM_RANKING", "Ranking"), ("DOMAIN_SUMMARY", "Domains"),
    ("CLUSTERS", "Clusters"), ("TAXONOMY", "Taxonomy"),
]

def new_sheet(name, index, title, subtitle):
    ws = wb.create_sheet(title=name, index=index)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 32
    ws["A2"] = subtitle
    ws["A2"].fill = SUB_FILL
    ws["A2"].font = SUB_FONT
    for i, (sheet, label) in enumerate(NAV_LINKS):
        c = ws.cell(row=3, column=i + 1, value=f'=HYPERLINK("#{sheet}!A1","{label}")')
        c.fill = NAV_FILL
        c.font = NAV_FONT
        c.alignment = Alignment(horizontal="center")
    return ws

def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX

def band_row(ws, row, ncols, start_col=1):
    if row % 2 == 0:
        for c in range(start_col, start_col + ncols):
            ws.cell(row=row, column=c).fill = BAND_FILL

def merge_title(ws, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

print("style kit ready")

# ================================================================= CLUSTERS
CLUSTER_NAMES = {
    "C1": "Performance", "C2": "Process Design", "C3": "Mechanical & Equipment",
    "C4": "Software & Control", "C5": "Infrastructure & Integration",
    "C6": "Reliability & Maintenance", "C7": "Quality, Testing & Risk",
    "C8": "Commercial & Execution",
}
with open("/tmp/offer_cluster.json") as f:
    OFFER_CLUSTER = json.load(f)   # OFFER-ID -> [cluster_id, cluster_name]

cl_idx = wb.sheetnames.index("DOMAIN_SUMMARY")
ws = new_sheet("CLUSTERS", cl_idx, "OFFER item clusters (C1-C8)",
               "Thematic grouping of the 50 OFFER items -- filter aid only, clusters do NOT rank (per the engineering handover). Source: canonical QPS_OFFER_Cluster_v3_6.xlsx, BT_Clusters sheet.")
merge_title(ws, 6)
header_row(ws, 5, ["Cluster ID", "Cluster Name", "OFFER items in cluster", "Count", "", ""])
r = 6
for cid in sorted(CLUSTER_NAMES, key=lambda x: int(x[1:])):
    members = sorted([oid for oid, (c, n) in OFFER_CLUSTER.items() if c == cid],
                      key=lambda o: int(o.split("-")[1]))
    ws.cell(row=r, column=1, value=cid).font = Font(name="Carlito", bold=True)
    ws.cell(row=r, column=2, value=CLUSTER_NAMES[cid]).font = BODY_FONT
    ws.cell(row=r, column=3, value="; ".join(members)).font = BODY_FONT
    ws.cell(row=r, column=3).alignment = BODY_FONT_WRAP
    ws.cell(row=r, column=4, value=len(members)).font = BODY_FONT
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = BOX
    band_row(ws, r, 6)
    r += 1
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 8
print("CLUSTERS sheet built")

# ================================================================ TAXONOMY
tx_idx = wb.sheetnames.index("CLUSTERS") + 1
ws = new_sheet("TAXONOMY", tx_idx, "Requirement taxonomy & definitions",
               "Reference material: Requirement Type, lifecycle supergroups, and the OFFER disposition vocabulary. Read this before using the Type/Category columns on RTM_RANKING and RTM_REVIEW_QUEUE.")
merge_title(ws, 4)

row = 5
ws.cell(row=row, column=1, value="Requirement Type").font = Font(name="Carlito", size=13, bold=True, color="17365D")
row += 1
header_row(ws, row, ["Type", "Definition", "Subcategories", "Example (T0 Gate)"])
row += 1
type_rows = [
    ("SYSTEM", "A technical property or behaviour the delivered QPS itself must exhibit, verified by calculation, test, or design review.",
     "Performance (efficiency/capacity/thermal figures) / Constraint (physical, material, interface, environmental boundary) / Control (control-system behaviour, interlock logic, functional-safety response)",
     "RTM-075 invCOP calculation (Performance)"),
    ("PROJECT", "An action or administrative responsibility the Contractor (the organisation) must carry out -- a WBS task, not a system property.",
     "Documentation & Compliance (submit/produce a specific record) / Governance & Responsibility (who is accountable) / Quality Assurance (QAP-driven process control)",
     "RTM-693 Technical File submission"),
    ("SAFETY", "The requirement's SUBJECT is elimination, reduction, or mitigation of a hazard to personnel or equipment.",
     "Personnel Protection / Equipment Protection / Interlock & Control Safety / Regulatory-Certification",
     "RTM-596 fail-safe hard-wired interlocks"),
]
for t, d, s, e in type_rows:
    ws.cell(row=row, column=1, value=t).font = Font(name="Carlito", bold=True)
    ws.cell(row=row, column=2, value=d).font = BODY_FONT
    ws.cell(row=row, column=2).alignment = BODY_FONT_WRAP
    ws.cell(row=row, column=3, value=s).font = BODY_FONT
    ws.cell(row=row, column=3).alignment = BODY_FONT_WRAP
    ws.cell(row=row, column=4, value=e).font = BODY_FONT
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BOX
    ws.row_dimensions[row].height = 44
    row += 1
row += 1
ws.cell(row=row, column=1, value=(
    "Note: many T0 items are dual-natured (e.g. \"submit a risk analysis\" is a PROJECT action about a SAFETY subject). "
    "SAFETY is used when the hazard itself is the requirement's substance; PROJECT is used when the ask is fundamentally "
    "administrative even if safety-adjacent. This is a judgment call, not a mechanical rule -- flagged here rather than hidden."
)).font = Font(name="Carlito", italic=True, size=10, color="555555")
ws.cell(row=row, column=1).alignment = BODY_FONT_WRAP
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.row_dimensions[row].height = 44
row += 2

ws.cell(row=row, column=1, value="Lifecycle supergroups").font = Font(name="Carlito", size=13, bold=True, color="17365D")
row += 1
header_row(ws, row, ["Supergroup", "Applicable phase(s) it covers", "What evidence looks like at this stage", ""])
row += 1
super_rows = [
    ("Design", "L0 Tender / Offer, L1 Conceptual Design, L2 Detailed Design",
     "A design artefact -- calculation, drawing, specification. Not yet a physical test."),
    ("Build", "L3 Procurement & Manufacturing / FAT",
     "Factory-floor test and inspection records."),
    ("Field", "L4 Installation, L5 Standalone Commissioning",
     "Site records and functional demonstration, ahead of witnessed acceptance."),
    ("Closeout", "L6 Site Acceptance Testing, PAC / Handover",
     "The final witnessed test and handover documentation -- many requirements can only be FORMALLY closed here, even though work toward them started much earlier."),
]
for sg, ph, ev in super_rows:
    ws.cell(row=row, column=1, value=sg).font = Font(name="Carlito", bold=True)
    ws.cell(row=row, column=2, value=ph).font = BODY_FONT
    ws.cell(row=row, column=2).alignment = BODY_FONT_WRAP
    ws.cell(row=row, column=3, value=ev).font = BODY_FONT
    ws.cell(row=row, column=3).alignment = BODY_FONT_WRAP
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = BOX
    ws.row_dimensions[row].height = 40
    row += 1
row += 1
ws.cell(row=row, column=1, value=(
    "A requirement can be partially/gradually fulfilled across supergroups and formally closed at a different point than "
    "where work on it started -- e.g. a calculation closes in Design while a witnessed test on the same RTM can only close "
    "at L6. This workbook does not yet track partial-closure state per phase; that is a possible future enhancement, "
    "flagged here as not built rather than silently assumed."
)).font = Font(name="Carlito", italic=True, size=10, color="555555")
ws.cell(row=row, column=1).alignment = BODY_FONT_WRAP
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.row_dimensions[row].height = 44
row += 2

ws.cell(row=row, column=1, value="OFFER disposition vocabulary").font = Font(name="Carlito", size=13, bold=True, color="17365D")
row += 1
header_row(ws, row, ["Disposition", "Definition", "", ""])
row += 1
disp_rows = [
    ("Deviation", "The OFFER response explicitly departs from the RTM's stated requirement, disclosed by the bidder as a deviation per §4.1 offer rules."),
    ("Clarification", "The OFFER response is ambiguous or incomplete such that a formal RFI/clarification is needed before a disposition can be assigned."),
    ("Not Fulfilled (NF)", "The OFFER response does not address the requirement at all, or contradicts it, with no evidence of compliance."),
    ("Partially Fulfilled", "Some but not all elements of a multi-part requirement are evidenced."),
    ("Conditionally Accepted", "Acceptable subject to a stated follow-up condition (e.g. a document still to be delivered, a test still to be witnessed)."),
    ("Unconditionally Accepted", "Fully evidenced and compliant with no outstanding condition."),
]
for d, defn in disp_rows:
    ws.cell(row=row, column=1, value=d).font = Font(name="Carlito", bold=True)
    ws.cell(row=row, column=2, value=defn).font = BODY_FONT
    ws.cell(row=row, column=2).alignment = BODY_FONT_WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = BOX
    ws.row_dimensions[row].height = 30
    row += 1
row += 1
ws.cell(row=row, column=1, value=(
    "Relationship to STATUS (Compliance_Legend): STATUS is the document-review triage state used per-bidder in "
    "EVALUATION_WORKSPACE (OK / OK-Minor / Minor disc. / Major disc. / NOK / MISS / ERROR / TBD). The disposition "
    "vocabulary above is the RTM-level compliance FINDING recorded once a reviewer has read the OFFER text against a "
    "specific requirement. The two track different things -- workspace-level triage vs. per-requirement compliance "
    "finding -- and are not meant to be forced into a 1:1 mapping."
)).font = Font(name="Carlito", italic=True, size=10, color="555555")
ws.cell(row=row, column=1).alignment = BODY_FONT_WRAP
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.row_dimensions[row].height = 58

ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 44
ws.column_dimensions["D"].width = 20
print("TAXONOMY sheet built")

wb.save("/tmp/wb_stage1.xlsx")
print("stage 1 saved")

# ================================================ RTM_RANKING taxonomy cols
with open("/tmp/rtm_cluster.json") as f:
    RTM_CLUSTER_RAW = json.load(f)  # RTM_ID -> [[cid,cname,ltype,oid], ...]

def resolve_cluster(rid):
    links = RTM_CLUSTER_RAW.get(rid)
    if not links:
        return None, None
    direct = [l for l in links if str(l[2]).startswith("Direct")]
    pick = direct[0] if direct else links[0]
    distinct = sorted(set((l[0], l[1]) for l in links))
    if len(distinct) == 1:
        return distinct[0][0], distinct[0][1]
    return pick[0] + " (+)", pick[1] + " + others"

ws = wb["RTM_RANKING"]
# insert 4 new columns right after 'Domain' (col E=5) so taxonomy sits next
# to the RTM's own thematic grouping -- Requirement Type, Category,
# Subcategory, Cluster
INSERT_AT = 6  # after column E (Domain)
ws.insert_cols(INSERT_AT, 4)
header_row(ws, 5, ["Requirement Type", "Category", "Subcategory", "Cluster"], start_col=INSERT_AT)
# Match the sheet's own pre-existing convention (every other data column is
# vertical="top" + wrap_text=True, since every row is a fixed 50.1pt tall) --
# without this, these 4 new columns default to Excel's bottom alignment and
# float in the lower half of each row while their neighbours sit at the top.
TAX_ALIGN = Alignment(vertical="top", wrap_text=True)
for col, w in zip("FGHI", (16, 20, 24, 26)):
    ws.column_dimensions[col].width = w
n_t0 = n_other = 0
for r in range(6, ws.max_row + 1):
    rid = ws.cell(row=r, column=2).value
    if not rid:
        continue
    cid, cname = resolve_cluster(rid)
    cluster_txt = f"{cid} — {cname}" if cid else "Not linked to an OFFER item"
    cell = ws.cell(row=r, column=INSERT_AT + 3, value=cluster_txt)
    cell.font = BODY_FONT
    cell.alignment = TAX_ALIGN
    if rid in T0:
        rtype, cat, subcat, _, _ = T0[rid]
        c0 = ws.cell(row=r, column=INSERT_AT, value=rtype)
        c0.font = Font(name="Carlito", bold=True)
        c0.alignment = TAX_ALIGN
        c1 = ws.cell(row=r, column=INSERT_AT + 1, value=cat)
        c1.font = BODY_FONT
        c1.alignment = TAX_ALIGN
        c2 = ws.cell(row=r, column=INSERT_AT + 2, value=subcat)
        c2.font = BODY_FONT
        c2.alignment = TAX_ALIGN
        n_t0 += 1
    else:
        c0 = ws.cell(row=r, column=INSERT_AT, value="TBD — not yet classified")
        c0.font = Font(name="Carlito", italic=True, color="9A9AA5")
        c0.alignment = TAX_ALIGN
        n_other += 1
print(f"RTM_RANKING: {n_t0} T0 rows classified, {n_other} left TBD, cluster resolved for all 722")

# rewrite T0 rows' deliverable (col shifted +4, was 26 now 30) and evidence
# (was 25 now 29) with the hand-authored nuanced text
for r in range(6, ws.max_row + 1):
    rid = ws.cell(row=r, column=2).value
    if rid in T0:
        _, _, _, deliv_override, evidence_txt = T0[rid]
        if deliv_override:
            ws.cell(row=r, column=30, value=deliv_override)
        ws.cell(row=r, column=29, value=evidence_txt)
        ws.cell(row=r, column=29).alignment = BODY_FONT_WRAP
print("RTM_RANKING: T0 evidence/deliverable text replaced with hand-authored nuance")

wb.save("/tmp/wb_stage2.xlsx")
print("stage 2 saved")

# ============================================= RTM_REVIEW_QUEUE rebuild ===
# Rebuilt column-by-column into GBO's requested logical sequence: nav/ID
# columns first (Type/Category/Subcategory/Domain/Cluster/Tier/Ranking),
# then verification & validation grouped together (Evidence, Deliverables,
# Codes & Standards sit adjacent instead of scattered), then reviewer-input
# disposition columns at the end.
old = wb["QPS_OFFER_Evaluation_FULL_v5.xlsx" in IN and "RTM_REVIEW_QUEUE" or "RTM_REVIEW_QUEUE"]
old_idx = wb.sheetnames.index("RTM_REVIEW_QUEUE")
old_data = []
for r in range(6, old.max_row + 1):
    rid = old.cell(row=r, column=2).value
    if not rid:
        continue
    old_data.append({
        "rank": old.cell(row=r, column=1).value,
        "rid": rid,
        "tier": old.cell(row=r, column=3).value,
        "domain": old.cell(row=r, column=4).value,
        "pdf_page": old.cell(row=r, column=5).value,
        "section": old.cell(row=r, column=6).value,
        "section_title": old.cell(row=r, column=7).value,
        "primary_dim": old.cell(row=r, column=8).value,
        "weighted_s": old.cell(row=r, column=9).value,
        "bt_win": old.cell(row=r, column=10).value,
        "shall": old.cell(row=r, column=11).value,
        "code_std": old.cell(row=r, column=12).value,
        "deliverable": old.cell(row=r, column=13).value,
        "evidence": old.cell(row=r, column=14).value,
        "phases": old.cell(row=r, column=15).value,
        "basis": old.cell(row=r, column=16).value,
    })
print(f"RTM_REVIEW_QUEUE: {len(old_data)} data rows captured before rebuild")

NEW_HEADERS = [
    "Rank", "RTM ID", "Requirement Type", "Category", "Subcategory", "Domain",
    "Cluster", "Tier", "Weighted S", "BT win %", "Section", "Section title",
    "PDF page", "Primary dimension", "Shall statement",
    "Actual explicit code / standard", "Explicit deliverable / proof",
    "Evidence / measurability (V&V)", "Applicable phase(s)", "Basis / review",
    "Disposition", "Comments", "Cross-link / check to perform",
]
N = len(NEW_HEADERS)

wb.remove(old)
ws = wb.create_sheet(title="RTM_REVIEW_QUEUE", index=old_idx)
ws.sheet_view.showGridLines = False
ws["A1"] = "RTM review queue — priority subset"
ws["A1"].fill = TITLE_FILL
ws["A1"].font = TITLE_FONT
ws.row_dimensions[1].height = 32
ws["A2"] = ("Standalone queue: all T0/T1 RTMs plus deliverable-heavy requirements. "
            "Columns run ID/navigation -> tier & ranking -> verification, deliverables & "
            "codes/standards (grouped) -> disposition (reviewer input, right-hand side).")
ws["A2"].fill = SUB_FILL
ws["A2"].font = SUB_FONT
for i, (sheet, label) in enumerate(NAV_LINKS):
    c = ws.cell(row=3, column=i + 1, value=f'=HYPERLINK("#{sheet}!A1","{label}")')
    c.fill = NAV_FILL
    c.font = NAV_FONT
    c.alignment = Alignment(horizontal="center")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
header_row(ws, 5, NEW_HEADERS)

DISP_VALUES = ["TBD", "Deviation", "Clarification", "Not Fulfilled (NF)",
               "Partially Fulfilled", "Conditionally Accepted", "Unconditionally Accepted"]
dv = DataValidation(type="list", formula1='"' + ",".join(DISP_VALUES) + '"', allow_blank=True)
ws.add_data_validation(dv)

# All the long-text columns wrap; everything else is short and stays single-line.
# Every cell (wrapped or not) gets vertical="top" -- otherwise, on a row that's
# been made tall by one wrapped long-text column, the short unwrapped columns
# (Rank, RTM ID, Tier...) default to Excel's bottom vertical alignment and end
# up floating in empty space at the bottom of the row. Row height is computed
# per-row from the widest wrapped column's actual text length so nothing is
# visually clipped.
WRAP_COLS = {15, 16, 17, 18, 19, 20, 22, 23}
COL_CHAR_WIDTH = {15: 44, 16: 26, 17: 34, 18: 40, 19: 30, 20: 22, 22: 30, 23: 30}

r = 6
n_t0 = 0
for d in old_data:
    rid = d["rid"]
    cid, cname = resolve_cluster(rid)
    cluster_txt = f"{cid} — {cname}" if cid else "Not linked to an OFFER item"
    if rid in T0:
        rtype, cat, subcat, deliv_override, evidence_txt = T0[rid]
        deliverable = deliv_override or d["deliverable"]
        evidence = evidence_txt
        n_t0 += 1
    else:
        rtype, cat, subcat = "TBD — not yet classified", "", ""
        deliverable = d["deliverable"]
        evidence = d["evidence"]
    row_vals = [
        d["rank"], rid, rtype, cat, subcat, d["domain"], cluster_txt, d["tier"],
        d["weighted_s"], d["bt_win"], d["section"], d["section_title"], d["pdf_page"],
        d["primary_dim"], d["shall"], d["code_std"], deliverable, evidence,
        d["phases"], d["basis"], "TBD", None, None,
    ]
    max_lines = 1
    for c, v in enumerate(row_vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Carlito", bold=True) if c == 3 and rid in T0 else BODY_FONT
        if c in WRAP_COLS:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if v:
                import math
                lines = math.ceil(len(str(v)) / COL_CHAR_WIDTH[c])
                max_lines = max(max_lines, lines)
        else:
            cell.alignment = Alignment(vertical="top")
        cell.border = BOX
    ws.row_dimensions[r].height = max(20, max_lines * 14 + 8)
    dv.add(ws.cell(row=r, column=21))
    band_row(ws, r, N)
    r += 1
print(f"RTM_REVIEW_QUEUE rebuilt: {len(old_data)} rows, {n_t0} with hand-authored T0 content, "
      f"{N} columns in the new sequence")

widths = {"A": 6, "B": 10, "C": 14, "D": 20, "E": 22, "F": 20, "G": 22, "H": 12,
          "I": 10, "J": 9, "K": 9, "L": 22, "M": 9, "N": 14, "O": 44, "P": 26,
          "Q": 34, "R": 40, "S": 30, "T": 22, "U": 20, "V": 30, "W": 30}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "C6"

wb.save("/tmp/wb_stage3.xlsx")
print("stage 3 saved")

# ======================================== EVALUATION_WORKSPACE: Z_comment ==
ws = wb["EVALUATION_WORKSPACE"]
ws.insert_cols(29, 1)  # before old col29 "Board / Negotiation Topic"
ws.cell(row=5, column=29, value="Z_comment (auto A vs B synthesis)")
ws.cell(row=5, column=29).fill = HDR_FILL
ws.cell(row=5, column=29).font = HDR_FONT
ws.cell(row=5, column=29).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(row=4, column=29, value="Formula-driven -- fills itself in once A_STATUS/B_STATUS are set; nothing fabricated while both are TBD.").font = Font(name="Carlito", italic=True, size=9, color="7A5B00")
n_z = 0
for r in range(6, ws.max_row + 1):
    if ws.cell(row=r, column=1).value is None:
        continue
    f = (
        f'=IF(AND(Q{r}="TBD",W{r}="TBD"),"Not yet evaluated — awaiting bidder review.",'
        f'IF(Q{r}="TBD",CONCATENATE("Bidder B: ",W{r}," — ",IF(AA{r}="","(no comment)",AA{r}),'
        f'" | Bidder A not yet evaluated."),'
        f'IF(W{r}="TBD",CONCATENATE("Bidder A: ",Q{r}," — ",IF(U{r}="","(no comment)",U{r}),'
        f'" | Bidder B not yet evaluated."),'
        f'CONCATENATE("Bidder A: ",Q{r}," — ",IF(U{r}="","(no comment)",U{r}),'
        f'" || Bidder B: ",W{r}," — ",IF(AA{r}="","(no comment)",AA{r}),'
        f'" | ",IF(Q{r}=W{r},"Comparable disposition.","Diverging disposition — flag for negotiation.")))))'
    )
    c = ws.cell(row=r, column=29, value=f)
    c.font = BODY_FONT
    c.alignment = BODY_FONT_WRAP
    n_z += 1
ws.column_dimensions[get_column_letter(29)].width = 42
print(f"EVALUATION_WORKSPACE: Z_comment formula added to {n_z} rows")

wb.save("/tmp/wb_stage4.xlsx")
print("stage 4 saved")

# ============================================ DOMAIN_SUMMARY: filter + cluster
ws = wb["DOMAIN_SUMMARY"]
ws.auto_filter.ref = "A5:L27"
print("DOMAIN_SUMMARY: AutoFilter enabled on A5:L27")

# cluster breakdown -- separate thematic axis, explicitly NOT a ranking
start = 29
ws.cell(row=start, column=1, value="Cluster breakdown (C1-C8) — thematic, not ranking").font = Font(name="Carlito", size=13, bold=True, color="17365D")
start += 1
ws.cell(row=start, column=1, value=(
    "Same 722 RTMs, regrouped by which OFFER cluster their linked OFFER item(s) belong to "
    "(via RTM_CROSSWALK). RTMs with no crosswalk link show as 'Not linked'. This is a filter aid, "
    "same as the CLUSTERS sheet -- it does not change any RTM's rank or tier."
)).font = Font(name="Carlito", italic=True, size=10, color="555555")
ws.cell(row=start, column=1).alignment = BODY_FONT_WRAP
ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=6)
start += 2
header_row(ws, start, ["Cluster", "RTM count"], )
start += 1
cluster_counts = {cid: 0 for cid in CLUSTER_NAMES}
cluster_counts["Not linked"] = 0
for rid in [wb["RTM_RANKING"].cell(row=r, column=2).value for r in range(6, wb["RTM_RANKING"].max_row + 1)]:
    if not rid:
        continue
    cid, _ = resolve_cluster(rid)
    key = cid.split(" ")[0] if cid else "Not linked"
    key = key.replace("(+)", "").strip()
    cluster_counts[key] = cluster_counts.get(key, 0) + 1
for cid in list(CLUSTER_NAMES) + ["Not linked"]:
    label = f"{cid} — {CLUSTER_NAMES[cid]}" if cid in CLUSTER_NAMES else cid
    ws.cell(row=start, column=1, value=label).font = BODY_FONT
    ws.cell(row=start, column=2, value=cluster_counts[cid]).font = BODY_FONT
    for c in (1, 2):
        ws.cell(row=start, column=c).border = BOX
    band_row(ws, start, 2)
    start += 1
print(f"DOMAIN_SUMMARY: cluster breakdown added ({sum(cluster_counts.values())} RTMs accounted for)")

wb.save("/tmp/wb_stage5.xlsx")
print("stage 5 saved")

# ================================================================ DASHBOARD_2
# FILTER() was tested against this sandbox's LibreOffice 24.2 headless
# recalc and returned #VALUE! -- not safe to ship (breaks the zero-formula-
# error QA bar used throughout this project). Built instead on COUNTIFS/
# AVERAGEIFS with wildcard "wrapper" cells for the "(ALL)" state -- classic
# formulas, proven safe, and give the same "pick a Domain or Cluster and see
# live counts" interaction GBO asked for.
DOMAINS = ['Subsystems', 'Control & Interlock', 'Acceptance Testing', 'Buildings & Utilities',
           'Process & Functional', 'Quality Assurance & Control', 'Design & Fabrication',
           'Technical Documentation', 'Global Design Criteria', 'Contract Performance',
           'Safety & Protection', 'Codes & Standards', 'Acceptance & Warranty',
           'Cryogenic Interfaces', 'Training', 'Schedule', 'Other Deliverables',
           'General / Compliance', 'Commissioning', 'Installation', 'After-Sales',
           'Transport & Logistics']
d2_idx = wb.sheetnames.index("DASHBOARD") + 1
ws = new_sheet("DASHBOARD_2", d2_idx, "Filter view — by Domain or Cluster",
               "Pick a Domain and/or a Cluster below; the counts update live. This is a filter aid over the same 722 RTMs in RTM_RANKING -- for row-level detail, use RTM_RANKING's own AutoFilter after picking your slice here, or open RTM_LOOKUP for one specific RTM.")
merge_title(ws, 6)

ws.cell(row=5, column=1, value="Domain").font = Font(name="Carlito", bold=True)
ws.cell(row=5, column=2, value="(ALL)").fill = INPUT_FILL
ws.cell(row=5, column=2).font = INPUT_FONT
dv_dom = DataValidation(type="list", formula1='"(ALL),' + ",".join(DOMAINS) + '"', allow_blank=False)
ws.add_data_validation(dv_dom)
dv_dom.add(ws.cell(row=5, column=2))

ws.cell(row=6, column=1, value="Cluster").font = Font(name="Carlito", bold=True)
ws.cell(row=6, column=2, value="(ALL)").fill = INPUT_FILL
ws.cell(row=6, column=2).font = INPUT_FONT
dv_cl = DataValidation(type="list", formula1='"(ALL),C1,C2,C3,C4,C5,C6,C7,C8"', allow_blank=False)
ws.add_data_validation(dv_cl)
dv_cl.add(ws.cell(row=6, column=2))

# hidden wildcard-wrapper helper cells
ws.cell(row=5, column=4, value='=IF(B5="(ALL)","*",B5)')
ws.cell(row=6, column=4, value='=IF(B6="(ALL)","*","*"&B6&" —*")')
ws.cell(row=5, column=3, value="Domain criteria (helper):").font = Font(italic=True, size=9, color="9A9AA5")
ws.cell(row=6, column=3, value="Cluster criteria (helper):").font = Font(italic=True, size=9, color="9A9AA5")

RR = "RTM_RANKING"
DOM_RNG = f"{RR}!$E$6:$E$728"
CLU_RNG = f"{RR}!$I$6:$I$728"
TIER_RNG = f"{RR}!$D$6:$D$728"
WS_RNG = f"{RR}!$V$6:$V$728"   # Weighted S is col 22 = V

def tile(row, label, formula):
    ws.cell(row=row, column=1, value=label).font = Font(name="Carlito", bold=True)
    c = ws.cell(row=row, column=2, value=formula)
    c.font = Font(name="Carlito", size=14, bold=True, color="17365D")
    for cc in range(1, 3):
        ws.cell(row=row, column=cc).border = BOX

row = 8
ws.cell(row=row, column=1, value="Live counts for this slice").font = Font(name="Carlito", size=13, bold=True, color="17365D")
row += 1
tile(row, "Matching RTMs", f'=COUNTIFS({DOM_RNG},$D$5,{CLU_RNG},$D$6)')
row += 1
tile(row, "  of which T0 Gate", f'=COUNTIFS({DOM_RNG},$D$5,{CLU_RNG},$D$6,{TIER_RNG},"T0 Gate")')
row += 1
tile(row, "  of which T1 Primary", f'=COUNTIFS({DOM_RNG},$D$5,{CLU_RNG},$D$6,{TIER_RNG},"T1 Primary")')
row += 1
tile(row, "  of which T2 Secondary", f'=COUNTIFS({DOM_RNG},$D$5,{CLU_RNG},$D$6,{TIER_RNG},"T2 Secondary")')
row += 1
tile(row, "  of which T3 Contextual", f'=COUNTIFS({DOM_RNG},$D$5,{CLU_RNG},$D$6,{TIER_RNG},"T3 Contextual")')
row += 1
tile(row, "Average Weighted S", f'=IFERROR(ROUND(AVERAGEIFS({WS_RNG},{DOM_RNG},$D$5,{CLU_RNG},$D$6),2),"n/a")')
row += 2
ws.cell(row=row, column=1, value=(
    "Row-level detail: switch to RTM_RANKING and use its AutoFilter (Domain / Cluster / Tier / Requirement Type "
    "columns) with the same picks you made above -- this page shows the count, RTM_RANKING's filter shows the rows."
)).font = Font(name="Carlito", italic=True, size=10, color="555555")
ws.cell(row=row, column=1).alignment = BODY_FONT_WRAP
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 16
ws.column_dimensions["C"].hidden = True
ws.column_dimensions["D"].hidden = True
print("DASHBOARD_2 built (COUNTIFS/AVERAGEIFS-driven, FILTER() avoided)")

# AutoFilter on RTM_RANKING and the rebuilt RTM_REVIEW_QUEUE for row-level use
wb["RTM_RANKING"].auto_filter.ref = f"A5:{get_column_letter(wb['RTM_RANKING'].max_column)}{wb['RTM_RANKING'].max_row}"
wb["RTM_REVIEW_QUEUE"].auto_filter.ref = f"A5:{get_column_letter(wb['RTM_REVIEW_QUEUE'].max_column)}{wb['RTM_REVIEW_QUEUE'].max_row}"
print("AutoFilter enabled on RTM_RANKING and RTM_REVIEW_QUEUE")

wb.save("/tmp/wb_stage6.xlsx")
print("stage 6 saved")

# ===================================================== RTM_RANKING: crosswalk
# TEXTJOIN(...,IF(array),...) and FILTER() both returned #VALUE! under this
# sandbox's LibreOffice headless recalc (tested standalone before committing
# to either) -- so "linked OFFER items" is precomputed here in Python from
# RTM_CROSSWALK's own (already-static) link table and written as a plain
# value, then read by RTM_LOOKUP via a normal INDEX/MATCH -- the same safe
# pattern already used throughout this workbook (START_HERE, EVALUATION_
# WORKSPACE's BT-score lookups, etc.), not a new formula class.
cw = wb["RTM_CROSSWALK"]
rtm_links = {}
for r in range(5, cw.max_row + 1):
    oid = cw.cell(row=r, column=1).value
    rid = cw.cell(row=r, column=5).value
    ltype = cw.cell(row=r, column=10).value
    if oid and rid:
        short = "Direct" if str(ltype).startswith("Direct") else (
            "Supporting" if str(ltype).startswith("Supporting") else str(ltype))
        rtm_links.setdefault(rid, []).append(f"{oid} ({short})")

ws = wb["RTM_RANKING"]
ws.cell(row=5, column=33, value="Linked OFFER items (crosswalk)")
ws.cell(row=5, column=33).fill = HDR_FILL
ws.cell(row=5, column=33).font = HDR_FONT
ws.cell(row=5, column=33).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
n_linked = 0
for r in range(6, ws.max_row + 1):
    rid = ws.cell(row=r, column=2).value
    if not rid:
        continue
    links = rtm_links.get(rid)
    txt = "; ".join(links) if links else "Not linked to an OFFER item"
    ws.cell(row=r, column=33, value=txt).font = BODY_FONT
    ws.cell(row=r, column=33).alignment = BODY_FONT_WRAP
    if links:
        n_linked += 1
ws.column_dimensions[get_column_letter(33)].width = 40
print(f"RTM_RANKING: Linked OFFER items precomputed for {n_linked}/722 RTMs")

wb.save("/tmp/wb_stage7.xlsx")
print("stage 7 saved")

# ================================================================ RTM_LOOKUP
# The "insert an RTM number, see all artefacts" page GBO called out as
# a great addition. Plain INDEX/MATCH against RTM_RANKING (all 722 RTMs,
# not just the 289-row priority queue) -- the same safe formula pattern
# already used throughout the workbook.
rl_idx = wb.sheetnames.index("RTM_LOOKUP") if "RTM_LOOKUP" in wb.sheetnames else wb.sheetnames.index("DASHBOARD_2") + 1
ws = new_sheet("RTM_LOOKUP", rl_idx, "RTM Lookup — enter any RTM ID",
               "Type an RTM ID (e.g. RTM-693) in the yellow box. Every field for that requirement pulls in below, across all 722 canonical RTMs -- not just the 289 in the priority queue.")
merge_title(ws, 4)

ws.cell(row=5, column=1, value="RTM ID:").font = Font(name="Carlito", size=13, bold=True)
inp = ws.cell(row=5, column=2, value="RTM-693")
inp.fill = INPUT_FILL
inp.font = INPUT_FONT
ws.cell(row=5, column=3, value=(
    "Format: RTM-### (3 digits, e.g. RTM-001, RTM-693). If the box below shows #N/A, check the ID exists on RTM_RANKING."
)).font = Font(italic=True, size=9, color="7A5B00")
ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=6)

RR = "RTM_RANKING"
MATCH = f'MATCH($B$5,{RR}!$B$6:$B$728,0)'
FIELDS = [
    ("Rank (of 722)", 1), ("Gate (T0?)", 3), ("Tier", 4), ("Domain", 5),
    ("Requirement Type", 6), ("Category", 7), ("Subcategory", 8), ("Cluster", 9),
    ("PDF page", 10), ("Section", 11), ("Section title", 12),
    ("Shall statement", 13), ("Full verbatim requirement", 14),
    ("Weighted S", 22), ("BT Win %", 23), ("Primary dimension", 25),
    ("Actual explicit code / standard", 27), ("Evidence / measurability", 29),
    ("Explicit deliverable / proof", 30), ("Applicable phase(s)", 31),
    ("Evidence basis / review", 32), ("Linked OFFER items (crosswalk)", 33),
]
row = 7
header_row(ws, row, ["Field", "Value", "", "", "", ""])
row += 1
for label, col in FIELDS:
    ws.cell(row=row, column=1, value=label).font = Font(name="Carlito", bold=True)
    col_letter = get_column_letter(col)
    formula = f'=IFERROR(INDEX({RR}!${col_letter}$6:${col_letter}$728,{MATCH}),"— RTM ID not found —")'
    c = ws.cell(row=row, column=2, value=formula)
    c.font = BODY_FONT
    c.alignment = BODY_FONT_WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    for cc in (1, 2):
        ws.cell(row=row, column=cc).border = BOX
    band_row(ws, row, 6)
    if label in ("Shall statement", "Full verbatim requirement", "Evidence / measurability"):
        ws.row_dimensions[row].height = 60
    row += 1

row += 1
ws.cell(row=row, column=1, value="Reviewer disposition for this RTM (from RTM_REVIEW_QUEUE, if it's in the priority subset)").font = Font(name="Carlito", size=12, bold=True, color="17365D")
row += 1
RQ = "RTM_REVIEW_QUEUE"
MATCH_Q = f'MATCH($B$5,{RQ}!$B$6:$B$294,0)'
DQ_FIELDS = [("Disposition", 21), ("Comments", 22), ("Cross-link / check to perform", 23)]
for label, col in DQ_FIELDS:
    ws.cell(row=row, column=1, value=label).font = Font(name="Carlito", bold=True)
    col_letter = get_column_letter(col)
    formula = f'=IFERROR(INDEX({RQ}!${col_letter}$6:${col_letter}$294,{MATCH_Q}),"— not in the priority queue —")'
    c = ws.cell(row=row, column=2, value=formula)
    c.font = BODY_FONT
    c.alignment = BODY_FONT_WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    for cc in (1, 2):
        ws.cell(row=row, column=cc).border = BOX
    band_row(ws, row, 6)
    row += 1

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 22
for col in ("C", "D", "E", "F"):
    ws.column_dimensions[col].width = 16
print("RTM_LOOKUP sheet built")

wb.save("/tmp/wb_stage8.xlsx")
print("stage 8 saved")

# =================================================== START_HERE / NAV wiring
sh = wb["START_HERE"]
new_links = [("RTM_LOOKUP", "RTM Lookup"), ("DASHBOARD_2", "Filter view"),
             ("CLUSTERS", "Clusters"), ("TAXONOMY", "Taxonomy")]
start_col = 11
for i, (sheet, label) in enumerate(new_links):
    c = sh.cell(row=3, column=start_col + i, value=f'=HYPERLINK("#{sheet}!A1","{label}")')
    c.font = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="0F6B78")
    c.alignment = Alignment(horizontal="center")

step_row = 15  # after the existing 9-step walkthrough (steps 1-9, rows 5-14)
steps = [
    (10, "RTM_LOOKUP", "Look up any single RTM by ID and see every field for it in one place, incl. its OFFER cross-links and (if in the priority queue) its disposition.", "Fastest way to answer 'what does RTM-### actually require and what counts as evidence?' during a review meeting."),
    (11, "DASHBOARD_2", "Pick a Domain and/or Cluster and see live counts (total, by tier, average Weighted S) for that slice.", "Use before assigning review workload -- see which slice is biggest before diving into row-level detail."),
    (12, "CLUSTERS", "Reference: which of the 50 OFFER items sit in each of the 8 thematic clusters (C1-C8). Filter aid only -- not a ranking.", "Cross-check with DASHBOARD_2's Cluster picker."),
    (13, "TAXONOMY", "Reference: Requirement Type (SYSTEM/PROJECT/SAFETY) definitions, lifecycle supergroups, and the OFFER disposition vocabulary.", "Read once before using the Type/Category columns or the Disposition column."),
]
for step_num, sheet, purpose, action in steps:
    r = step_row + (step_num - 10)
    sh.cell(row=r, column=1, value=str(step_num)).font = NAV_FONT if False else Font(name="Carlito")
    sh.cell(row=r, column=2, value=f'=HYPERLINK("#{sheet}!A1","{sheet}")').font = Font(name="Carlito")
    sh.cell(row=r, column=3, value=purpose).font = Font(name="Carlito")
    sh.cell(row=r, column=3).alignment = BODY_FONT_WRAP
    sh.cell(row=r, column=4, value=action).font = Font(name="Carlito")
    sh.cell(row=r, column=4).alignment = BODY_FONT_WRAP
print("START_HERE: nav bar + step list extended with the 4 new sheets")

nm = wb["NAVIGATION_MAP"]
nm_start = nm.max_row + 1
nm_rows = [
    ("RTM_LOOKUP", "Single-RTM detail view across all 722 canonical requirements", "Look up one RTM during review or negotiation", "All reviewers"),
    ("DASHBOARD_2", "Live filtered counts by Domain / Cluster", "Scope review workload before diving into rows", "Lead evaluator"),
    ("CLUSTERS", "C1-C8 thematic grouping of the 50 OFFER items (filter aid, not a ranking)", "Cross-reference with Domain/Cluster filters", "All reviewers"),
    ("TAXONOMY", "Requirement Type, lifecycle supergroup, and disposition-vocabulary definitions", "Read before using Type/Category/Disposition columns", "All reviewers"),
]
for offset, (sheet, purpose, action, owner) in enumerate(nm_rows):
    r = nm_start + offset
    nm.cell(row=r, column=1, value=f'=HYPERLINK("#{sheet}!A1","{sheet}")').font = Font(name="Carlito")
    nm.cell(row=r, column=2, value=purpose).font = Font(name="Carlito")
    nm.cell(row=r, column=3, value=action).font = Font(name="Carlito")
    nm.cell(row=r, column=4, value=owner).font = Font(name="Carlito")
print(f"NAVIGATION_MAP extended with {len(nm_rows)} rows")

# tab colours for the new sheets, matching the v5 palette convention
NEW_TAB_COLOR = {
    "CLUSTERS": "1FA7A0", "TAXONOMY": "1FA7A0",
    "RTM_LOOKUP": "562873", "DASHBOARD_2": "562873",
}
for name, hexcolor in NEW_TAB_COLOR.items():
    wb[name].sheet_properties.tabColor = Color(rgb="FF" + hexcolor)

wb.save(OUT)
print(f"FINAL saved: {OUT} -- {len(wb.sheetnames)} sheets")
print(wb.sheetnames)
