"""
build_workbook_slim_v6.py -- produces the v6 reviewer-shareable "Lite"
workbook from QPS_OFFER_Evaluation_FULL_v13.xlsx.

KEEP grows from v5's 11 sheets to 17: this round added four sheets that are
themselves reviewer-facing tools (RTM_LOOKUP, DASHBOARD_2, CLUSTERS,
TAXONOMY) rather than internal artefacts, plus RTM_REVIEW_QUEUE (now carries
the live Disposition/Comments/Cross-link tracking columns -- genuinely
reviewer input, not just an internal audit extract) and DOMAIN_SUMMARY (now
carries the AutoFilter + cluster breakdown GBO asked for) are promoted from
"drop" to "keep" for the same reason.

Still dropped (unchanged reasoning from v5): STANDARDS, DELIVERABLES,
AUDIT_NOTES, DMAIC_AUDIT, EVALUATION_INPUT, NAVIGATION_MAP, CODING_HANDOVER,
WEIGHTS_METHOD -- internal/audit/method-reproduction artefacts per
NAVIGATION_MAP's own role column.

Same dead-hyperlink neutralisation and formula-dependency discipline as v5's
slim-down script.
"""
import warnings
warnings.filterwarnings("ignore")
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

FULL = "QPS_OFFER_Evaluation_FULL_v13.xlsx"
OUT = "QPS_OFFER_Evaluation_LITE_v13.xlsx"

KEEP = [
    "START_HERE", "DASHBOARD", "DASHBOARD_2", "RTM_LOOKUP", "COMPLIANCE_LEGEND",
    "EVALUATION_WORKSPACE", "NEGOTIATION_AGENDA", "OFFER_RANKING", "RTM_CROSSWALK",
    "RTM_RANKING", "CLUSTERS", "TAXONOMY", "DOMAIN_SUMMARY", "RTM_REVIEW_QUEUE",
    "QUALITY_CHECKS", "LISTS", "OFFER_CANONICAL",
]
HYPERLINK_RE = re.compile(r'=HYPERLINK\("#([A-Za-z0-9_]+)!', re.IGNORECASE)
DEAD_FILL = PatternFill("solid", fgColor="E8E8EC")
DEAD_FONT = Font(name="Carlito", size=11, italic=True, color="9A9AA5")

wb = openpyxl.load_workbook(FULL, data_only=False)

# ---- pre-flight: full cross-sheet formula-dependency scan before dropping -
CELL_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!\$?[A-Z]{1,3}\$?\d+")
dropped_candidates = [n for n in wb.sheetnames if n not in KEEP]
live_deps = set()
for name in KEEP:
    ws = wb[name]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=") and "HYPERLINK" not in cell.value:
                for m in CELL_REF_RE.finditer(cell.value):
                    ref = m.group(1) or m.group(2)
                    if ref in dropped_candidates:
                        live_deps.add((name, ref, cell.coordinate))
if live_deps:
    print("WARNING -- live formula dependencies on sheets about to be dropped:")
    for d in live_deps:
        print("  ", d)
else:
    print("pre-flight OK: no kept-sheet formula depends on a dropped sheet")

dropped = [n for n in wb.sheetnames if n not in KEEP]
for name in dropped:
    del wb[name]
print(f"dropped {len(dropped)} sheets: {dropped}")
print(f"kept {len(wb.sheetnames)} sheets: {wb.sheetnames}")

wb["LISTS"].sheet_state = "hidden"

sh = wb["START_HERE"]
if sh["A1"].value and "DMAIC-ready" in str(sh["A1"].value):
    sh["A1"] = str(sh["A1"].value).replace(
        "QPS OFFER evaluation workbook — v5 DMAIC-ready",
        "QPS OFFER evaluation workbook — v6 LITE (reviewer edition)",
    )
sh["A2"] = (
    "Reviewer-shareable subset of the full evaluation workbook -- ranking, crosswalk, "
    "workspace, quality, lookup, filter-dashboard, cluster and taxonomy tabs. Internal "
    "audit/handover/method-control tabs live in the full v6 workbook only."
)

n_neutralised = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue
            m = HYPERLINK_RE.match(cell.value)
            if not m:
                continue
            target_sheet = m.group(1)
            if target_sheet in wb.sheetnames:
                continue
            label_match = re.search(r',"([^"]*)"\)$', cell.value)
            label = label_match.group(1) if label_match else target_sheet
            cell.value = label
            cell.font = DEAD_FONT
            cell.fill = DEAD_FILL
            n_neutralised += 1
print(f"neutralised {n_neutralised} dead cross-sheet hyperlinks (dropped-tab references)")

for row_num in (7, 8):
    for col in ("A", "C", "D"):
        cell = sh[f"{col}{row_num}"]
        if cell.value:
            cell.font = DEAD_FONT

wb.save(OUT)
print(f"saved {OUT}")
