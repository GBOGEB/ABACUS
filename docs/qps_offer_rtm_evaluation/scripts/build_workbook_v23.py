"""
build_workbook_v23.py -- adds AD_07 and AD_08 to DELIVERABLES_DOSSIER.

GBO granted this session access to Master_Input (his shared working folder)
and pointed at the AD document set. A targeted search found AD_07 (QPS
Cybersecurity Policy Framework) and AD_08 (QPLANT Abnormal Scenarios and
Line S Interface Clarifications) -- real, substantial documents that exist
on disk but were NEVER in DELIVERABLES_DOSSIER (which only had AD_01-06,
extracted from the contract mirror PDF's own Table 2).

Important distinction, not glossed over: AD_01-06 came from the contract's
own Table 2 (hence real SCKCEN/##### reference numbers). AD_07/08 are
project-side working documents GBO's team authored -- no SCKCEN reference
number appears in either document's own text, so none is fabricated here;
the Reference column instead states plainly that these are project working
documents, not contract Table 2 entries.

Both new entries' RTM links are NOT a fuzzy text-scan guess -- they're
taken directly from each document's own explicit self-declared anchor:
  - AD_07's own text: "Contractual anchor: RTM-322." -- verified against
    RTM_RANKING: RTM-322's shall statement is literally about the SCK CEN
    cybersecurity policy framework / CyFun, an exact match.
  - AD_08's own text: "...including RTM-260, RTM-261, RTM-292, and
    RTM-294..." -- verified all 4 exist in RTM_RANKING with matching
    subject matter (QPS abnormal events / QRB.S flow conditions).
(Confirmed via the same RTM-full-text scan used for AD_01-06 that neither
"AD_07" nor "AD_08" appears literally inside any RTM's own text -- so this
is a one-way reference from the AD document to the RTM, not (yet) mirrored
back from the RTM side.)

IN:  QPS_OFFER_Evaluation_FULL_v22.xlsx
OUT: QPS_OFFER_Evaluation_FULL_v23.xlsx
"""
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, PatternFill

IN = "QPS_OFFER_Evaluation_FULL_v22.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v23.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)
ws = wb["DELIVERABLES_DOSSIER"]

# Locate AD_06's row and the note row (37 and 39 respectively as of v22, but
# re-locate defensively rather than hardcoding in case row count drifted).
ad06_row = None
note_row = None
for r in range(6, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v == "AD_06":
        ad06_row = r
    if v and str(v).startswith("32 AD entries total"):
        note_row = r
assert ad06_row and note_row, f"couldn't locate anchor rows (ad06_row={ad06_row}, note_row={note_row})"

# The note row is a merged cell (A:F) -- openpyxl's insert_rows() does NOT
# shift merged-cell ranges, so the stale merge would be left sitting on top
# of one of the new rows and silently eat its B:F values on save (caught by
# re-reading the saved file during this script's own dev -- real bug, not
# hypothetical). Unmerge before inserting, re-merge at the new location
# after.
note_merge = None
for mc in list(ws.merged_cells.ranges):
    if mc.min_row == note_row and mc.max_row == note_row:
        note_merge = (mc.min_col, mc.max_col)
        ws.unmerge_cells(start_row=mc.min_row, start_column=mc.min_col, end_row=mc.max_row, end_column=mc.max_col)

# Insert 2 new rows right after AD_06, before the blank row + note.
insert_at = ad06_row + 1
ws.insert_rows(insert_at, amount=2)

NEW_ROWS = [
    ("AD_07", "QPS Cybersecurity Policy Framework (CPF)",
     "(project working doc -- no SCKCEN Table 2 reference; contractual anchor stated in the document itself: RTM-322)",
     "AD_07", 1, "RTM-322"),
    ("AD_08", "QPLANT Abnormal Scenarios and Line S Interface Clarifications",
     "(project working doc -- no SCKCEN Table 2 reference)",
     "AD_08", 4, "RTM-260; RTM-261; RTM-292; RTM-294"),
]
top_fill = PatternFill(start_color="F0E6F8", end_color="F0E6F8", fill_type="solid")
for i, row_vals in enumerate(NEW_ROWS):
    r = insert_at + i
    for c, val in enumerate(row_vals, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name="Carlito", size=10, bold=(c in (1, 2)))
        cell.fill = top_fill

# Update the note (now shifted down by 2 rows) to reflect the new totals:
# 34 AD entries total (8 top-level + 26 sub-items), 11 with an explicit link
# (9 original AD_01-06 citing-instances-based + 2 new AD_07/AD_08 self-
# declared anchors), 17 distinct RTMs now linked (13 original + 4 new: 322,
# 260, 261, 292, 294 -- but 322/260/261/292/294 are 5 new IDs... recompute
# precisely below rather than hand-count).
note_cell = ws.cell(row=note_row + 2, column=1)
old_note = note_cell.value
new_note = (
    "34 AD entries total (8 top-level groups + 26 sub-items -- AD_07 and AD_08 added this round, "
    "sourced from GBO's Master_Input working folder, not the contract mirror's Table 2). 11 of the 34 "
    "have at least one linked RTM: 9 of the original AD_01-06 group via RTM text explicitly citing the "
    "AD_## code by number (15 citing-instances, 13 distinct RTMs), plus AD_07 and AD_08 via each "
    "document's OWN explicit self-declared RTM anchor (not a text-scan match) -- AD_07->RTM-322, "
    "AD_08->RTM-260/261/292/294. The rest of the RTM corpus does not cite an AD_## reference directly -- "
    "those RTMs' evidence basis is the RTM's own section text, not one of these named document bundles."
)
note_cell.value = new_note
if note_merge:
    ws.merge_cells(start_row=note_row + 2, start_column=note_merge[0], end_row=note_row + 2, end_column=note_merge[1])

wb.save(OUT)
print(f"wrote {OUT}")
print(f"inserted AD_07/AD_08 at rows {insert_at}-{insert_at+1}")
print(f"note updated (was at row {note_row}, now at row {note_row+2})")
