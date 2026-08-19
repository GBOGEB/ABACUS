"""
build_workbook_slim_v24.py -- produces the v24 reviewer-shareable "Lite"
workbook from QPS_OFFER_Evaluation_FULL_v24.xlsx.

KEEP list is unchanged from v23's 23 sheets, PLUS one addition this round:
PCA_ANALYSIS. Per this script's own standing inclusion criterion (see
build_workbook_slim_v6.py's docstring, carried forward unchanged since):
a sheet is promoted to "keep" when it is itself a reviewer-facing
analytical view/finding/reference (like CLUSTERS, TAXONOMY, DASHBOARD_2),
not an internal audit/method-reproduction artefact (like WEIGHTS_METHOD,
DMAIC_AUDIT, CODING_HANDOVER). PCA_ANALYSIS (new in FULL_v24, built by
build_workbook_v24.py from compute_pca.py's output) is a disclosed findings
sheet -- explained-variance table, loadings table, domain-level PC1/PC2
positions, "bulk vs distinctive" split -- squarely a reviewer-facing
analytical view, so it is KEPT here. The 4th weight-scenario toggle table
that also landed in FULL_v24 lives on WEIGHTS_METHOD, which stays dropped
(unchanged reasoning: it's the live-formula method-reproduction sheet, not
a findings view) -- reviewers instead see the same scenario-4 numbers via
the kept PCA_ANALYSIS/CLUSTERS/TAXONOMY tabs and the Navigator, not by
exposing the toggle mechanism itself.

Still dropped (unchanged reasoning from v5/v6/v23): STANDARDS, DELIVERABLES,
AUDIT_NOTES, DMAIC_AUDIT, EVALUATION_INPUT, NAVIGATION_MAP, CODING_HANDOVER,
WEIGHTS_METHOD -- internal/audit/method-reproduction artefacts per
NAVIGATION_MAP's own role column. RECOMPUTE_LOG (present in both FULL_v23
and FULL_v24, unrelated to this round's PCA/weight-scenario-4 build) is also
dropped, consistent with v23's treatment -- it's an audit trail of a data
correction, not a reviewer-facing view.

Same dead-hyperlink neutralisation and formula-dependency discipline as the
v6/v23 slim-down scripts.
"""
import warnings
warnings.filterwarnings("ignore")
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

FULL = "QPS_OFFER_Evaluation_FULL_v24.xlsx"
OUT = "QPS_OFFER_Evaluation_LITE_v24.xlsx"

KEEP = [
    "README", "START_HERE", "DASHBOARD", "DASHBOARD_2", "RTM_LOOKUP", "COMPLIANCE_LEGEND",
    "EVALUATION_WORKSPACE", "NEGOTIATION_AGENDA", "OFFER_RANKING", "RTM_CROSSWALK",
    "RTM_RANKING", "CLUSTERS", "TAXONOMY", "DOMAIN_SUMMARY", "RTM_REVIEW_QUEUE",
    "QUALITY_CHECKS", "LISTS", "OFFER_CANONICAL", "CATEGORY_FOCUS",
    # v20 additions -- all 4 are reviewer-facing tools (worklist / analytical
    # view / reference / heuristic flag list), same promotion logic as the
    # v6 round's RTM_LOOKUP/DASHBOARD_2/CLUSTERS/TAXONOMY additions.
    "REVIEW_FOCUS", "RTM_PHASE_EXPANSION", "DELIVERABLES_DOSSIER", "CONFLICT_CANDIDATES",
    # v24 addition -- PCA_ANALYSIS is a disclosed-findings/analytical view
    # (explained variance, loadings, domain PC1/PC2 positions), same
    # promotion logic as above; see module docstring for the full reasoning.
    "PCA_ANALYSIS",
]
HYPERLINK_RE = re.compile(r'=HYPERLINK\("#([A-Za-z0-9_]+)!', re.IGNORECASE)
DEAD_FILL = PatternFill("solid", fgColor="E8E8EC")
DEAD_FONT = Font(name="Carlito", size=11, italic=True, color="9A9AA5")

wb = openpyxl.load_workbook(FULL, data_only=False)

# ---- pre-flight: full cross-sheet formula-dependency scan before dropping -
CELL_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!\$?[A-Z]{1,3}\$?\d+")
dropped_candidates = [n for n in wb.sheetnames if n not in KEEP]
live_deps = set()
for name in KEEP:
    ws = wb[name]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=") and "HYPERLINK" not in cell.value:
                for m in CELL_REF_RE.finditer(cell.value):
                    ref = m.group(1) or m.group(2)
                    if ref in dropped_candidates:
                        live_deps.add((name, ref, cell.coordinate))
if live_deps:
    print("WARNING -- live formula dependencies on sheets about to be dropped:")
    for d in live_deps:
        print("  ", d)
else:
    print("pre-flight OK: no kept-sheet formula depends on a dropped sheet")

dropped = [n for n in wb.sheetnames if n not in KEEP]
for name in dropped:
    del wb[name]
print(f"dropped {len(dropped)} sheets: {dropped}")
print(f"kept {len(wb.sheetnames)} sheets: {wb.sheetnames}")

wb["LISTS"].sheet_state = "hidden"

sh = wb["START_HERE"]
if sh["A1"].value and "DMAIC-ready" in str(sh["A1"].value):
    sh["A1"] = "QPS OFFER evaluation workbook — v24 LITE (reviewer edition)"
sh["A2"] = (
    "Reviewer-shareable subset of the full evaluation workbook -- ranking, crosswalk, "
    "workspace, quality, lookup, filter-dashboard, cluster, taxonomy and PCA-findings tabs. "
    "Internal audit/handover/method-control tabs (incl. WEIGHTS_METHOD's live toggle "
    "mechanism) live in the full v24 workbook only."
)

n_neutralised = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue
            m = HYPERLINK_RE.match(cell.value)
            if not m:
                continue
            target_sheet = m.group(1)
            if target_sheet in wb.sheetnames:
                continue
            label_match = re.search(r',"([^"]*)"\)$', cell.value)
            label = label_match.group(1) if label_match else target_sheet
            cell.value = label
            cell.font = DEAD_FONT
            cell.fill = DEAD_FILL
            n_neutralised += 1
print(f"neutralised {n_neutralised} dead cross-sheet hyperlinks (dropped-tab references)")

for row_num in (7, 8):
    for col in ("A", "C", "D"):
        cell = sh[f"{col}{row_num}"]
        if cell.value:
            cell.font = DEAD_FONT

wb.save(OUT)
print(f"saved {OUT}")
