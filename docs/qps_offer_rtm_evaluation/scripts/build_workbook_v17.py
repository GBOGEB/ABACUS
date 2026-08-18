"""
build_workbook_v17.py -- four fixes/additions from GBO's latest round:

1. STANDARDS!B ("Scope / Description") lazy-boilerplate fix -- GBO asked
   "Is the standards and proofs fully utilized and understood or can be
   dropped?" Investigation: the sheet IS real, used cross-reference data
   (44 codes/standards, each with real linked RTM IDs, OFFER IDs and
   section numbers pulled from the canonical documents) -- NOT droppable.
   But 26 of 44 rows (59%) share the exact same generic boilerplate
   "Scope defined by cited canonical requirement/section" in the
   description column -- the same low-information pattern already found
   and fixed twice before in this workbook (DELIVERABLES!F in v8,
   RTM_CROSSWALK!L in v11). Fixed the same way: derive a real, grounded
   description from the linked RTM's own Category/Subcategory/shall-text
   (already-curated data), never invented external knowledge about the
   standard itself.

2. Weight-dimension letter legend + visual emphasis -- GBO: "make mor
   visual enmpahsuys in workbook ... the Category Weights L, R, P F Q LC
   and C". OFFER_RANKING!H:N and RTM_RANKING!O:U both use bare
   single/double-letter headers (L, R, P, F, Q, LC, C) with no legend
   anywhere in the workbook explaining what they stand for. Added: (a) a
   colour-coded "Weight dimension legend" table to WEIGHTS_METHOD showing
   code -> full name -> frozen weight -> colour swatch, (b) the same
   colour applied to the matching header cell fill on both ranking
   sheets, (c) a cell comment on each header spelling out the full name,
   so hovering/looking at either sheet directly answers "what is L".

3. Scenario toggle ("bizarro world" live what-if) -- GBO: "toggle (would
   be final) but where you switch the output based on Cost 0.03 or 0.7 of
   total to see impact immediate". The v14 Cost=70% analysis was a STATIC
   top-12-movers table. This adds a real interactive toggle: a
   data-validation dropdown cell (Base / Cost-heavy 70%) that drives a
   live-recomputed, live-re-ranked 50-row OFFER table via formulas
   (SUMPRODUCT against a 2-row scenario-weight lookup keyed by the
   toggle) -- switching the dropdown recomputes Scenario S and Scenario
   rank immediately, no macro, no re-run needed. Still clearly labelled
   illustrative / does not touch the frozen official ranking.

4. README + CHANGELOG sheets -- GBO asked (repeatedly, across several
   messages) for something explaining "which button to push", tab
   colour/grouping conventions, and "log for version and major changes".
   START_HERE partly covers navigation already; README adds a single
   one-stop orientation page (what every sheet is, in the 6-group tab-
   colour taxonomy from v13) and CHANGELOG lists v8-v17 in one place --
   both were previously only reconstructable by reading commit-style
   docstrings in the build scripts, which GBO cannot see.
"""
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

IN = "QPS_OFFER_Evaluation_FULL_v16.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v17.xlsx"

wb = openpyxl.load_workbook(IN, data_only=False)

NAVY = "1F4E78"
CARLITO = "Carlito"

# =========================================================== 1. STANDARDS!B fix
rtm_meta = {}
rws = wb["RTM_RANKING"]
for r in range(6, rws.max_row + 1):
    rid = rws.cell(r, 2).value
    if not rid:
        continue
    rtm_meta[rid] = {
        "category": rws.cell(r, 7).value or "",
        "subcategory": rws.cell(r, 8).value or "",
        "shall": rws.cell(r, 13).value or "",
    }

BOILER = "Scope defined by cited canonical requirement/section"
sws = wb["STANDARDS"]
n_fixed = 0
for r in range(6, sws.max_row + 1):
    desc = sws.cell(r, 2).value
    if desc != BOILER:
        continue
    rtm_ids_raw = sws.cell(r, 3).value or ""
    first_id = str(rtm_ids_raw).split(";")[0].strip()
    meta = rtm_meta.get(first_id)
    if not meta:
        continue
    shall = str(meta["shall"]).strip()
    shall_snip = (shall[:110] + "…") if len(shall) > 110 else shall
    cat = meta["category"]
    sub = meta["subcategory"]
    new_desc = f"{cat} / {sub}: “{shall_snip}”" if shall_snip else f"{cat} / {sub}"
    cell = sws.cell(r, 2, value=new_desc)
    cell.font = Font(name=CARLITO, italic=True, color="5B7FA6")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    n_fixed += 1
sws.column_dimensions["B"].width = 48
for r in range(6, sws.max_row + 1):
    sws.row_dimensions[r].height = 32
print(f"STANDARDS!B rewritten: {n_fixed} rows (grounded in linked RTM's own Category/Subcategory/shall text)")

last_row = sws.max_row + 2
note = sws.cell(row=last_row, column=1,
    value=("Column B methodology (Phase 17): rows that previously read only \"Scope defined by cited "
           "canonical requirement/section\" were rewritten from the first linked RTM's own Category / "
           "Subcategory and a clipped snippet of its shall-statement (RTM_RANKING) -- not invented "
           "knowledge about the standard itself. This sheet is real, used cross-reference data (44 "
           "codes/standards, each with genuine linked RTM/OFFER IDs and section numbers) -- kept, not dropped."))
note.font = Font(name=CARLITO, italic=True, size=10.5, color="444444")
note.alignment = Alignment(wrap_text=True, vertical="top")
sws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=7)
sws.row_dimensions[last_row].height = 60

# =========================================================== 2. Weight-dimension legend + colour
DIM_INFO = [
    # code, full name, weight, colour
    ("L",  "Safety / Legal",         0.20, "F4CCCC"),
    ("R",  "Reliability",            0.22, "D9EAD3"),
    ("P",  "Performance",            0.20, "D6E9F8"),
    ("F",  "Functional",             0.16, "E6DFF6"),
    ("Q",  "Quality / Verifiability",0.12, "FCE5CD"),
    ("LC", "Lifecycle",              0.07, "D9D2E9"),
    ("C",  "Cost",                   0.03, "FFF2CC"),
]

wm = wb["WEIGHTS_METHOD"]
r0 = wm.max_row + 3
wm.cell(row=r0, column=1, value="Weight dimension legend (code -> full name, as used on OFFER_RANKING!H:N and RTM_RANKING!O:U)").font = Font(name=CARLITO, size=13, bold=True, color=NAVY)
r0 += 1
hdrs = ["Code", "Full dimension name", "Frozen weight", "Colour"]
for c, h in enumerate(hdrs, start=1):
    cell = wm.cell(row=r0, column=c, value=h)
    cell.font = Font(name=CARLITO, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
r0 += 1
legend_start = r0
for code, name, weight, colour in DIM_INFO:
    wm.cell(row=r0, column=1, value=code).font = Font(name=CARLITO, bold=True)
    wm.cell(row=r0, column=2, value=name).font = Font(name=CARLITO)
    wm.cell(row=r0, column=3, value=weight).number_format = "0%"
    swatch = wm.cell(row=r0, column=4, value="")
    swatch.fill = PatternFill("solid", fgColor=colour)
    r0 += 1
wm.column_dimensions["D"].width = 10
print("WEIGHTS_METHOD legend table added at row", legend_start)

# apply colour + comment to header row on both ranking sheets
TARGETS = [
    ("OFFER_RANKING", 8),   # H
    ("RTM_RANKING", 15),    # O
]
for sheet, start_col in TARGETS:
    ws = wb[sheet]
    for i, (code, name, weight, colour) in enumerate(DIM_INFO):
        col = start_col + i
        cell = ws.cell(row=5, column=col)
        cell.fill = PatternFill("solid", fgColor=colour)
        cell.font = Font(name=CARLITO, bold=True, color="1A1A1A")
        cell.comment = Comment(f"{name} — frozen weight {weight:.0%} of Weighted S. See WEIGHTS_METHOD for the full legend and construction formula.", "Workbook")
    print(f"{sheet}: header colours + comments applied to {get_column_letter(start_col)}:{get_column_letter(start_col+6)}")

# =========================================================== 2b. Uniform row heatmap across all 7 raw-score columns
# GBO: "all colours and single rows to be filled by colour to see the
# 'coverage' ... way to look at RTM and OFFER in slight different BT lens".
# Before this fix only 3 of the 7 raw 0-3 relevance columns had ANY visual
# treatment (F had a colorScale, Q/LC had dataBars, L/R/P/C had none) --
# inconsistent and made a row hard to scan at a glance. Replaced with ONE
# consistent 2-colour scale (white=0 -> that dimension's own legend colour
# at 3) on all 7 columns, both sheets -- so a full row now reads as a
# little heatmap: which of the 7 dimensions this item scores strongly on
# jumps out by colour, not just by number, and the colour identity matches
# the legend (same hue family as the header fill / WEIGHTS_METHOD swatch).
from openpyxl.formatting.rule import ColorScaleRule

STRONG_COLORS = {
    "L": "CC0000",   # Safety/Legal -- strong red (matches pale F4CCCC header)
    "R": "2E7D32",   # Reliability -- strong green
    "P": "2B6CB0",   # Performance -- strong blue
    "F": "674EA7",   # Functional -- strong purple
    "Q": "B45F06",   # Quality/Verifiability -- strong orange
    "LC": "6A3D9A",  # Lifecycle -- strong violet (distinct from Functional's purple)
    "C": "BF9000",   # Cost -- strong gold
}
for sheet, start_col in TARGETS:
    ws = wb[sheet]
    last_row = ws.max_row
    for i, (code, name, weight, colour) in enumerate(DIM_INFO):
        col = start_col + i
        col_letter = get_column_letter(col)
        rng = f"{col_letter}6:{col_letter}{last_row}"
        # drop any pre-existing CF on this exact range so rules don't stack/conflict
        try:
            del ws.conditional_formatting[rng]
        except KeyError:
            pass
        rule = ColorScaleRule(
            start_type="num", start_value=0, start_color="FFFFFF",
            end_type="num", end_value=3, end_color=STRONG_COLORS[code],
        )
        ws.conditional_formatting.add(rng, rule)
    print(f"{sheet}: uniform 0-3 heatmap applied to all 7 dimension columns ({get_column_letter(start_col)}:{get_column_letter(start_col+6)})")

# =========================================================== 3. Scenario toggle (live what-if)
wm = wb["WEIGHTS_METHOD"]
r0 = wm.max_row + 3
wm.cell(row=r0, column=1, value="Live scenario toggle -- switch Cost weight and see the impact immediately").font = Font(name=CARLITO, size=13, bold=True, color=NAVY)
r0 += 1
wm.cell(row=r0, column=1,
    value=("Pick a scenario in the yellow cell below; the table recomputes live via formulas (no macro). "
           "Illustrative only -- never changes the frozen official ranking on OFFER_RANKING. Both scenarios "
           "split the non-Cost weight equally across the other 6 dimensions, the simplest scheme to explain. "
           "'Live rank' deliberately ranks by score alone (no gate precedence) -- same method as the static "
           "Cost=70% table below -- so it shows what happens to today's gate items on merit score alone, "
           "which is the whole point of the toggle. That means even under 'Base' weights, Live rank will "
           "differ slightly from the official Base rank (col C) for gate items and exact ties -- expected, "
           "not an error; col C is always the real official rank.")
).font = Font(name=CARLITO, italic=True, size=10.5, color="444444")
wm.merge_cells(start_row=r0, start_column=1, end_row=r0+1, end_column=8)
wm.row_dimensions[r0].height = 28
wm.row_dimensions[r0+1].height = 28
r0 += 3

toggle_label_row = r0
wm.cell(row=r0, column=1, value="Scenario:").font = Font(name=CARLITO, bold=True)
toggle_cell = wm.cell(row=r0, column=2, value="Base (Cost=3%)")
toggle_cell.fill = PatternFill("solid", fgColor="FFF2CC")
toggle_cell.font = Font(name=CARLITO, bold=True, color="7A4E00")
toggle_ref = f"WEIGHTS_METHOD!$B${toggle_label_row}"
dv = DataValidation(type="list", formula1='"Base (Cost=3%),Cost-heavy (Cost=70%)"', allow_blank=False)
wm.add_data_validation(dv)
dv.add(toggle_cell)
r0 += 2

# hidden scenario-weight lookup table (2 rows x 7 dims), used by the live formulas
lut_hdr_row = r0
wm.cell(row=r0, column=1, value="(scenario weight lookup -- feeds the formulas below, not for manual editing)").font = Font(name=CARLITO, italic=True, size=9, color="999999")
r0 += 1
lut_start = r0
codes = [d[0] for d in DIM_INFO]
for c, code in enumerate(codes, start=2):
    wm.cell(row=r0, column=c, value=code).font = Font(name=CARLITO, bold=True, size=9, color="999999")
wm.cell(row=r0, column=1, value="Base").font = Font(name=CARLITO, size=9, color="999999")
r0 += 1
base_row = r0
BASE_W = {"L":0.20,"R":0.22,"P":0.20,"F":0.16,"Q":0.12,"LC":0.07,"C":0.03}
for c, code in enumerate(codes, start=2):
    wm.cell(row=r0, column=c, value=BASE_W[code]).number_format = "0.00"
    wm.cell(row=r0, column=c).font = Font(size=9, color="999999")
wm.cell(row=r0, column=1, value="Base").font = Font(name=CARLITO, size=9, color="999999")
r0 += 1
scen_row = r0
SCEN_W = {"L":0.05,"R":0.05,"P":0.05,"F":0.05,"Q":0.05,"LC":0.05,"C":0.70}
for c, code in enumerate(codes, start=2):
    wm.cell(row=r0, column=c, value=SCEN_W[code]).number_format = "0.00"
    wm.cell(row=r0, column=c).font = Font(size=9, color="999999")
wm.cell(row=r0, column=1, value="Cost-heavy").font = Font(name=CARLITO, size=9, color="999999")
r0 += 2

# live table: 50 OFFER rows, weight per dim picked live via IF(toggle=...) per column
table_hdr_row = r0
headers2 = ["OFFER ID", "Title", "Base rank", "Live rank", "Base Weighted S", "Live Weighted S", "Rank delta"]
for c, h in enumerate(headers2, start=1):
    cell = wm.cell(row=r0, column=c, value=h)
    cell.font = Font(name=CARLITO, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
r0 += 1
data_start = r0

ow = wb["OFFER_RANKING"]
n_offer = 0
for i in range(50):
    orow = 6 + i          # OFFER_RANKING data row
    wrow = data_start + i  # WEIGHTS_METHOD table row
    wm.cell(row=wrow, column=1, value=f"=OFFER_RANKING!B{orow}")
    wm.cell(row=wrow, column=2, value=f"=OFFER_RANKING!C{orow}")
    wm.cell(row=wrow, column=3, value=f"=OFFER_RANKING!A{orow}")
    # live weighted S: SUMPRODUCT(raw scores H:N, live weight row selected by toggle)/3*100
    live_w_row = f"IF($B${toggle_label_row}=\"Base (Cost=3%)\",{base_row},{scen_row})"
    s_formula = (f"=100*SUMPRODUCT(OFFER_RANKING!H{orow}:N{orow},"
                 f"INDEX($B${base_row}:$H${scen_row},{live_w_row}-{base_row}+1,0))/3")
    wm.cell(row=wrow, column=6, value=s_formula).number_format = "0.00"
    wm.cell(row=wrow, column=5, value=f"=OFFER_RANKING!O{orow}").number_format = "0.00"
    n_offer += 1

# rank + delta formulas (need full column range known up front). Deliberately IGNORES gate
# precedence (pure Live-S rank, same methodology as the v14 static Cost=70% table) -- the whole
# point of this toggle is to show where each item would land on merit score alone, INCLUDING
# what happens to today's gate items once Cost dominates. So even under "Base" weights, Live
# rank differs slightly from the official Base rank (which applies gate precedence + a fixed
# tie-break); see the note above the table.
data_end = data_start + n_offer - 1
for i in range(n_offer):
    wrow = data_start + i
    wm.cell(row=wrow, column=4, value=f"=RANK(F{wrow},$F${data_start}:$F${data_end},0)")
    wm.cell(row=wrow, column=7, value=f"=C{wrow}-D{wrow}")
    wm.cell(row=wrow, column=7).font = Font(name=CARLITO, bold=True)

wm.column_dimensions["B"].width = 34
for c in ("A","C","D","E","F","G"):
    wm.column_dimensions[c].width = 14
print(f"Live scenario toggle table added: rows {data_start}-{data_end}, driven by toggle at B{toggle_label_row}")

# conditional colour on rank delta (reuse the same green/red convention as v14's static table)
from openpyxl.formatting.rule import CellIsRule
delta_rng = f"G{data_start}:G{data_end}"
wm.conditional_formatting.add(delta_rng, CellIsRule(operator="greaterThan", formula=["0"],
    fill=PatternFill("solid", fgColor="D9EAD3"), font=Font(color="2E7D32", bold=True)))
wm.conditional_formatting.add(delta_rng, CellIsRule(operator="lessThan", formula=["0"],
    fill=PatternFill("solid", fgColor="F4CCCC"), font=Font(color="A30000", bold=True)))

# =========================================================== 4. README + CHANGELOG
GROUPS = {
    "Navigate":            ("1F4E78", ["START_HERE", "NAVIGATION_MAP", "LISTS"]),
    "Overview":            ("562873", ["DASHBOARD", "DASHBOARD_2"]),
    "Live review / input": ("C99A00", ["EVALUATION_WORKSPACE", "EVALUATION_INPUT", "NEGOTIATION_AGENDA"]),
    "Rankings & trace":    ("1FA7A0", ["OFFER_RANKING", "RTM_RANKING", "RTM_CROSSWALK", "RTM_REVIEW_QUEUE",
                                        "RTM_LOOKUP", "DOMAIN_SUMMARY", "CLUSTERS"]),
    "Reference / canonical": ("0B5FA5", ["OFFER_CANONICAL", "STANDARDS", "DELIVERABLES", "TAXONOMY",
                                          "COMPLIANCE_LEGEND", "WEIGHTS_METHOD"]),
    "Governance / QC":     ("A64CA6", ["QUALITY_CHECKS", "AUDIT_NOTES", "DMAIC_AUDIT", "CODING_HANDOVER"]),
}
SHEET_ONE_LINERS = {
    "START_HERE": "Landing page: what this workbook is, where to start.",
    "NAVIGATION_MAP": "Full sheet index with hyperlinks, one row per sheet, coloured by group.",
    "LISTS": "Raw dropdown/reference lists used by data validation elsewhere.",
    "DASHBOARD": "Top-line charts: Top-15 OFFER by S, Tier split, Category averages, full 50-item rank profile (with the item-7 spike explainer).",
    "DASHBOARD_2": "Secondary dashboard views.",
    "EVALUATION_WORKSPACE": "Main interaction sheet -- static rank read-only, reviewer input starts at STATUS columns.",
    "EVALUATION_INPUT": "Raw bidder input capture, feeds EVALUATION_WORKSPACE.",
    "NEGOTIATION_AGENDA": "Talking points generated from low-status / high-S items, for negotiation prep.",
    "OFFER_RANKING": "Static BT ranking of all 50 OFFER items -- the SSOT for OFFER importance.",
    "RTM_RANKING": "Static BT ranking of all 722 RTM requirements -- the SSOT for RTM importance.",
    "RTM_CROSSWALK": "One row per OFFER<->RTM relationship (up to 66 per item) -- the full link detail behind OFFER_RANKING's condensed 'Primary RTMs' summary.",
    "RTM_REVIEW_QUEUE": "RTM rows flagged for reviewer attention, ordered by priority.",
    "RTM_LOOKUP": "Pick one RTM, see everything about it on one screen.",
    "DOMAIN_SUMMARY": "RTM counts/tiers/avg S rolled up by Domain (22 domains).",
    "CLUSTERS": "The 8 classification clusters (C1-C8) and their member OFFER items.",
    "OFFER_CANONICAL": "Original hand-curated source data for all 50 OFFER items -- read-only upstream of OFFER_RANKING.",
    "STANDARDS": "Codes/standards register -- which RTM/OFFER items cite which external standard.",
    "DELIVERABLES": "Per-OFFER-item evidence/fulfilment expectations.",
    "TAXONOMY": "Controlled vocabularies used throughout (Requirement Type, EPIC evidence-types, disposition, etc.) -- the single place these are DEFINED, reused elsewhere.",
    "COMPLIANCE_LEGEND": "STATUS colour/label legend used on EVALUATION_WORKSPACE.",
    "WEIGHTS_METHOD": "The 7 frozen BT weights (L/R/P/F/Q/LC/C), how Weighted S/BT Win %/BT λ are built, and the weight-sensitivity scenarios (static + live toggle).",
    "QUALITY_CHECKS": "Automated formula-integrity and cross-reference checks over the whole workbook.",
    "AUDIT_NOTES": "Free-text audit trail of judgement calls made during data curation.",
    "DMAIC_AUDIT": "DMAIC-cycle traceability for this evaluation process itself.",
    "CODING_HANDOVER": "Technical handover notes for whoever maintains this workbook next.",
}

if "README" in wb.sheetnames:
    del wb["README"]
readme = wb.create_sheet("README", 0)
readme.sheet_view.showGridLines = False
readme.column_dimensions["A"].width = 26
readme.column_dimensions["B"].width = 70
readme.column_dimensions["C"].width = 12
r = 1
readme.cell(row=r, column=1, value="README — how this workbook is organised").font = Font(name=CARLITO, size=16, bold=True, color="FFFFFF")
readme.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
readme.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
readme.row_dimensions[r].height = 26
r += 2
readme.cell(row=r, column=1, value=("25 sheets, grouped into 6 colour-coded families (tab colour = group). "
    "Start with START_HERE, then NAVIGATION_MAP for the full clickable index. This page is the one-stop "
    "explanation of what every sheet is and why it's coloured the way it is -- see CHANGELOG for what changed and when.")
).font = Font(name=CARLITO, italic=True, color="444444")
readme.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
readme.row_dimensions[r].height = 30
r += 2

for group, (colour, sheets) in GROUPS.items():
    ghdr = readme.cell(row=r, column=1, value=group)
    ghdr.font = Font(name=CARLITO, bold=True, color="FFFFFF")
    ghdr.fill = PatternFill("solid", fgColor=colour)
    readme.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        # HYPERLINK() formula (not a native .hyperlink attribute) so the LITE-workbook
        # builder's dead-link neutralisation (which regex-matches HYPERLINK formulas)
        # can find and grey out links to sheets LITE drops -- same convention as every
        # other nav row in this workbook.
        c1 = readme.cell(row=r, column=1, value=f'=HYPERLINK("#{sheet}!A1","{sheet}")')
        c1.font = Font(name=CARLITO, bold=True, color=colour)
        c1.alignment = Alignment(vertical="top")
        c2 = readme.cell(row=r, column=2, value=SHEET_ONE_LINERS.get(sheet, ""))
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        one_liner = SHEET_ONE_LINERS.get(sheet, "")
        lines = max(1, -(-len(one_liner) // 60))  # ceil division, ~60 chars/line at this col width
        readme.row_dimensions[r].height = max(18, lines * 14 + 6)
        r += 1
    r += 1
readme.sheet_properties.tabColor = NAVY
print("README sheet created with", r, "rows")

if "CHANGELOG" in wb.sheetnames:
    del wb["CHANGELOG"]
changelog = wb.create_sheet("CHANGELOG", 1)
changelog.sheet_view.showGridLines = False
changelog.column_dimensions["A"].width = 10
changelog.column_dimensions["B"].width = 90
r = 1
changelog.cell(row=r, column=1, value="CHANGELOG").font = Font(name=CARLITO, size=16, bold=True, color="FFFFFF")
changelog.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
changelog.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
changelog.row_dimensions[r].height = 26
r += 2
changelog.cell(row=r, column=1, value="Version").font = Font(name=CARLITO, bold=True, color="FFFFFF")
changelog.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
changelog.cell(row=r, column=2, value="Major changes").font = Font(name=CARLITO, bold=True, color="FFFFFF")
changelog.cell(row=r, column=2).fill = PatternFill("solid", fgColor=NAVY)
r += 1
CHANGES = [
    ("v8",  "DELIVERABLES!F evidence text rewritten from generic boilerplate to per-item RTM cross-reference + EPIC-keyed evidence type."),
    ("v9",  "Full 722 RTM / 50 OFFER value-weighted Pareto and Focus Score analysis added to the HTML Navigator (Domain/Category/Cluster, for both RTM and OFFER)."),
    ("v10", "RTM_CROSSWALK: added a real header row (row 4 was blank), fixed a misleading dark-navy fill on row 5 that looked like a header, added AutoFilter/frozen panes, and colour-coded relation types (Direct/Supporting/Broad/Contextual)."),
    ("v11", "RTM_CROSSWALK: added headers for columns M/N (QC check, Source basis); rewrote 233 of 377 rows (62%) of column L from generic boilerplate to genuinely per-row text."),
    ("v12", "Excel window/zoom defaults set for 16:9 full-screen opening; per-sheet autozoom by column-count."),
    ("v13", "Tab-colour taxonomy: all 25 sheets assigned to 6 named groups; NAVIGATION_MAP extended with a colour-swatch legend column."),
    ("v14", "Nav-row column-clipping fixed across multiple sheets; Cost=70% weight-sensitivity scenario (static table) added to WEIGHTS_METHOD."),
    ("v15", "Technical Depth (EVALUATION_WORKSPACE R/X) colour-coded (soft pink/orange/green), deliberately paler than STATUS."),
    ("v16", "Gate column (Yes/No) recoloured to reuse the exact T0 Gate red across EVALUATION_WORKSPACE/OFFER_RANKING/RTM_RANKING; DASHBOARD's rank-profile chart got a callout explaining the item-7 spike (OFFER-09's S=55.67 entering rank 7, right after the 6 gate-precedence items)."),
    ("v17", "STANDARDS!B lazy-boilerplate fixed (26 of 44 rows, grounded in linked RTM data); weight-dimension (L/R/P/F/Q/LC/C) colour legend + header comments added to OFFER_RANKING/RTM_RANKING/WEIGHTS_METHOD; all 7 raw dimension columns given a uniform 0-3 heatmap (previously only 3 of 7 had any colour) on both ranking sheets; live Cost-weight scenario toggle added (dropdown + formula-driven re-rank, no macro); new CATEGORY_FOCUS sheet (per-category top-ranked 'hidden champions' + Category x Cluster density heat maps, RTM and OFFER); README + this CHANGELOG sheet added."),
]
for ver, desc in CHANGES:
    changelog.cell(row=r, column=1, value=ver).font = Font(name=CARLITO, bold=True, color=NAVY)
    c2 = changelog.cell(row=r, column=2, value=desc)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    changelog.row_dimensions[r].height = 34
    r += 1
changelog.sheet_properties.tabColor = NAVY
print("CHANGELOG sheet created,", len(CHANGES), "entries")

# =========================================================== 5. CATEGORY_FOCUS -- "hidden champions" + density heat maps
# GBO: "hidden champions or local focus points, ... per category ranking -
# top per rank? ... density of items across all or specific combinations -
# new sheet". A globally-low-ranked item can still be the clear #1 choice
# within its own Category -- worth flagging separately, since a pure global
# rank list buries that. Built directly from OFFER_RANKING/RTM_RANKING (no
# new source data, pure aggregation), same convention as DOMAIN_SUMMARY.
from openpyxl.formatting.rule import ColorScaleRule as _CSR

if "CATEGORY_FOCUS" in wb.sheetnames:
    del wb["CATEGORY_FOCUS"]
cf = wb.create_sheet("CATEGORY_FOCUS")
cf.sheet_properties.tabColor = "1FA7A0"  # same teal as the other Rankings & trace sheets
cf.sheet_view.showGridLines = False

r = 1
cf.cell(row=r, column=1, value="Category focus — hidden champions & density").font = Font(name=CARLITO, size=16, bold=True, color="FFFFFF")
cf.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
cf.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cf.row_dimensions[r].height = 26
r += 1
cf.cell(row=r, column=1, value=("A 'hidden champion' is the top-ranked item WITHIN its own Category, even if its "
    "GLOBAL rank is much lower -- a pure global top-N list buries these. Density heat maps below show where "
    "items concentrate across Category x Cluster.")).font = Font(name=CARLITO, italic=True, color="444444")
cf.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cf.row_dimensions[r].height = 28
r += 1
navlinks = ["START_HERE","DASHBOARD","OFFER_RANKING","RTM_RANKING","CLUSTERS"]
for i, sn in enumerate(navlinks, start=1):
    cf.cell(row=r, column=i, value=f'=HYPERLINK("#{sn}!A1","{sn.title().replace("_"," ")}")')
cf.row_dimensions[r].height = 18
r += 2

def section_title(ws, r, text):
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=CARLITO, size=13, bold=True, color=NAVY)
    return r + 1

# ---- gather data ----
offer_rows, rtm_rows = [], []
ow = wb["OFFER_RANKING"]
for rr in range(6, ow.max_row + 1):
    oid = ow.cell(rr, 2).value
    if not oid: continue
    offer_rows.append({"id": oid, "title": ow.cell(rr,3).value, "gate": ow.cell(rr,4).value,
                        "tier": ow.cell(rr,5).value, "cat": ow.cell(rr,6).value,
                        "s": ow.cell(rr,15).value or 0, "rank": ow.cell(rr,1).value})
rw = wb["RTM_RANKING"]
for rr in range(6, rw.max_row + 1):
    rid = rw.cell(rr, 2).value
    if not rid: continue
    rtm_rows.append({"id": rid, "shall": rw.cell(rr,13).value, "tier": rw.cell(rr,4).value,
                      "cat": rw.cell(rr,7).value, "cluster": rw.cell(rr,9).value,
                      "s": rw.cell(rr,22).value or 0, "rank": rw.cell(rr,1).value})

offer_cluster_map = {}
for crow in range(6, wb["CLUSTERS"].max_row + 1):
    cid = wb["CLUSTERS"].cell(crow, 1).value
    items = wb["CLUSTERS"].cell(crow, 3).value or ""
    if not cid: continue
    for tok in str(items).split(";"):
        tok = tok.strip()
        if tok: offer_cluster_map[tok] = cid

import re as _re
def cluster_key(c):
    if not c: return "Not linked"
    m = _re.search(r"C[1-8]", str(c))
    return m.group(0) if m else "Not linked"

# ---- section A: hidden champions, OFFER (top 3 per category) ----
r = section_title(cf, r, "Hidden champions — OFFER (top 3 by Weighted S, within each Category)")
hdrs = ["Category","Local rank","OFFER ID","Title","Weighted S","Global rank","Flag"]
for c,h in enumerate(hdrs, start=1):
    cell = cf.cell(row=r, column=c, value=h); cell.font=Font(name=CARLITO,bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor=NAVY)
r += 1
from collections import defaultdict, Counter
by_cat = defaultdict(list)
for row in offer_rows: by_cat[row["cat"]].append(row)
n_global = len(offer_rows)
for cat in sorted(by_cat):
    items = sorted(by_cat[cat], key=lambda x: -x["s"])[:3]
    for i, it in enumerate(items, start=1):
        hidden = (i == 1 and it["rank"] and it["rank"] > n_global * 0.4)
        cf.cell(row=r, column=1, value=cat if i==1 else "")
        cf.cell(row=r, column=2, value=i)
        cf.cell(row=r, column=3, value=it["id"])
        cf.cell(row=r, column=4, value=it["title"])
        cf.cell(row=r, column=5, value=round(it["s"],2))
        cf.cell(row=r, column=6, value=it["rank"])
        flagcell = cf.cell(row=r, column=7, value=("Hidden champion — locally #1, globally rank "+str(it["rank"])+" of "+str(n_global)) if hidden else "")
        if hidden:
            for c in range(1,8):
                cf.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFF2CC")
            flagcell.font = Font(name=CARLITO, bold=True, color="7A4E00")
        r += 1
r += 1

# ---- section B: hidden champions, RTM (top 5 per category) ----
r = section_title(cf, r, "Hidden champions — RTM (top 5 by Weighted S, within each Category)")
for c,h in enumerate(["Category","Local rank","RTM ID","Shall (clipped)","Weighted S","Global rank","Flag"], start=1):
    cell = cf.cell(row=r, column=c, value=h); cell.font=Font(name=CARLITO,bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor=NAVY)
r += 1
by_cat_r = defaultdict(list)
for row in rtm_rows: by_cat_r[row["cat"]].append(row)
n_global_r = len(rtm_rows)
for cat in sorted(by_cat_r):
    items = sorted(by_cat_r[cat], key=lambda x: -x["s"])[:5]
    for i, it in enumerate(items, start=1):
        hidden = (i == 1 and it["rank"] and it["rank"] > n_global_r * 0.4)
        shall = str(it["shall"] or "")
        shall_snip = (shall[:90] + "…") if len(shall) > 90 else shall
        cf.cell(row=r, column=1, value=cat if i==1 else "")
        cf.cell(row=r, column=2, value=i)
        cf.cell(row=r, column=3, value=it["id"])
        cf.cell(row=r, column=4, value=shall_snip)
        cf.cell(row=r, column=5, value=round(it["s"],2))
        cf.cell(row=r, column=6, value=it["rank"])
        flagcell = cf.cell(row=r, column=7, value=("Hidden champion — locally #1, globally rank "+str(it["rank"])+" of "+str(n_global_r)) if hidden else "")
        if hidden:
            for c in range(1,8):
                cf.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFF2CC")
            flagcell.font = Font(name=CARLITO, bold=True, color="7A4E00")
        r += 1
r += 1

# ---- section C: density heat map, RTM Category x Cluster ----
r = section_title(cf, r, "Density heat map — RTM Category x Cluster (item counts)")
rtm_cats = sorted(by_cat_r.keys())
cluster_ids = ["C1","C2","C3","C4","C5","C6","C7","C8","Not linked"]
cf.cell(row=r, column=1, value="Category")
for j, cid in enumerate(cluster_ids, start=2):
    cf.cell(row=r, column=j, value=cid).font = Font(name=CARLITO, bold=True, color="FFFFFF")
    cf.cell(row=r, column=j).fill = PatternFill("solid", fgColor=NAVY)
cf.cell(row=r, column=1).font = Font(name=CARLITO, bold=True, color="FFFFFF")
cf.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
r += 1
heat_start = r
for cat in rtm_cats:
    cf.cell(row=r, column=1, value=cat).font = Font(name=CARLITO, bold=True)
    counts = Counter(cluster_key(it["cluster"]) for it in by_cat_r[cat])
    for j, cid in enumerate(cluster_ids, start=2):
        cf.cell(row=r, column=j, value=counts.get(cid, 0))
    r += 1
heat_end = r - 1
cf.conditional_formatting.add(f"B{heat_start}:J{heat_end}", _CSR(
    start_type="min", start_color="FFFFFF", end_type="max", end_color="1F6B5C"))
r += 1

# ---- section D: density heat map, OFFER Category x Cluster ----
r = section_title(cf, r, "Density heat map — OFFER Category x Cluster (item counts)")
offer_cats = sorted(by_cat.keys())
cf.cell(row=r, column=1, value="Category")
for j, cid in enumerate(cluster_ids, start=2):
    cf.cell(row=r, column=j, value=cid).font = Font(name=CARLITO, bold=True, color="FFFFFF")
    cf.cell(row=r, column=j).fill = PatternFill("solid", fgColor=NAVY)
cf.cell(row=r, column=1).font = Font(name=CARLITO, bold=True, color="FFFFFF")
cf.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
r += 1
heat_start2 = r
for cat in offer_cats:
    cf.cell(row=r, column=1, value=cat).font = Font(name=CARLITO, bold=True)
    counts = Counter(cluster_key(offer_cluster_map.get(it["id"])) for it in by_cat[cat])
    for j, cid in enumerate(cluster_ids, start=2):
        cf.cell(row=r, column=j, value=counts.get(cid, 0))
    r += 1
heat_end2 = r - 1
cf.conditional_formatting.add(f"B{heat_start2}:J{heat_end2}", _CSR(
    start_type="min", start_color="FFFFFF", end_type="max", end_color="1F6B5C"))
r += 2

cf.cell(row=r, column=1, value=("Method: 'Hidden champion' = the #1-by-Weighted-S item within its Category whose "
    "GLOBAL rank falls in the bottom 60% overall (rank > 40% mark) -- i.e. genuinely easy to miss on a global list. "
    "Density heat maps are plain item counts (darker = more items in that Category x Cluster cell), not S-weighted -- "
    "see the Domains tab in the HTML Navigator for the S-weighted version of this same breakdown."
    )).font = Font(name=CARLITO, italic=True, size=10, color="666666")
cf.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
cf.row_dimensions[r].height = 42

cf.column_dimensions["A"].width = 26
cf.column_dimensions["D"].width = 42
for c in ("B","C","E","F","G","H","I","J"):
    cf.column_dimensions[c].width = 12
print(f"CATEGORY_FOCUS sheet built, {r} rows")

wb.save(OUT)
print(f"saved {OUT}")
