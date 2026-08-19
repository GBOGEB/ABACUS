"""
fix_pdf_page_numbers.py -- corrects RTM_RANKING!J ("PDF page") against the
ACTUAL source PDF (uploads_v5/QPS_Contract_mirror_DOCX.pdf, 137 pages),
rather than trusting whatever process originally populated that column.

Found this round: every one of the 722 RTM IDs is literally printed in the
PDF text (e.g. "RTM-693" appears as a label directly above its shall
statement on page 127), giving a 722/722 authoritative ground truth for
"which page is this RTM actually on." Compared against the workbook's own
pdfPage column: only 247/722 (34%) matched: 475 rows (66%) were off, almost
always by 1-2 pages (avg drift 1.40, max 35), most likely from an earlier
pagination pass (e.g. before/after a cover-page or TOC insertion) that
wasn't re-synced when the PDF was finalised.

This matters beyond just the "PDF page" field being wrong on-screen: GBO
asked this round about generating a verbatim-text/PDF-snippet view per RTM
keyed off this exact field -- building that on the current (66% wrong)
values would silently show the wrong page for most RTMs. Fix the field
first, before anything gets built on top of it.

IN:  QPS_OFFER_Evaluation_FULL_v20.xlsx, uploads_v5/QPS_Contract_mirror_DOCX.pdf
OUT: QPS_OFFER_Evaluation_FULL_v21.xlsx
"""
import re, warnings
warnings.filterwarnings("ignore")
import openpyxl
import fitz  # PyMuPDF

PDF_PATH = "uploads_v5/QPS_Contract_mirror_DOCX.pdf"
IN = "QPS_OFFER_Evaluation_FULL_v20.xlsx"
OUT = "QPS_OFFER_Evaluation_FULL_v21.xlsx"

doc = fitz.open(PDF_PATH)
assert doc.page_count == 137, f"expected 137 pages, got {doc.page_count}"

found_pages = {}
for i in range(doc.page_count):
    text = doc[i].get_text()
    for m in re.finditer(r"RTM-(\d{2,3})\b", text):
        rid = f"RTM-{m.group(1).zfill(3)}"
        if rid not in found_pages:
            found_pages[rid] = i + 1  # 1-indexed page number, first occurrence

wb = openpyxl.load_workbook(IN, data_only=False)
ws = wb["RTM_RANKING"]

fixed = 0
already_correct = 0
no_pdf_hit = 0
diffs = []
for r in range(6, ws.max_row + 1):
    rid = ws.cell(row=r, column=2).value
    if not rid:
        continue
    pdf_page = found_pages.get(rid)
    if pdf_page is None:
        no_pdf_hit += 1
        continue
    old_val = ws.cell(row=r, column=10).value
    try:
        old_n = int(old_val)
    except (TypeError, ValueError):
        old_n = None
    if old_n == pdf_page:
        already_correct += 1
        continue
    ws.cell(row=r, column=10).value = pdf_page
    fixed += 1
    if old_n is not None:
        diffs.append(abs(pdf_page - old_n))

wb.save(OUT)
print(f"wrote {OUT}")
print(f"RTM IDs matched literally in PDF: {len(found_pages)} / 722")
print(f"rows already correct: {already_correct}")
print(f"rows corrected: {fixed}")
print(f"rows with no literal RTM-ID hit in PDF (left unchanged): {no_pdf_hit}")
if diffs:
    print(f"correction size: min={min(diffs)} max={max(diffs)} avg={sum(diffs)/len(diffs):.2f}")
