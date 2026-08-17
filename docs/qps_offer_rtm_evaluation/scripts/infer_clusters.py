"""
infer_clusters.py -- infer a best-guess C1-C8 cluster for the 429 RTMs that
have no direct RTM_CROSSWALK link to an OFFER item (currently shown as
"Not linked to an OFFER item").

Method (disclosed everywhere this is surfaced as "Inferred", never
presented as equivalent to a "Direct"/crosswalk-verified cluster):
  1. For each Domain, compute the cluster distribution among the 293 RTMs
     that DO have a resolvable crosswalk-derived cluster (ground truth).
  2. Most domains have one dominant cluster (e.g. Buildings & Utilities is
     100% C5, Acceptance Testing is 100% C7) -- assign that domain's
     plurality cluster to every unlinked RTM in it, tagged "high" or
     "medium" confidence by vote share.
  3. "Subsystems" (the largest domain, 202 items, only 71 with a direct
     link, and those splitting roughly C2/C3/C7 with no >50% majority) gets
     a keyword tie-break instead of a flat plurality vote, since a flat
     vote there would be barely better than a coin flip.
  4. Every domain with zero linked ground-truth (shouldn't happen given the
     data, but guarded) falls back to "Unclassified" rather than a guess.
"""
import warnings, json, re
warnings.filterwarnings("ignore")
import openpyxl
from collections import Counter, defaultdict

wb = openpyxl.load_workbook("QPS_OFFER_Evaluation_FULL_v6.xlsx", data_only=False)
ws = wb["RTM_RANKING"]
rtm_cluster = json.load(open("/tmp/rtm_cluster.json"))

CLUSTER_NAMES = {
    "C1": "Performance", "C2": "Process Design", "C3": "Mechanical & Equipment",
    "C4": "Software & Control", "C5": "Infrastructure & Integration",
    "C6": "Reliability & Maintenance", "C7": "Quality, Testing & Risk",
    "C8": "Commercial & Execution",
}

def resolve_direct_cluster(rid):
    links = rtm_cluster.get(rid)
    if not links:
        return None
    direct = [l for l in links if str(l[2]).startswith("Direct")]
    pick = direct[0] if direct else links[0]
    distinct = sorted(set((l[0], l[1]) for l in links))
    if len(distinct) == 1:
        return distinct[0][0]
    return pick[0]

# ---- ground truth: domain -> cluster vote distribution --------------------
domain_votes = defaultdict(Counter)
rows = []
for r in range(6, ws.max_row + 1):
    rid = ws.cell(row=r, column=2).value
    if not rid:
        continue
    domain = ws.cell(row=r, column=5).value
    shall = str(ws.cell(row=r, column=13).value or "")
    full = str(ws.cell(row=r, column=14).value or "")
    cid = resolve_direct_cluster(rid)
    rows.append((rid, domain, shall + " " + full, cid))
    if cid:
        domain_votes[domain][cid] += 1

domain_plurality = {}
for dom, votes in domain_votes.items():
    total = sum(votes.values())
    top_cid, top_n = votes.most_common(1)[0]
    share = top_n / total
    conf = "high" if share >= 0.75 else ("medium" if share >= 0.5 else "low")
    domain_plurality[dom] = (top_cid, conf, share)

# ---- Subsystems keyword tie-break -----------------------------------------
def rx(*words):
    return re.compile(r"\b(?:" + "|".join(re.escape(w) for w in words) + r")\b", re.IGNORECASE)

KW_C4 = rx("software", "plc", "control system", "scada", "firmware", "hmi", "cybersecurity")
KW_C3 = rx("compressor", "turbine", "bearing", "valve", "pump", "mechanical", "motor",
           "coupling", "vessel", "piping", "flange", "gasket")
KW_C7 = rx("test", "quality", "inspection", "calibrat", "audit", "verification", "acceptance criteria")
KW_C6 = rx("maintenance", "spare part", "reliab", "mtbf", "mttr", "obsolescence")

def subsystems_tiebreak(text):
    if KW_C4.search(text):
        return "C4", "medium"
    if KW_C3.search(text):
        return "C3", "medium"
    if KW_C7.search(text):
        return "C7", "medium"
    if KW_C6.search(text):
        return "C6", "medium"
    return "C2", "low"  # domain plurality fallback (C2 was the largest single vote)

# ---- assign ----------------------------------------------------------------
inferred = {}
n_direct = n_inferred = n_unclassified = 0
for rid, domain, text, cid in rows:
    if cid:
        n_direct += 1
        continue
    if domain == "Subsystems":
        icid, conf = subsystems_tiebreak(text)
    elif domain in domain_plurality:
        icid, conf, share = domain_plurality[domain]
    else:
        n_unclassified += 1
        continue
    inferred[rid] = {"cluster": icid, "clusterName": CLUSTER_NAMES[icid],
                      "confidence": conf, "method": "inferred"}
    n_inferred += 1

print(f"direct (crosswalk): {n_direct}")
print(f"inferred: {n_inferred}")
print(f"unclassified (no domain signal): {n_unclassified}")

conf_counts = Counter(v["confidence"] for v in inferred.values())
print("confidence distribution:", dict(conf_counts))

cluster_counts = Counter(v["cluster"] for v in inferred.values())
print("inferred cluster distribution:", dict(sorted(cluster_counts.items())))

with open("/tmp/inferred_clusters.json", "w") as f:
    json.dump(inferred, f)
print("saved /tmp/inferred_clusters.json")
