"""
build_workbook_slim_v5.py -- produces the reviewer-shareable "Lite" workbook
from QPS_OFFER_Evaluation_FULL_v5.xlsx, per GBO's request for one slimmed
file "for external or Lite version" rather than two full-detail files.

Kept tabs (the "really need / need to know" set -- matches the curated
9-item nav bar GBO had already built into START_HERE row 3, plus the new
RTM_RANKING tab and the LISTS support sheet dropdowns depend on):
  START_HERE, DASHBOARD, COMPLIANCE_LEGEND, EVALUATION_WORKSPACE,
  NEGOTIATION_AGENDA, OFFER_RANKING, RTM_CROSSWALK, RTM_RANKING,
  QUALITY_CHECKS, LISTS (kept but hidden -- EVALUATION_WORKSPACE's dropdown
  validations point at named ranges defined on this sheet; deleting it would
  break every STATUS/Depth/Negotiation dropdown).

Dropped (still in the FULL workbook): OFFER_CANONICAL, DOMAIN_SUMMARY,
RTM_REVIEW_QUEUE, STANDARDS, DELIVERABLES, AUDIT_NOTES, DMAIC_AUDIT,
EVALUATION_INPUT, NAVIGATION_MAP, CODING_HANDOVER, WEIGHTS_METHOD -- these
are internal/audit/reproduction artefacts (per the engineering handover's
own §4 role column) rather than reviewer-facing content.

Any HYPERLINK("#SheetName!...") cell left pointing at a dropped sheet is
neutralised (link removed, label kept, styled as unavailable) rather than
left as a dead link -- checked across every kept sheet, not just the ones
known to have a nav bar.
"""
import warnings
warnings.filterwarnings("ignore")
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

FULL = "QPS_OFFER_Evaluation_FULL_v5.xlsx"
OUT = "QPS_OFFER_Evaluation_LITE_v5.xlsx"

KEEP = [
    "START_HERE", "DASHBOARD", "COMPLIANCE_LEGEND", "EVALUATION_WORKSPACE",
    "NEGOTIATION_AGENDA", "OFFER_RANKING", "RTM_CROSSWALK", "RTM_RANKING",
    "QUALITY_CHECKS", "LISTS", "OFFER_CANONICAL",
    # OFFER_CANONICAL is reference-only per NAVIGATION_MAP's own role column
    # ("Reference only -- All reviewers"), and QUALITY_CHECKS!C9 depends on
    # its column M (per-item RTM link count) -- dropping it would both
    # remove content the reviewer role calls out as theirs and break a
    # live formula (#REF!). Verified via a full cross-sheet formula scan
    # that no other dropped sheet is a formula dependency of a kept sheet.
]
HYPERLINK_RE = re.compile(r'=HYPERLINK\("#([A-Za-z0-9_]+)!', re.IGNORECASE)
DEAD_FILL = PatternFill("solid", fgColor="E8E8EC")
DEAD_FONT = Font(name="Carlito", size=11, italic=True, color="9A9AA5")

wb = openpyxl.load_workbook(FULL, data_only=False)
dropped = [n for n in wb.sheetnames if n not in KEEP]

for name in dropped:
    del wb[name]
print(f"dropped {len(dropped)} sheets: {dropped}")
print(f"kept {len(wb.sheetnames)} sheets: {wb.sheetnames}")

# LISTS stays for dropdown integrity but is not reviewer-facing content
wb["LISTS"].sheet_state = "hidden"

# Update the workbook title on START_HERE to signal this is the Lite edition
sh = wb["START_HERE"]
if sh["A1"].value and "DMAIC-ready" in str(sh["A1"].value):
    sh["A1"] = str(sh["A1"].value).replace(
        "QPS OFFER evaluation workbook — v5 DMAIC-ready",
        "QPS OFFER evaluation workbook — v5 LITE (reviewer edition)",
    )
sh["A2"] = (
    "Reviewer-shareable subset of the full evaluation workbook -- ranking, "
    "crosswalk, workspace and quality tabs only. Internal audit/handover/"
    "method-control tabs live in the full DMAIC-ready workbook."
)

# ---- neutralise dead cross-sheet hyperlinks left over from dropped tabs --
n_neutralised = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue
            m = HYPERLINK_RE.match(cell.value)
            if not m:
                continue
            target_sheet = m.group(1)
            if target_sheet in wb.sheetnames:
                continue  # still valid, leave untouched
            # extract the human-readable label ("...","Label") to keep as plain text
            label_match = re.search(r',"([^"]*)"\)$', cell.value)
            label = label_match.group(1) if label_match else target_sheet
            cell.value = label
            cell.font = DEAD_FONT
            cell.fill = DEAD_FILL
            n_neutralised += 1
print(f"neutralised {n_neutralised} dead cross-sheet hyperlinks (dropped-tab references)")

# START_HERE's numbered step table (rows 6-14) describes one sheet per row
# across 4 columns (Step/Sheet/Purpose/Action) -- the loop above only
# neutralised the "Sheet" cell (the one holding the HYPERLINK), leaving the
# Purpose/Action prose for DMAIC_AUDIT (row 7, step 7) and CODING_HANDOVER
# (row 8, step 8) describing a tab that's no longer there. Greying those out
# too avoids a half-disabled-looking row.
for row_num in (7, 8):
    for col in ("A", "C", "D"):
        cell = sh[f"{col}{row_num}"]
        if cell.value:
            cell.font = DEAD_FONT

wb.save(OUT)
print(f"saved {OUT}")
