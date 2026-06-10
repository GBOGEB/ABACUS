#!/usr/bin/env python3
"""Build MINERVA_NOMENCLATURE_MASTER.xlsx (Phase 4).

Consolidates: original nomenclature (instruments+equipment), PPT/QSYS
re-allocations, new DIS/LS components, layer-to-component mapping and the
colour / line-weight specification.  Pure openpyxl.
"""
import os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import symbols as SYM
import build_pid_v3 as B

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)
SEG = os.path.join(PROJECT, "segmentation", "data")
OUT = os.path.join(PROJECT, "output_v3")

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(horizontal="center", vertical="center")

ISA_FIRST = SYM.ISA_FIRST
ISA_SUCCEED = SYM.ISA_SUCCEED


def isa_meaning(prefix):
    p = prefix.upper()
    if not p:
        return ""
    first = ISA_FIRST.get(p[0], "?")
    rest = " / ".join(ISA_SUCCEED.get(c, "?") for c in p[1:]) if len(p) > 1 else ""
    return f"{first}" + (f" - {rest}" if rest else "")


def style_header(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w


def fill_rows(ws, start, rows):
    for r, row in enumerate(rows, start):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP


def signal_kind(prefix):
    return B.SIGNAL_KIND.get(prefix.upper(), "-")


def sheet_for(prefix):
    role_valve = prefix.upper() in B.VALVE_PREFIX
    return "Instrumentation (+ Process)" if not role_valve else "Process / Valves"


def main():
    wb = openpyxl.Workbook()

    # ---- 1. Overview ----
    ws = wb.active
    ws.title = "00_Overview"
    ws["A1"] = "MINERVA CryoCell P&ID v3 - Nomenclature Master"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    info = [
        ("Project", "MINERVA CryoCell - SCK CEN (MYRRHA / MINERVA Phase 1)"),
        ("Consultant", "Mott MacDonald"),
        ("Client", "SCK CEN, Boeretang 200, 2400 Mol, Belgium"),
        ("Standards", "ANSI/ISA-5.1-2022; ISO 10628; IEC 60617; SCK CEN AD_01.16"),
        ("Tag scheme", "W-X:Y-Z-1  (Section-Subsection : Discipline-DeviceClass-Index)"),
        ("Drawings", "QCELL (Sheet1 Cryogenic, Sheet2 Instrumentation); "
                     "RFCELL (Sheet1 Process, Sheet2 Instrumentation)"),
        ("Variants", "STANDARD, STANDARD_MONO, CONTROL-CENTRIC, CONTROL-CENTRIC_MONO"),
        ("Worksheets", "01 Instruments | 02 Equipment | 03 Re-allocations | "
                       "04 New components | 05 Layer map | 06 Colour/line-weight | "
                       "07 Scope categories"),
    ]
    r = 3
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    ws.column_dimensions["A"].width = 16
    for col in "BCD":
        ws.column_dimensions[col].width = 34

    # ---- 1+. load segmentation, apply reallocation ----
    segs = {}
    for key, meta in B.SHEETS.items():
        seg = json.load(open(os.path.join(SEG, meta["seg"])))
        B.apply_reallocation(seg)
        segs[key] = seg

    # ---- 2. Instruments ----
    ws = wb.create_sheet("01_Instruments")
    ws["A1"] = "Instrument Nomenclature (original + re-allocated)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    style_header(ws, 3,
                 ["Tag", "Prefix", "Number", "ISA meaning", "Drawing",
                  "Source layer", "Signal type", "Safety?", "Note"],
                 [14, 8, 10, 30, 12, 18, 12, 9, 26])
    rows = []
    for key, seg in segs.items():
        draw = B.SHEETS[key]["sub"]
        for i in seg.get("instruments", []):
            p = (i.get("prefix") or "").upper()
            note = ""
            if i.get("_realloc_from"):
                note = f"re-allocated from {i['_realloc_from']} ({i.get('_sensor_type','')})"
            rows.append([i.get("tag", ""), p, i.get("number", ""), isa_meaning(p),
                         draw, i.get("layer", ""), signal_kind(p),
                         "Y" if i.get("is_safety") else "", note])
    rows.sort(key=lambda x: (x[4], x[1], str(x[2])))
    fill_rows(ws, 4, rows)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{3+len(rows)}"

    # ---- 3. Equipment ----
    ws = wb.create_sheet("02_Equipment")
    ws["A1"] = "Equipment & Terminal Points"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    style_header(ws, 3, ["Drawing", "Kind", "Label", "Source layer", "Glyph"],
                 [12, 26, 22, 20, 16])
    erows = []
    for key, seg in segs.items():
        draw = B.SHEETS[key]["sub"]
        for e in seg.get("equipment", []):
            erows.append([draw, e.get("kind", ""), e.get("label", ""),
                          e.get("layer", ""), B.EQUIP_GLYPH.get(e.get("kind", ""), "node")])
    erows.sort(key=lambda x: (x[0], x[1]))
    fill_rows(ws, 4, erows)
    ws.freeze_panes = "A4"

    # ---- 4. Re-allocations ----
    ws = wb.create_sheet("03_Re-allocations")
    ws["A1"] = "Sensor Re-allocations (QSYS instrumentation study / PPT)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    style_header(ws, 3, ["Original tag", "New tag", "Sensor type", "Position", "Basis"],
                 [14, 14, 14, 34, 30])
    rr = []
    for old, (np_, st, pos) in B.REALLOC.items():
        rr.append([old, np_ + old[2:], st, pos, "Piezo cold/warm extremity mapping"])
    for st, port, note in B.MAG_COUPLER_SENSORS:
        rr.append(["(new)", "-", st, port, note])
    fill_rows(ws, 4, rr)

    # ---- 5. New DIS/LS components ----
    ws = wb.create_sheet("04_New_components")
    ws["A1"] = "New Instrumentation / Control Components (v2+v3 additions)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    style_header(ws, 3, ["Component", "Qty", "Layer", "Description"],
                 [26, 6, 26, 50])
    nc = [
        ["DIS - Device Interlock System", "1", "14_Instruments_Control_DIS",
         "Aggregates Vacuum/Cryo/Utilities OK -> master interlock to RF"],
        ["Tuner limit switch (LS-T1..T3)", "3", "14_Instruments_Control_DIS",
         "Mounted OUTSIDE the vacuum vessel; mechanical link to tuner"],
        ["Lemo B-series connector", "3", "14_Instruments_Control_DIS",
         "Patch-panel HV pins for Piezo (PZ) drive"],
        ["MV bellows (anti thermal short)", "2", "09_Valves_Mechanical",
         "Bellows element on manual-valve (MV) lines"],
        ["Buffer-volume annotation", "1", "17_Notes_TOGGLEABLE",
         "Liquid buffer 7 L min / vapour buffer 5 L min"],
        ["Scope diamonds (TP / cat / next)", "24", "01_Scope_Boundaries",
         "AD_01.16 3-compartment termination points (last-metre hand-over)"],
    ]
    fill_rows(ws, 4, nc)

    # ---- 6. Layer-to-component mapping ----
    ws = wb.create_sheet("05_Layer_map")
    ws["A1"] = "Hierarchical Layer -> Component Mapping"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    style_header(ws, 3, ["Layer name", "Components", "Colour", "Line weight"],
                 [34, 40, 16, 16])
    lm = [
        ("00_Background_TitleBlock", "Frame, header, compact bottom title block", "Black", "0.7-2.0 mm"),
        ("01_Scope_Boundaries", "TP scope diamonds (AD_01.16)", "Per category", "1.2 mm"),
        ("02_Structure_Reference", "Faded reference geometry", "Grey", "0.5 mm"),
        ("03_Equipment_Vessels", "Cavities, couplers, HX, vessels, tuners, antennae", "Black/brown", "0.5 mm"),
        ("04A_Piping_PRIMARY_40K", "40 K shield trunk lines", "Red #e00000", "1.0 mm"),
        ("04B_Piping_BRANCHES_40K", "40 K shield branch lines", "Red #e00000", "0.7 mm"),
        ("05A_Piping_PRIMARY_4p5K", "4.5 K supply trunk lines", "Blue #0033cc", "1.0 mm"),
        ("05B_Piping_BRANCHES_4p5K", "4.5 K supply branch lines", "Blue #0033cc", "0.7 mm"),
        ("06A_Piping_PRIMARY_2K", "2 K return trunk lines", "Cyan #00a6bd", "1.0 mm"),
        ("06B_Piping_BRANCHES_2K", "2 K return branch lines", "Cyan #00a6bd", "0.7 mm"),
        ("07_Piping_SECONDARY_Water", "DI cooling-water secondary circuit", "Green #00a000", "0.5 mm"),
        ("08_Piping_OUTSIDE_SCOPE", "Guard/infra/instrument-air services", "Grey dashed", "0.35 mm"),
        ("09_Valves_Mechanical", "CV/HV/SV/RV/MV/PL valves + bellows", "Black", "1.0 mm"),
        ("10_Signals_Pneumatic", "Pneumatic signals (dash + cross-tick)", "Purple #7a00a0", "0.25 mm"),
        ("11_Signals_Electric", "Electric signals (dotted)", "Blue #00529b", "0.25 mm"),
        ("12_Signals_Hydraulic", "Hydraulic signals (dash-dot)", "Amber #a06a00", "0.25 mm"),
        ("13_Instruments_Sensors", "ISA instrument bubbles + heat loads", "Per family", "0.3 mm"),
        ("14_Instruments_Control_DIS", "DIS, tuner LS, Lemo connectors", "Black/red", "0.3-1.6 mm"),
        ("15_Tags_Instruments", "Tag text (>=2.5 mm main / 2.0 mm bubble)", "Black", "n/a"),
        ("16_Legend_TOGGLEABLE", "Compact legend overlay (off by default)", "Mixed", "n/a"),
        ("17_Notes_TOGGLEABLE", "Buffer/scope notes, callouts", "Green/red", "n/a"),
    ]
    fill_rows(ws, 4, lm)

    # ---- 7. Colour / line-weight spec ----
    ws = wb.create_sheet("06_Colour_LineWeight")
    ws["A1"] = "Colour & Line-Weight Specification"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    style_header(ws, 3,
                 ["Class / item", "Hex", "Primary (mm)", "Branch (mm)",
                  "Mono style", "Notes"],
                 [22, 12, 12, 12, 22, 30])
    cw = [
        ["40 K shield (D)", "#e00000", "1.0", "0.7", "black solid", "Red per source legend"],
        ["4.5 K supply (A)", "#0033cc", "1.0", "0.7", "black solid", "Blue"],
        ["2 K return (B)", "#00a6bd", "1.0", "0.7", "black solid", "Cyan"],
        ["DI water (secondary)", "#00a000", "0.5", "-", "black solid", "Green"],
        ["60 K guard / infra / air", "grey", "0.35", "-", "black dashed", "Out of scope"],
        ["Pneumatic signal", "#7a00a0", "0.25", "-", "dash + cross-tick", "AD_01.16"],
        ["Electric signal", "#00529b", "0.25", "-", "dotted", "AD_01.16 (electrical)"],
        ["Hydraulic signal", "#a06a00", "0.25", "-", "dash-dot", "AD_01.16 (L mark)"],
        ["Instrument bubble LB", "#ffffff", "-", "-", "white fill", "Cryo field instrument"],
        ["Instrument bubble RF", "#ffe2e2", "-", "-", "white fill", "RFCELL instrument"],
        ["Instrument bubble LBI", "#dbe9ff", "-", "-", "white fill", "LBI-specific"],
        ["Safety / SIS bubble", "dashed", "-", "-", "dashed outline", "Interlock instruments"],
    ]
    fill_rows(ws, 4, cw)
    ws["A17"] = ("Text minimums @ A3:  main tags >= 2.5 mm,  bubble text 2.0 mm,  "
                 "line callouts 2.2 mm,  legend 1.8 mm.")
    ws["A17"].font = Font(italic=True)
    ws.merge_cells("A17:F17")

    # ---- 8. Scope categories ----
    ws = wb.create_sheet("07_Scope_categories")
    ws["A1"] = "Termination-Point Scope Categories (SCK CEN AD_01.16)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    style_header(ws, 3, ["Category letter", "Meaning", "Diamond format"], [16, 26, 46])
    sc = []
    for k, (name, col) in SYM.SCOPE_CATEGORY.items():
        sc.append([k, name, "TP / <letter><unique no.> / <next system>"])
    fill_rows(ws, 4, sc)
    ws["A14"] = ("Diamond = 3 stacked compartments: top 'TP', middle category+number, "
                 "bottom next system/process (ZZZ). Marks last-metre in/out-of-scope hand-over.")
    ws["A14"].font = Font(italic=True)
    ws.merge_cells("A14:C14")

    out_path = os.path.join(OUT, "MINERVA_NOMENCLATURE_MASTER.xlsx")
    wb.save(out_path)
    print("wrote", out_path)
    print("instruments:", len(rows), "equipment:", len(erows))


if __name__ == "__main__":
    main()
