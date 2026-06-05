"""Phase 1 / Phase 6 - LINE_SPECIFICATION_MASTER.xlsx

Comprehensive line database for the revised MINERVA CryoCell P&ID (v5),
built from line_spec_data.LINES plus the extracted source nomenclature.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import line_spec_data as L

OUT = os.path.join(os.path.dirname(__file__), "..", "output_v5")

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
GRP_FILL = {"cold": "DCE6F1", "thermal": "FCE4D6", "warm": "E2EFDA", "scope": "EDEDED"}
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hex(c):
    return c.replace("#", "").upper()


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def sheet_lines(wb):
    ws = wb.active
    ws.title = "01_Line_Specification"
    ws["A1"] = "MINERVA CryoCell - LINE SPECIFICATION MASTER (v5)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = "Project NA.PS01 - QCELL/RFCELL - revised cryogenic colour scheme & nomenclature"
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws.merge_cells("A2:I2")

    cols = ["Line", "Group", "Colour (hex)", "Temp", "Pressure", "Flow",
            "Size (DN)", "MOC", "Description"]
    hr = 4
    for i, h in enumerate(cols, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header(ws, hr, len(cols))

    r = hr + 1
    for row in L.spec_rows():
        ws.cell(row=r, column=1, value=row["Line"]).font = Font(bold=True, size=11)
        ws.cell(row=r, column=2, value=row["name"])
        ccell = ws.cell(row=r, column=3, value=row["Colour"])
        ccell.fill = PatternFill("solid", fgColor=_hex(row["Colour"]))
        # white text on dark fills
        if row["Colour"].lower() in ("#0000ff", "#000080", "#ff0000", "#cc0000",
                                     "#808000", "#808080", "#008b8b"):
            ccell.font = Font(color="FFFFFF", bold=True)
        else:
            ccell.font = Font(bold=True)
        ccell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=row["Temp"])
        ws.cell(row=r, column=5, value=row["Pressure"])
        ws.cell(row=r, column=6, value=row["Flow"])
        ws.cell(row=r, column=7, value=row["Size (DN)"])
        ws.cell(row=r, column=8, value=row["MOC"])
        ws.cell(row=r, column=9, value=row["Description"])
        fill = GRP_FILL.get(row["group"])
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c != 3 and fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 9),
                                       horizontal=cell.alignment.horizontal or "left")
        r += 1

    widths = [8, 26, 13, 12, 12, 12, 10, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{hr}:I{r-1}"


def sheet_branch_logic(wb):
    ws = wb.create_sheet("02_Main_vs_Branch")
    ws["A1"] = "MAIN vs BRANCH hierarchy"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    cols = ["Main line", "Branch", "Main size", "Branch size", "Note"]
    for i, h in enumerate(cols, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(cols))
    pairs = [("A", "A'", "DN50", "DN25", "4.5 K supply taps to cavities/HX"),
             ("B", "B'", "DN40", "DN25", "2 K return collectors"),
             ("D", "D'", "DN32", "DN20", "40 K shield distribution"),
             ("E", "E'", "DN32", "DN20", "60 K shield collection")]
    r = 4
    for m, b, ms, bs, note in pairs:
        for i, v in enumerate([m, b, ms, bs, note], 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = BORDER
        r += 1
    for i, w in enumerate([12, 12, 12, 12, 48], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_source(wb):
    """Equipment/circuit references drawn from the extracted nomenclature xlsx."""
    ws = wb.create_sheet("03_Source_Circuits")
    ws["A1"] = "Source nomenclature - circuit references (extracted)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    cols = ["Circuit", "Maps to line", "Example components", "Source sheet"]
    for i, h in enumerate(cols, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(cols))
    data = [
        ("40 K circuit", "D / E (shield in/out)", "CV100, CV101, FV100, PT100, TT100-103", "valve box-jumper"),
        ("4.5 K circuit", "A", "CV200-202, FV200, PT200, TT200-203", "valve box-jumper"),
        ("2 K circuit", "B", "CV300-302, FV300, TT300-310, SV300", "cryomodule"),
        ("Cooling water", "W (warm)", "CV003/004, FT001/002, PT005, TT003-006", "cryomodule"),
        ("GHe supply (U)", "U", "WPS supply inlet (NA.CP03)", "Auxiliary lines"),
        ("Warm return (W)", "W", "WPS / CPLR return (NA.CP03)", "Auxiliary lines"),
        ("Safety line (S)", "S", "WPS safety valve relief (NA.CP03)", "Auxiliary lines"),
    ]
    r = 4
    for row in data:
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
    for i, w in enumerate([18, 22, 42, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.Workbook()
    sheet_lines(wb)
    sheet_branch_logic(wb)
    sheet_source(wb)
    path = os.path.join(OUT, "LINE_SPECIFICATION_MASTER.xlsx")
    wb.save(path)
    print("wrote", os.path.relpath(path, os.path.join(os.path.dirname(__file__), "..")))


if __name__ == "__main__":
    main()
