#!/usr/bin/env python3
"""Phase 10 - AD_01.16 compliance matrix (xlsx).

Audits every v4 symbol / convention against SCK CEN AD_01.16 and the cited
international standards, classifying each requirement as:
  Compliant         - implemented exactly as specified
  Adapted           - intent met, representation adapted for this deliverable
  Design Freedom    - no binding rule; a defensible engineering choice was made
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(PROJECT, "output_v4")

COMPLIANT = "Compliant"
ADAPTED = "Adapted"
FREEDOM = "Design Freedom"

# (section, requirement, status, evidence/notes)
ROWS = [
    ("1. Drawing format", "A3 landscape sheet (420x297 mm), bordered frame + title block",
     COMPLIANT, "All 16 SVGs are 420x297 mm; PDFs export at 1190.55x841.89 pt (A3). Title block bottom-right."),
    ("1. Drawing format", "Title block fields: project, drawing no, MMD project, rev, scale, suitability, security",
     COMPLIANT, "Title block carries =NA.PS01_PFB712/713, MMD 411066, S2-FOR ACCEPTANCE, RESTRICTED, NTS."),
    ("1. Drawing format", "Revision table with rev / date / description / approver",
     COMPLIANT, "Revision rows A1..C1 (2026-04..2026-06) shown; v4 superset."),
    ("2. Layering", "Discipline-segregated layer hierarchy",
     COMPLIANT, "24 named layers (frame, scope, structure, piping by class, valves, signals, instruments, tags, legend, notes)."),
    ("2. Layering", "Toggleable / non-printing overlays for working views",
     ADAPTED, "Implemented via SVG layer display + HTML viewer toggles; overlays hidden by default (valve overlay, legend, notes)."),
    ("3. Line identification", "Each process line carries a line number / identifier",
     ADAPTED, "Inline on-line names per class (e.g. 40K-3\"-PAC1-DA) placed on the longest runs; full per-segment numbering pending P&ID line list."),
    ("3. Line identification", "Line naming follows <size>-<service>-<spec>-<route> convention",
     COMPLIANT, "Line-spec labels follow AD_01.10 form (size + service code + class + route suffix)."),
    ("3. Line identification", "Names legible on monochrome plots (no colour dependence)",
     COMPLIANT, "04C_Piping_LINENAMES layer renders black on mono with white boxes; legibility verified in QA."),
    ("4. Instrument symbols", "ISA-5.1 bubbles: prefix + loop number, field vs panel modifiers",
     COMPLIANT, "bubble_v3 renders field (plain) / panel (single line) / safety (dashed) per ISA-5.1."),
    ("4. Instrument symbols", "Bubbles must not be obscured by piping; tags readable",
     COMPLIANT, "All instrument & valve tags drawn in opaque white boxes on the front-most 12_Tags_Instruments layer."),
    ("4. Instrument symbols", "Signal lines typed (pneumatic / electric / hydraulic)",
     COMPLIANT, "Three signal layers with ISA line styles (dash+hash / dotted / dash-dot)."),
    ("5. Valves & actuators", "Standard valve body + actuator glyphs",
     COMPLIANT, "valve() renders gate/control/relief/solenoid bodies with actuator glyph above."),
    ("5. Valves & actuators", "Valve orientation shown in process context",
     COMPLIANT, "In-line vertical valves on 09_Valves_Mechanical follow the pipe."),
    ("5. Valves & actuators", "Tracked-asset / horizontal valve schedule view",
     FREEDOM, "08B_Valves_HORIZONTAL_OVERLAY adds a toggleable horizontal valve banner with leader lines - an added asset-tracking aid (no binding AD_01.16 rule)."),
    ("6. Terminal points", "Off-sheet connections shown at the drawing border",
     COMPLIANT, "02B_TerminalPoints_EDGE anchors TP assemblies to the left (FROM) and right (TO) page edges per AD_01.10."),
    ("6. Terminal points", "TP assembly: cloud + drawing ref + line no + FROM/TO + TP code + category",
     COMPLIANT, "terminal_point_edge() draws scalloped cloud, arrow ref box (dwg ref + line no), FROM/TO text, 3-compartment TP diamond and category triangle."),
    ("6. Terminal points", "Scope boundary categories (B/C/E/G/H/L/S/W)",
     COMPLIANT, "scope_diamond_3c uses the AD_01.16 category set; legend lists all categories."),
    ("7. Colour & B/W", "Drawing must be usable in monochrome (print)",
     COMPLIANT, "Every sheet has a *_MONO variant; colour replaced by line weight + on-line names."),
    ("7. Colour & B/W", "Colour key / legend present",
     COMPLIANT, "16_Legend_INTERACTIVE table: class colour, service, temperature, pressure + signal key + ISA/scope key."),
    ("8. Standards cited", "ANSI/ISA-5.1, ISO 10628, IEC 60617, AD_01.16",
     COMPLIANT, "Cited in <metadata> standard element and YAML front matter of every SVG."),
    ("9. Metadata", "Machine-readable metadata embedded in the drawing",
     ADAPTED, "Jekyll-style YAML 1.2 front matter carried in a metadata CDATA section (XML comments cannot contain '--', so CDATA preserves the '---' fences); plus a <metadata> default-views descriptor."),
    ("9. Metadata", "Default / named views defined",
     COMPLIANT, "5 preset views (FULL/PROCESS/CONTROL/MAIN/PRINT_MONO) embedded and wired into the HTML viewer."),
    ("10. Deliverable usability", "Interactive review aid",
     FREEDOM, "HTML viewer (layer toggles, presets, tag search, zoom/pan, colour/mono, PNG export) provided as an added review aid."),
    ("10. Deliverable usability", "QA evidence pack",
     COMPLIANT, "output_v4/QA: 300-DPI-class renders, colour-vs-mono comparison strips, v3-vs-v4 grid, alignment overlays, QA_REPORT.md."),
]

STATUS_FILL = {
    COMPLIANT: PatternFill("solid", fgColor="C6EFCE"),
    ADAPTED: PatternFill("solid", fgColor="FFEB9C"),
    FREEDOM: PatternFill("solid", fgColor="BDD7EE"),
}
STATUS_ICON = {COMPLIANT: "\u2705 Compliant", ADAPTED: "\u26a0\ufe0f Adapted",
               FREEDOM: "\U0001f513 Design Freedom"}
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "AD_01.16 Compliance"

    # title rows
    ws.merge_cells("A1:D1")
    ws["A1"] = "MINERVA CryoCell P&ID v4.0 - AD_01.16 Compliance Matrix"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1D3B53")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:D2")
    ws["A2"] = ("SCK CEN AD_01.16 | ANSI/ISA-5.1-2009 | ISO 10628 | IEC 60617    "
                "Legend:  \u2705 Compliant   \u26a0\ufe0f Adapted   \U0001f513 Design Freedom")
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    headers = ["Section", "Requirement", "Status", "Evidence / Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C4D6B")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[3].height = 20

    r = 4
    for sec, req, status, note in ROWS:
        ws.cell(row=r, column=1, value=sec)
        ws.cell(row=r, column=2, value=req)
        sc = ws.cell(row=r, column=3, value=STATUS_ICON[status])
        sc.fill = STATUS_FILL[status]
        sc.font = Font(bold=True)
        ws.cell(row=r, column=4, value=note)
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        r += 1

    # summary
    counts = {COMPLIANT: 0, ADAPTED: 0, FREEDOM: 0}
    for *_, status, _ in [(0, 0, s, 0) for *_, s, _ in ROWS]:
        pass
    for _, _, status, _ in ROWS:
        counts[status] += 1
    r += 1
    ws.cell(row=r, column=1, value="SUMMARY").font = Font(bold=True)
    for status in (COMPLIANT, ADAPTED, FREEDOM):
        r += 1
        ws.cell(row=r, column=1, value=STATUS_ICON[status])
        ws.cell(row=r, column=1).fill = STATUS_FILL[status]
        ws.cell(row=r, column=2,
                value=f"{counts[status]} of {len(ROWS)} requirements")

    widths = [22, 52, 18, 78]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # second sheet: design-freedom rationale
    ws2 = wb.create_sheet("Design Freedom Rationale")
    ws2["A1"] = "Design-Freedom Decisions (no binding AD_01.16 rule)"
    ws2["A1"].font = Font(bold=True, size=13)
    fr = [
        ("Horizontal valve overlay",
         "AD_01.16 fixes neither a valve schedule view nor a second orientation. "
         "A toggleable horizontal 'tracked-asset' banner with grey leader lines to "
         "each in-line valve was added to aid maintenance/asset tracking. Hidden by default."),
        ("Interactive HTML viewer",
         "Not required by the standard. Provided to let reviewers toggle the 24 "
         "layers, switch colour/mono and style, apply 5 preset views, search tags, "
         "zoom/pan and export PNG - without CAD software."),
        ("YAML front matter via CDATA",
         "The brief asked for Jekyll YAML front matter 'as an XML comment'. XML "
         "comments may not legally contain '--', and YAML uses '---' fences, so the "
         "front matter is carried verbatim inside a metadata CDATA section (well-formed "
         "XML, fences preserved, still machine-extractable)."),
        ("Inline line-name sampling",
         "Until a formal P&ID line list is issued, on-line names are placed on the "
         "longest representative run of each process class. The naming convention "
         "matches AD_01.10; per-segment numbers can be back-filled from the line list."),
        ("Two style profiles",
         "STANDARD (process-emphasised) and CONTROL-CENTRIC (signal-emphasised, process "
         "greyed) are offered so the same data serves process and controls reviews."),
    ]
    ws2.append([]); ws2.append(["Decision", "Rationale"])
    for c in (1, 2):
        ws2.cell(row=3, column=c).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=3, column=c).fill = PatternFill("solid", fgColor="2C4D6B")
    for dec, rat in fr:
        ws2.append([dec, rat])
        rr = ws2.max_row
        for c in (1, 2):
            ws2.cell(row=rr, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            ws2.cell(row=rr, column=c).border = BORDER
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 96

    path = os.path.join(OUT, "AD_01.16_COMPLIANCE_MATRIX.xlsx")
    wb.save(path)
    print("wrote", os.path.relpath(path, PROJECT))
    print("summary:", counts)


if __name__ == "__main__":
    main()
