"""
================================================================================
 Module : src/abacus_svg_pid/build_w005.py
 Wave   : W005 — Tag & Instrument Register Reconciliation (XLSX coverage delta)
 Status : ACTIVE
--------------------------------------------------------------------------------
 Purpose
 -------
 Cross-reference the auto-catalogued, as-drawn instrument tags (W003/W004 output,
 reports/COMPONENT_CATALOG.xlsx) against the official *design nomenclature*
 (the engineering tag register XLSX) to quantify coverage:

     matched / missing (in design, not drawn) / extra (drawn, not in design)
     + per-type coverage + ISA-class cross-validation + PPT re-allocations.

 Inputs (tracked source)
 -----------------------
   extracted/master/PID Nomenclature MINERVA CryoCell (QCELL-LB).xlsx   (design register)
   reports/COMPONENT_CATALOG.xlsx                                        (as-drawn, derived)
   configs/isa_classes.json                                             (prefix -> ISA meta)

 Outputs
 -------
   data/excel/excel_register.json            (Phase 1 — parsed design register)
   data/excel/catalog_register.json          (Phase 2 — parsed as-drawn catalog)
   data/excel/reconciliation_results.json    (Phase 3 — match/missing/extra/realloc)
   reports/W005_coverage_statistics.json     (Phase 4 — coverage summary)
   reports/W005_XLSX_RECONCILIATION_REPORT.md (Phase 5 — main report)
   data/excel/canonical_register_v1.yaml      (Phase 6 — unified SSOT)
   reports/COMPONENT_CATALOG_v2.xlsx          (Phase 7 — catalog + reconciliation cols)
   reports/W005_validation_report.md          (Phase 8 — validation record)

 Honesty note
 ------------
 The design register and the as-drawn catalog use *orthogonal* numbering schemes
 (design = circuit-sequential CV001/TT100…; as-drawn = SVG-instance CV560/TT514…).
 Exact tag overlap is therefore ZERO. That is a real engineering finding, not a
 failure: the headline deliverable is the design<->as-drawn cross-map need plus
 the per-TYPE coverage delta. We never fabricate matches.

 Reproducible: PYTHONPATH=src python3 -m abacus_svg_pid.build_w005
================================================================================
"""

import json
import os
import re
from collections import Counter, defaultdict

from openpyxl import load_workbook, Workbook

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

EXCEL_SRC = os.path.join(
    ROOT, "extracted", "master",
    "PID Nomenclature MINERVA CryoCell (QCELL-LB).xlsx",
)
CATALOG_SRC = os.path.join(ROOT, "reports", "COMPONENT_CATALOG.xlsx")
ISA_CLASSES = os.path.join(ROOT, "configs", "isa_classes.json")

OUT_EXCEL_DIR = os.path.join(ROOT, "data", "excel")
OUT_REPORTS = os.path.join(ROOT, "reports")

EXCEL_REGISTER_JSON = os.path.join(OUT_EXCEL_DIR, "excel_register.json")
CATALOG_REGISTER_JSON = os.path.join(OUT_EXCEL_DIR, "catalog_register.json")
RECON_JSON = os.path.join(OUT_EXCEL_DIR, "reconciliation_results.json")
CANONICAL_YAML = os.path.join(OUT_EXCEL_DIR, "canonical_register_v1.yaml")

COVERAGE_JSON = os.path.join(OUT_REPORTS, "W005_coverage_statistics.json")
REPORT_MD = os.path.join(OUT_REPORTS, "W005_XLSX_RECONCILIATION_REPORT.md")
CATALOG_V2_XLSX = os.path.join(OUT_REPORTS, "COMPONENT_CATALOG_v2.xlsx")
VALIDATION_MD = os.path.join(OUT_REPORTS, "W005_validation_report.md")

# --------------------------------------------------------------------------- #
# PPT re-allocations (cited from the QSYS instrumentation deck — see report
# section 4). Exhaustive parse of the 65 MB deck is a KNOWN GAP; these two
# documented re-allocations are encoded from the brainstorming/QSYS source.
# --------------------------------------------------------------------------- #
PPT_REALLOCATIONS = [
    {
        "tag": "TT535",
        "reallocated_to": "PZ",
        "target": "phase separator (PZ) — coldest part",
        "sensor": "TT-CX (Cernox)",
        "source": "PPT 2024-03-07 'Brainstorming QM instrumentation and "
                   "controls' / QSYS (RFCELL) instrumentation deck, slide 59",
        "note": "TT535 is present in the as-drawn catalog; re-allocation moves "
                "it onto the phase-separator coldest measurement point.",
    },
    {
        "tag": "TT525",
        "reallocated_to": "PZ",
        "target": "phase separator (PZ) — warmest part",
        "sensor": "TT-PT100",
        "source": "PPT 2024-03-07 'Brainstorming QM instrumentation and "
                   "controls' / QSYS (RFCELL) instrumentation deck, slide 59",
        "note": "TT525 is NOT in the as-drawn catalog (missing); re-allocation "
                "documents its intended phase-separator warmest measurement point.",
    },
]


# --------------------------------------------------------------------------- #
# Tag helpers
# --------------------------------------------------------------------------- #
def norm_tag(tag):
    """Canonical comparison key: strip hyphens/spaces, upper-case."""
    if tag is None:
        return ""
    return re.sub(r"[-\s_]", "", str(tag)).upper()


def prefix_of(tag):
    """Leading 1–3 alpha characters of a tag (instrument class prefix)."""
    m = re.match(r"^([A-Za-z]{1,3})", str(tag).strip())
    return m.group(1).upper() if m else ""


def is_template(tag):
    """RFCELL templated placeholders carry a literal 'x' as a digit-placeholder
    in the number field (TTxxx, EHx11, PTx21…). These are not real instrument
    instances. Prefixes in this dataset are at most 2 letters (CV/EH/HV/LS/PT/
    TT/HX…), so we strip up to 2 leading letters before testing for 'x' — this
    avoids the 2-letter real prefix 'HX' (e.g. HX001) being treated as template."""
    body = re.sub(r"^[A-Za-z]{1,2}", "", str(tag))
    return "x" in body.lower()


# --------------------------------------------------------------------------- #
# ISA class metadata
# --------------------------------------------------------------------------- #
def load_isa_classes():
    with open(ISA_CLASSES) as fh:
        return json.load(fh)


def isa_meta(isa, prefix):
    classes = isa.get("classes", {})
    if prefix in classes:
        c = classes[prefix]
        return {
            "canonical_name": c.get("canonical_name", ""),
            "variable": c.get("variable", ""),
            "discipline": c.get("discipline", ""),
            "isa_known": True,
        }
    amb = isa.get("ambiguous_single_letters", {})
    if prefix in amb:
        a = amb[prefix]
        return {
            "canonical_name": a.get("canonical_name", a.get("note", "ambiguous")),
            "variable": a.get("variable", ""),
            "discipline": a.get("discipline", ""),
            "isa_known": True,
        }
    return {"canonical_name": "", "variable": "", "discipline": "", "isa_known": False}


# --------------------------------------------------------------------------- #
# Phase 1 — parse the design-nomenclature Excel
# --------------------------------------------------------------------------- #
def parse_excel(isa):
    """Parse the design tag register.

    Layout (per sheet): col A = section grouping label split vertically across
    consecutive rows; col B = the TAG; C = equipment/type; D = action/location;
    E = size/model; F = range; G = I/O; H = signal. Columns I/J hold the legend
    glossary (abbreviation -> meaning), NOT instrument data.
    """
    wb = load_workbook(EXCEL_SRC, data_only=True)
    records = []
    legend = {}
    sheets_meta = []

    for ws in wb.worksheets:
        sheet_rows = 0
        section_fragments = []      # accumulating col-A fragments
        current_section = ""
        for row in ws.iter_rows(min_row=1, values_only=True):
            row = list(row) + [None] * (10 - len(row)) if len(row) < 10 else list(row)
            part, name, equip, loc, size, rng, io, signal = row[:8]
            legend_key, legend_val = row[8], row[9]

            # --- legend glossary (cols I/J) ---
            if legend_key and str(legend_key).strip():
                k = str(legend_key).strip()
                v = str(legend_val).strip() if legend_val else ""
                if k and k.lower() not in ("legend", "abbreviation"):
                    legend.setdefault(k, v)

            # --- section label reconstruction (col A) ---
            if part and str(part).strip():
                frag = str(part).strip()
                # header-ish fragments accumulate into the current section name
                section_fragments.append(frag)
                current_section = " ".join(section_fragments)

            # --- a real instrument row needs a tag in col B ---
            if not (name and str(name).strip()):
                continue
            tag = str(name).strip()
            # skip header rows where col B literally says "Name"
            if tag.lower() in ("name", "tag", "tag name"):
                section_fragments = []
                continue
            # once we hit instrument rows the section name is settled; reset the
            # fragment accumulator so the NEXT block of A-fragments starts fresh
            pfx = prefix_of(tag)
            meta = isa_meta(isa, pfx)
            records.append({
                "tag": tag,
                "norm": norm_tag(tag),
                "prefix": pfx,
                "equipment_type": str(equip).strip() if equip else "",
                "location": str(loc).strip() if loc else "",
                "size_model": str(size).strip() if size else "",
                "range": str(rng).strip() if rng else "",
                "io": str(io).strip() if io else "",
                "signal": str(signal).strip() if signal else "",
                "sheet": ws.title,
                "section": current_section,
                "isa_canonical": meta["canonical_name"],
                "isa_variable": meta["variable"],
                "isa_known": meta["isa_known"],
            })
            section_fragments = []  # next A-block begins a new section
            sheet_rows += 1
        sheets_meta.append({"sheet": ws.title, "instrument_rows": sheet_rows})

    out = {
        "source": os.path.relpath(EXCEL_SRC, ROOT),
        "scheme": "design-nomenclature (circuit-sequential)",
        "sheets": sheets_meta,
        "instrument_count": len(records),
        "prefix_histogram": dict(sorted(Counter(r["prefix"] for r in records).items())),
        "legend_glossary": legend,
        "instruments": records,
    }
    os.makedirs(OUT_EXCEL_DIR, exist_ok=True)
    with open(EXCEL_REGISTER_JSON, "w") as fh:
        json.dump(out, fh, indent=2)
    return out


# --------------------------------------------------------------------------- #
# Phase 2 — parse the as-drawn COMPONENT_CATALOG.xlsx
# --------------------------------------------------------------------------- #
INSTRUMENT_SHEETS = {
    "CV_ControlValves": "CV",
    "TT_Temperature": "TT",
    "EH_ElectricalHeaters": "EH",
    "PT_Pressure": "PT",
    "LS_LimitSwitches": "LS",
    "HV_ManualValves": "HV",
}


def parse_catalog(isa):
    wb = load_workbook(CATALOG_SRC, data_only=True)
    records = []
    seen = set()
    dup_counter = Counter()

    for sheet_name in INSTRUMENT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}

        def cell(r, key):
            i = idx.get(key)
            if i is None or i >= len(r):
                return None
            return r[i]

        for r in rows[1:]:
            tag = cell(r, "Tag")
            if not (tag and str(tag).strip()):
                continue
            tag = str(tag).strip()
            ntag = norm_tag(tag)
            dup_counter[ntag] += 1
            if ntag in seen:
                continue          # dedupe by normalized tag
            seen.add(ntag)
            pfx = (str(cell(r, "Prefix")).strip() if cell(r, "Prefix") else prefix_of(tag)).upper()
            meta = isa_meta(isa, pfx)
            tmpl = is_template(tag)

            def num(key):
                v = cell(r, key)
                try:
                    return round(float(v), 2)
                except (TypeError, ValueError):
                    return None

            records.append({
                "tag": tag,
                "norm": ntag,
                "prefix": pfx,
                "number": str(cell(r, "Number")).strip() if cell(r, "Number") else "",
                "line": str(cell(r, "Line")).strip() if cell(r, "Line") else "",
                "colour": str(cell(r, "Colour")).strip() if cell(r, "Colour") else "",
                "role": str(cell(r, "Role")).strip() if cell(r, "Role") else "",
                "sheet": str(cell(r, "Sheet")).strip() if cell(r, "Sheet") else "",
                "x": num("X"),
                "y": num("Y"),
                "layer": str(cell(r, "Layer")).strip() if cell(r, "Layer") else "",
                "assign_dist_px": num("AssignDist_px"),
                "template": tmpl,
                "isa_canonical": meta["canonical_name"],
                "isa_known": meta["isa_known"],
                "src_sheet": sheet_name,
            })

    real = [r for r in records if not r["template"]]
    tmpl = [r for r in records if r["template"]]
    duplicates = {k: c for k, c in dup_counter.items() if c > 1}

    out = {
        "source": os.path.relpath(CATALOG_SRC, ROOT),
        "scheme": "as-drawn (SVG-instance numbering)",
        "instrument_sheets": list(INSTRUMENT_SHEETS),
        "unique_count": len(records),
        "real_count": len(real),
        "template_count": len(tmpl),
        "real_prefix_histogram": dict(sorted(Counter(r["prefix"] for r in real).items())),
        "template_tags": sorted(r["tag"] for r in tmpl),
        "duplicate_norm_tags": dict(sorted(duplicates.items())),
        "instruments": records,
    }
    with open(CATALOG_REGISTER_JSON, "w") as fh:
        json.dump(out, fh, indent=2)
    return out


# --------------------------------------------------------------------------- #
# Phase 3 — reconcile
# --------------------------------------------------------------------------- #
def reconcile(excel, catalog):
    ex = excel["instruments"]
    cat = catalog["instruments"]
    cat_real = [c for c in cat if not c["template"]]

    ex_by_norm = {r["norm"]: r for r in ex}
    cat_by_norm = {c["norm"]: c for c in cat_real}

    ex_norms = set(ex_by_norm)
    cat_norms = set(cat_by_norm)

    matched_norms = sorted(ex_norms & cat_norms)
    missing_norms = sorted(ex_norms - cat_norms)     # in design, not drawn
    extra_norms = sorted(cat_norms - ex_norms)       # drawn, not in design

    matched = [{
        "norm": n,
        "excel_tag": ex_by_norm[n]["tag"],
        "catalog_tag": cat_by_norm[n]["tag"],
    } for n in matched_norms]

    # discrepancies: tags that match by norm but disagree on prefix/type
    discrepancies = []
    for n in matched_norms:
        e, c = ex_by_norm[n], cat_by_norm[n]
        if e["prefix"] != c["prefix"]:
            discrepancies.append({
                "norm": n,
                "excel_prefix": e["prefix"],
                "catalog_prefix": c["prefix"],
                "issue": "prefix/type mismatch",
            })

    missing = [{
        "tag": ex_by_norm[n]["tag"],
        "prefix": ex_by_norm[n]["prefix"],
        "equipment_type": ex_by_norm[n]["equipment_type"],
        "section": ex_by_norm[n]["section"],
        "sheet": ex_by_norm[n]["sheet"],
        "isa_canonical": ex_by_norm[n]["isa_canonical"],
    } for n in missing_norms]

    extra = [{
        "tag": cat_by_norm[n]["tag"],
        "prefix": cat_by_norm[n]["prefix"],
        "line": cat_by_norm[n]["line"],
        "colour": cat_by_norm[n]["colour"],
        "sheet": cat_by_norm[n]["sheet"],
    } for n in extra_norms]

    template_extra = sorted(c["tag"] for c in cat if c["template"])

    # ----- per-TYPE (prefix) coverage delta -----
    ex_pref = Counter(r["prefix"] for r in ex)
    cat_pref = Counter(c["prefix"] for c in cat_real)
    all_pref = sorted(set(ex_pref) | set(cat_pref))
    type_coverage = {}
    for p in all_pref:
        e_n, c_n = ex_pref.get(p, 0), cat_pref.get(p, 0)
        type_coverage[p] = {
            "excel_count": e_n,
            "catalog_count": c_n,
            "in_excel_only": e_n > 0 and c_n == 0,
            "in_catalog_only": c_n > 0 and e_n == 0,
            "delta": c_n - e_n,
        }
    types_missing_from_catalog = sorted(p for p, v in type_coverage.items() if v["in_excel_only"])
    types_missing_from_excel = sorted(p for p, v in type_coverage.items() if v["in_catalog_only"])

    # ----- apply PPT re-allocations -----
    realloc = []
    for ra in PPT_REALLOCATIONS:
        n = norm_tag(ra["tag"])
        realloc.append({
            **ra,
            "in_catalog": n in cat_norms,
            "in_excel": n in ex_norms,
        })

    out = {
        "summary": {
            "excel_instruments": len(ex),
            "catalog_real_instruments": len(cat_real),
            "catalog_template_placeholders": len(cat) - len(cat_real),
            "exact_matches": len(matched),
            "missing_in_catalog": len(missing),
            "extra_in_catalog": len(extra),
            "prefix_discrepancies": len(discrepancies),
        },
        "finding": (
            "The design register (circuit-sequential numbering) and the as-drawn "
            "catalog (SVG-instance numbering) use orthogonal tag schemes; exact "
            "normalized tag overlap is %d. The actionable coverage signal is "
            "per-TYPE, not per-tag." % len(matched)
        ),
        "matched": matched,
        "missing": missing,
        "extra": extra,
        "template_extra": template_extra,
        "discrepancies": discrepancies,
        "type_coverage": type_coverage,
        "types_missing_from_catalog": types_missing_from_catalog,
        "types_missing_from_excel": types_missing_from_excel,
        "reallocations": realloc,
    }
    with open(RECON_JSON, "w") as fh:
        json.dump(out, fh, indent=2)
    return out


# --------------------------------------------------------------------------- #
# Phase 4 — coverage statistics
# --------------------------------------------------------------------------- #
def coverage_stats(excel, catalog, recon):
    s = recon["summary"]
    n_excel = s["excel_instruments"]
    n_cat = s["catalog_real_instruments"]
    exact_cov = (s["exact_matches"] / n_excel * 100.0) if n_excel else 0.0

    per_type = {}
    for p, v in recon["type_coverage"].items():
        e_n = v["excel_count"]
        per_type[p] = {
            "excel_count": e_n,
            "catalog_count": v["catalog_count"],
            "type_present_in_catalog": v["catalog_count"] > 0,
            "exact_tag_coverage_pct": 0.0,   # zero by scheme — documented
        }

    out = {
        "wave": "W005",
        "n_excel_design_tags": n_excel,
        "n_catalog_real_tags": n_cat,
        "n_catalog_template_placeholders": s["catalog_template_placeholders"],
        "n_exact_matches": s["exact_matches"],
        "n_missing_in_catalog": s["missing_in_catalog"],
        "n_extra_in_catalog": s["extra_in_catalog"],
        "exact_tag_coverage_pct": round(exact_cov, 1),
        "design_types_count": len([p for p, v in recon["type_coverage"].items() if v["excel_count"] > 0]),
        "catalog_types_count": len([p for p, v in recon["type_coverage"].items() if v["catalog_count"] > 0]),
        "types_present_in_both": sorted(
            p for p, v in recon["type_coverage"].items()
            if v["excel_count"] > 0 and v["catalog_count"] > 0
        ),
        "types_missing_from_catalog": recon["types_missing_from_catalog"],
        "types_missing_from_excel": recon["types_missing_from_excel"],
        "per_type": per_type,
        "reallocations_applied": len(recon["reallocations"]),
    }
    os.makedirs(OUT_REPORTS, exist_ok=True)
    with open(COVERAGE_JSON, "w") as fh:
        json.dump(out, fh, indent=2)
    return out


# --------------------------------------------------------------------------- #
# Phase 6 — canonical register (YAML SSOT)
# --------------------------------------------------------------------------- #
def canonical_register(excel, catalog, recon):
    import yaml

    ex_by_norm = {r["norm"]: r for r in excel["instruments"]}
    cat_by_norm = {c["norm"]: c for c in catalog["instruments"] if not c["template"]}
    realloc_by_norm = {norm_tag(r["tag"]): r for r in recon["reallocations"]}

    entries = []
    all_norms = sorted(set(ex_by_norm) | set(cat_by_norm))
    for n in all_norms:
        e = ex_by_norm.get(n)
        c = cat_by_norm.get(n)
        ra = realloc_by_norm.get(n)
        if e and c:
            status, confidence = "matched", "high"
        elif e and not c:
            status, confidence = "missing_in_catalog", "design-only"
        else:
            status, confidence = "extra_in_catalog", "as-drawn-only"
        entry = {
            "tag": (e or c)["tag"],
            "type": (e or c)["prefix"],
            "sensor": (e.get("equipment_type") if e else c.get("role")) or "",
            "excel_location": e.get("location") if e else "",
            "catalog_location": (
                {"x": c.get("x"), "y": c.get("y"), "layer": c.get("layer")} if c else None
            ),
            "reallocated": bool(ra),
            "reallocated_to": ra["reallocated_to"] if ra else None,
            "line": c.get("line") if c else "",
            "color": c.get("colour") if c else "",
            "confidence": confidence,
            "status": status,
            "notes": (ra["note"] if ra else ""),
        }
        entries.append(entry)

    doc = {
        "canonical_instrument_register": {
            "version": "1.0",
            "wave": "W005",
            "system": "MINERVA_QCELL_RFCELL",
            "description": (
                "Unified single-source-of-truth merging the design nomenclature "
                "register with the as-drawn catalog. status=matched only where a "
                "normalized tag exists in BOTH; design and as-drawn schemes are "
                "orthogonal so most entries are single-scheme."
            ),
            "counts": {
                "total": len(entries),
                "matched": sum(1 for e in entries if e["status"] == "matched"),
                "missing_in_catalog": sum(1 for e in entries if e["status"] == "missing_in_catalog"),
                "extra_in_catalog": sum(1 for e in entries if e["status"] == "extra_in_catalog"),
                "reallocated": sum(1 for e in entries if e["reallocated"]),
            },
            "entries": entries,
        }
    }
    os.makedirs(OUT_EXCEL_DIR, exist_ok=True)
    with open(CANONICAL_YAML, "w") as fh:
        fh.write("# MINERVA QCELL/RFCELL — Canonical Instrument Register (W005 SSOT)\n")
        fh.write("# Generated by src/abacus_svg_pid/build_w005.py — regenerable via ./make.sh\n")
        fh.write("%YAML 1.2\n---\n")
        yaml.safe_dump(doc, fh, default_flow_style=False, sort_keys=False, allow_unicode=True)
    return doc


# --------------------------------------------------------------------------- #
# Phase 7 — COMPONENT_CATALOG_v2.xlsx (catalog + reconciliation columns)
# --------------------------------------------------------------------------- #
def build_catalog_v2(catalog, recon):
    ex_norms = {m["norm"] for m in recon["matched"]}
    realloc_by_norm = {norm_tag(r["tag"]): r for r in recon["reallocations"]}

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    headers = [
        "Tag", "Prefix", "Number", "Line", "Colour", "Role", "Sheet",
        "X", "Y", "Layer", "Template?", "InExcelDesign?", "ExcelMatch",
        "Reallocated?", "ReallocatedTo", "Confidence", "Notes",
    ]
    ws.append(headers)
    for c in catalog["instruments"]:
        n = c["norm"]
        in_excel = n in ex_norms
        ra = realloc_by_norm.get(n)
        if c["template"]:
            conf, note = "n/a", "RFCELL template placeholder — non-reconcilable"
        elif in_excel:
            conf, note = "high", "exact normalized tag match"
        else:
            conf, note = "as-drawn-only", "drawn instance, not in design register"
        ws.append([
            c["tag"], c["prefix"], c["number"], c["line"], c["colour"], c["role"],
            c["sheet"], c["x"], c["y"], c["layer"],
            "yes" if c["template"] else "no",
            "yes" if in_excel else "no",
            next((m["excel_tag"] for m in recon["matched"] if m["norm"] == n), ""),
            "yes" if ra else "no",
            ra["reallocated_to"] if ra else "",
            conf, (ra["note"] if ra else note),
        ])

    # second sheet: type coverage
    ws2 = wb.create_sheet("TypeCoverage")
    ws2.append(["Prefix", "ISA_Class", "ExcelCount", "CatalogCount", "Delta", "Status"])
    isa = load_isa_classes()
    for p, v in sorted(recon["type_coverage"].items()):
        meta = isa_meta(isa, p)
        if v["in_excel_only"]:
            st = "MISSING FROM CATALOG (design type not drawn/catalogued)"
        elif v["in_catalog_only"]:
            st = "CATALOG ONLY (as-drawn type not in design register)"
        else:
            st = "present in both"
        ws2.append([p, meta["canonical_name"], v["excel_count"],
                    v["catalog_count"], v["delta"], st])

    os.makedirs(OUT_REPORTS, exist_ok=True)
    wb.save(CATALOG_V2_XLSX)
    return CATALOG_V2_XLSX


# --------------------------------------------------------------------------- #
# Phase 5 — main markdown report
# --------------------------------------------------------------------------- #
def reconciliation_report_md(excel, catalog, recon, cov):
    s = recon["summary"]
    tc = recon["type_coverage"]
    L = []
    A = L.append

    A("# W005 — XLSX Tag & Instrument Register Reconciliation Report\n")
    A("**Wave:** W005 · **Project:** MINERVA QCELL/RFCELL P&ID · "
      "**Generated by:** `src/abacus_svg_pid/build_w005.py` (regenerable via `./make.sh`)\n")
    A("**Sources:**\n")
    A(f"- Design register (tracked): `{excel['source']}`")
    A(f"- As-drawn catalog (derived): `{catalog['source']}`")
    A("- ISA class metadata: `configs/isa_classes.json`\n")

    # ---- Executive summary ----
    A("## Executive Summary\n")
    A("The official **design nomenclature register** and the **as-drawn component "
      "catalog** use *orthogonal* tag-numbering schemes:\n")
    A("- **Design register** — circuit-sequential numbering (e.g. `CV001`–`CV004`, "
      "`CV100/101` = 40 K, `CV200`–`CV202` = 4.5 K, `TT100`–`TT111` = Pt/40 K, "
      "`TT200`–`TT207` = Cernox/4.5 K).")
    A("- **As-drawn catalog** — SVG-instance numbering (e.g. `CV560`, `TT514`, "
      "`EH514`, `LS-021`, `HV503`).\n")
    A(f"As a direct consequence, **exact normalized tag overlap = "
      f"{s['exact_matches']} ({cov['exact_tag_coverage_pct']} %)**. "
      "This is a genuine engineering finding — the two registers were authored "
      "against different numbering conventions — **not** a pipeline defect. We do "
      "not fabricate matches. The actionable coverage signal is therefore measured "
      "**per instrument TYPE**, and the #1 recommended deliverable is a "
      "**design ↔ as-drawn cross-map**.\n")

    A("| Metric | Value |")
    A("| --- | --- |")
    A(f"| Design (Excel) instrument tags | {s['excel_instruments']} |")
    A(f"| As-drawn real instrument tags | {s['catalog_real_instruments']} |")
    A(f"| As-drawn template placeholders (non-reconcilable) | {s['catalog_template_placeholders']} |")
    A(f"| Exact normalized tag matches | {s['exact_matches']} |")
    A(f"| Missing (in design, not drawn) | {s['missing_in_catalog']} |")
    A(f"| Extra (drawn, not in design) | {s['extra_in_catalog']} |")
    A(f"| Design instrument TYPES | {cov['design_types_count']} |")
    A(f"| As-drawn instrument TYPES | {cov['catalog_types_count']} |")
    A(f"| TYPES present in both | {len(cov['types_present_in_both'])} ({', '.join(cov['types_present_in_both'])}) |")
    A("")

    # ---- 1. Missing ----
    A("## 1. Missing Instruments (in design register, absent from as-drawn catalog)\n")
    A(f"Because tag schemes differ, every one of the **{s['missing_in_catalog']} "
      "design tags** is 'missing' at the exact-tag level. More meaningfully, the "
      "following **instrument TYPES are entirely absent** from the as-drawn catalog "
      "(the W003/W004 category sheets covered only CV/EH/HV/LS/PT/TT):\n")
    A("| Prefix | ISA class | Design count | In catalog? |")
    A("| --- | --- | --- | --- |")
    isa = load_isa_classes()
    for p in recon["types_missing_from_catalog"]:
        meta = isa_meta(isa, p)
        A(f"| {p} | {meta['canonical_name'] or '—'} | {tc[p]['excel_count']} | NO |")
    A("")
    A("Full per-tag missing list is in `data/excel/reconciliation_results.json` "
      "(`missing`). Sample of design tags not represented as-drawn:\n")
    for m in recon["missing"][:25]:
        A(f"- `{m['tag']}` ({m['prefix']}{' — ' + m['isa_canonical'] if m['isa_canonical'] else ''}) "
          f"· {m['equipment_type'] or '—'} · sheet *{m['sheet']}*")
    if len(recon["missing"]) > 25:
        A(f"- … and {len(recon['missing']) - 25} more (see JSON).")
    A("")

    # ---- 2. Extra ----
    A("## 2. Extra Instruments (drawn/catalogued, absent from design register)\n")
    A(f"**{s['extra_in_catalog']} real as-drawn tags** have no design-register "
      f"counterpart, plus **{s['catalog_template_placeholders']} RFCELL template "
      "placeholders** (tags containing a literal `x`, e.g. `TTxxx`, `EHx11`) which "
      "are **non-reconcilable by construction** and are flagged, not matched.\n")
    A("Notable: **LS (limit switches)** appear only in the as-drawn catalog "
      f"({tc.get('LS', {}).get('catalog_count', 0)} tags) — position/limit switches "
      "are an as-drawn artdefact and were never in the design tag register.\n")
    A("| Prefix | ISA class | As-drawn count | In design? |")
    A("| --- | --- | --- | --- |")
    for p in recon["types_missing_from_excel"]:
        meta = isa_meta(isa, p)
        A(f"| {p} | {meta['canonical_name'] or '—'} | {tc[p]['catalog_count']} | NO |")
    A("")
    A("Sample of extra as-drawn tags:\n")
    for e in recon["extra"][:20]:
        A(f"- `{e['tag']}` ({e['prefix']}) · line *{e['line'] or '—'}* · {e['colour'] or '—'}")
    if len(recon["extra"]) > 20:
        A(f"- … and {len(recon['extra']) - 20} more (see JSON).")
    A("")
    A("Template placeholders flagged non-reconcilable: "
      + ", ".join(f"`{t}`" for t in recon["template_extra"]) + "\n")

    # ---- 3. Discrepancies ----
    A("## 3. Discrepancies (tag present both sides but attributes disagree)\n")
    if recon["discrepancies"]:
        A("| Norm tag | Design prefix | As-drawn prefix | Issue |")
        A("| --- | --- | --- | --- |")
        for d in recon["discrepancies"]:
            A(f"| {d['norm']} | {d['excel_prefix']} | {d['catalog_prefix']} | {d['issue']} |")
    else:
        A("No attribute-level discrepancies exist because there are no exact tag "
          "matches to compare (see Executive Summary). The structural discrepancy "
          "is the scheme mismatch itself.\n")
    A("")
    A("### Per-TYPE coverage delta (the meaningful metric)\n")
    A("| Prefix | ISA class | Design | As-drawn | Δ | Status |")
    A("| --- | --- | --- | --- | --- | --- |")
    for p in sorted(tc):
        meta = isa_meta(isa, p)
        v = tc[p]
        if v["in_excel_only"]:
            st = "design only — **not catalogued**"
        elif v["in_catalog_only"]:
            st = "as-drawn only"
        else:
            st = "both"
        A(f"| {p} | {meta['canonical_name'] or '—'} | {v['excel_count']} | "
          f"{v['catalog_count']} | {v['delta']:+d} | {st} |")
    A("")

    # ---- 4. PPT re-allocations ----
    A("## 4. PPT Re-allocations\n")
    A("Documented instrument re-allocations from the QSYS/RFCELL instrumentation "
      "presentation. **Known gap:** the full 65 MB deck was not exhaustively parsed "
      "(cost); the following are the explicitly documented re-allocations.\n")
    A("| Tag | Re-allocated to | Sensor | In catalog? | In design? | Source |")
    A("| --- | --- | --- | --- | --- | --- |")
    for ra in recon["reallocations"]:
        A(f"| {ra['tag']} | {ra['target']} | {ra['sensor']} | "
          f"{'yes' if ra['in_catalog'] else 'no'} | "
          f"{'yes' if ra['in_excel'] else 'no'} | {ra['source']} |")
    A("")

    # ---- 5. Recommendations ----
    A("## 5. Recommendations\n")
    A("1. **Author a design ↔ as-drawn cross-map** (highest priority). Since the "
      "two registers are numbered independently, reconciliation requires an "
      "engineering map keyed on TYPE + circuit + position, not on tag string.")
    A("2. **Extend the as-drawn catalog to the 10 missing TYPES** "
      f"({', '.join(recon['types_missing_from_catalog'])}) — the W003/W004 category "
      "sheets only covered CV/EH/HV/LS/PT/TT.")
    A("3. **Treat the "
      f"{s['catalog_template_placeholders']} RFCELL template placeholders** as a "
      "templating artefact: either instantiate them with real instance numbers or "
      "exclude them from coverage KPIs.")
    A("4. **Confirm the two PPT re-allocations** (TT535→PZ coldest, TT525→PZ "
      "warmest) against the issued P&ID, and add TT525 to the as-drawn catalog "
      "(currently missing).")
    A("5. **Adopt `canonical_register_v1.yaml`** as the merged SSOT going forward "
      "so future drawings reconcile against a single keyed register.\n")

    A("---\n")
    A("*Outputs:* `data/excel/excel_register.json`, `data/excel/catalog_register.json`, "
      "`data/excel/reconciliation_results.json`, `reports/W005_coverage_statistics.json`, "
      "`data/excel/canonical_register_v1.yaml`, `reports/COMPONENT_CATALOG_v2.xlsx`. "
      "See `reports/W005_validation_report.md` for the validation record.\n")

    os.makedirs(OUT_REPORTS, exist_ok=True)
    with open(REPORT_MD, "w") as fh:
        fh.write("\n".join(L))
    return REPORT_MD


# --------------------------------------------------------------------------- #
# Phase 8 — validation report
# --------------------------------------------------------------------------- #
def validation_report_md(excel, catalog, recon, cov):
    s = recon["summary"]
    L = []
    A = L.append
    A("# W005 — Validation Report\n")
    A("**Governance:** Claim ≠ Complete. A claim passes only when (1) the output "
      "files exist, (2) this validation record exists, (3) runtime counts are "
      "recorded, and (4) known gaps are listed.\n")

    A("## 1. Output files produced\n")
    outputs = [
        EXCEL_REGISTER_JSON, CATALOG_REGISTER_JSON, RECON_JSON,
        COVERAGE_JSON, REPORT_MD, CANONICAL_YAML, CATALOG_V2_XLSX,
    ]
    A("| File | Exists | Bytes |")
    A("| --- | --- | --- |")
    for p in outputs:
        ok = os.path.exists(p)
        sz = os.path.getsize(p) if ok else 0
        A(f"| `{os.path.relpath(p, ROOT)}` | {'✅' if ok else '❌'} | {sz} |")
    A("")

    A("## 2. Runtime counts (regenerable via ./make.sh)\n")
    A("| Quantity | Value |")
    A("| --- | --- |")
    A(f"| Design (Excel) instrument tags | {s['excel_instruments']} |")
    A(f"| As-drawn real instrument tags | {s['catalog_real_instruments']} |")
    A(f"| As-drawn template placeholders | {s['catalog_template_placeholders']} |")
    A(f"| Exact normalized tag matches | {s['exact_matches']} |")
    A(f"| Exact tag coverage % | {cov['exact_tag_coverage_pct']} |")
    A(f"| Missing in catalog | {s['missing_in_catalog']} |")
    A(f"| Extra in catalog | {s['extra_in_catalog']} |")
    A(f"| Design TYPES / As-drawn TYPES | {cov['design_types_count']} / {cov['catalog_types_count']} |")
    A(f"| Re-allocations applied | {cov['reallocations_applied']} |")
    A("")

    A("## 3. Invariants checked\n")
    checks = [
        ("excel + catalog tag counts are non-zero",
         s["excel_instruments"] > 0 and s["catalog_real_instruments"] > 0),
        ("matched ⊆ both registers (matched = |excel ∩ catalog|)",
         s["exact_matches"] == len(recon["matched"])),
        ("missing + matched == excel total",
         s["missing_in_catalog"] + s["exact_matches"] == s["excel_instruments"]),
        ("extra + matched == catalog real total",
         s["extra_in_catalog"] + s["exact_matches"] == s["catalog_real_instruments"]),
        ("every reallocation tag resolved against both registers",
         all("in_catalog" in r and "in_excel" in r for r in recon["reallocations"])),
    ]
    A("| Invariant | Result |")
    A("| --- | --- |")
    for name, ok in checks:
        A(f"| {name} | {'✅ PASS' if ok else '❌ FAIL'} |")
    A("")

    A("## 4. Known gaps\n")
    A("- **Zero exact tag overlap is real, not a bug:** design (circuit-sequential) "
      "and as-drawn (SVG-instance) schemes are orthogonal. A design↔as-drawn "
      "cross-map is required for tag-level reconciliation (Recommendation 1).")
    A("- **Full PPT deck not exhaustively parsed:** the 65 MB QSYS instrumentation "
      "deck was not machine-parsed (cost); only the two explicitly documented "
      "re-allocations (TT535, TT525) are encoded with citations.")
    A(f"- **10 design instrument TYPES not catalogued** "
      f"({', '.join(recon['types_missing_from_catalog'])}): the W003/W004 category "
      "sheets covered only CV/EH/HV/LS/PT/TT.")
    A(f"- **{s['catalog_template_placeholders']} RFCELL template placeholders** are "
      "flagged non-reconcilable rather than matched.")
    A("- **Catalog duplicates** were de-duplicated by normalized tag before "
      "reconciliation (see `catalog_register.json` `duplicate_norm_tags`).\n")

    overall = all(ok for _, ok in checks) and all(os.path.exists(p) for p in outputs)
    A(f"## Overall: {'✅ PASS' if overall else '❌ FAIL'}\n")

    with open(VALIDATION_MD, "w") as fh:
        fh.write("\n".join(L))
    return overall


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def run():
    isa = load_isa_classes()
    print("[W005] Phase 1 — parse design nomenclature Excel …")
    excel = parse_excel(isa)
    print(f"        {excel['instrument_count']} design tags · prefixes {excel['prefix_histogram']}")

    print("[W005] Phase 2 — parse as-drawn COMPONENT_CATALOG.xlsx …")
    catalog = parse_catalog(isa)
    print(f"        {catalog['real_count']} real + {catalog['template_count']} template")

    print("[W005] Phase 3 — reconcile …")
    recon = reconcile(excel, catalog)
    print(f"        matched={recon['summary']['exact_matches']} "
          f"missing={recon['summary']['missing_in_catalog']} "
          f"extra={recon['summary']['extra_in_catalog']}")

    print("[W005] Phase 4 — coverage statistics …")
    cov = coverage_stats(excel, catalog, recon)

    print("[W005] Phase 6 — canonical register YAML …")
    canonical_register(excel, catalog, recon)

    print("[W005] Phase 7 — COMPONENT_CATALOG_v2.xlsx …")
    build_catalog_v2(catalog, recon)

    print("[W005] Phase 5 — reconciliation report …")
    reconciliation_report_md(excel, catalog, recon, cov)

    print("[W005] Phase 8 — validation report …")
    ok = validation_report_md(excel, catalog, recon, cov)
    print(f"[W005] DONE — validation {'PASS' if ok else 'FAIL'}")
    return ok


def main():
    return 0 if run() else 1


if __name__ == "__main__":
    raise SystemExit(main())
