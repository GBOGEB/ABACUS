"""
================================================================================
 Module : build_catalog.py
 Purpose: PHASE 7 - generate the component catalog (Excel + interactive HTML)
          and assign each instrument to its nearest colour process line so the
          catalog carries Tag / Colour / Line / Location traceability.
 Current Wave : W003 + W004
 Status : ACTIVE
 Inputs  : segmentation/data/*_segmentation.json ; GeometryModel lines ;
           data/model/spec_dots_catalog.json ; data/model/paired_elements.json
 Outputs : reports/COMPONENT_CATALOG.xlsx
           publish/component_catalog.html
 Notes   : Line assignment is nearest-colour-line heuristic (confidence noted).
================================================================================
"""

from __future__ import annotations

import json
import math
import os
from collections import defaultdict

from abacus_svg_pid import build_w003_w004 as B
from abacus_svg_pid import parser as P

ROOT = B.ROOT
MODEL = B.MODEL
REPORTS = os.path.join(ROOT, "reports")
PUBLISH = os.path.join(ROOT, "publish")

CODE_COLOUR = {"A": "#0000FF", "A_prime": "#000080", "B": "#00FFFF",
               "W": "#00FF00", "S": "#808000", "V": "#999999",
               "D": "#FF8000", "E": "#FF0000"}

CATEGORY_SHEETS = {
    "CV": "Control Valves", "TT": "Temperature", "EH": "Electrical Heaters",
    "PT": "Pressure", "LS": "Limit Switches", "HV": "Manual Valves",
}


def _dist(ax, ay, bx, by):
    return math.hypot(ax - bx, ay - by)


def assign_lines(models, instruments):
    """For each instrument, find nearest colour process line -> (code, colour)."""
    line_index = {}
    for key, m in models.items():
        line_index[key] = [e for e in m.elements
                           if e.process_code in B.PROC_CODES
                           and e.shape in ("line", "path") and e.cx is not None]
    enriched = []
    for key, seg in instruments.items():
        lines = line_index.get(key, [])
        for inst in seg.get("instruments", []):
            x, y = inst.get("x"), inst.get("y")
            code, colour, dist = "TBD", None, None
            if lines and x is not None and y is not None:
                nearest = min(lines, key=lambda l: _dist(x, y, l.cx, l.cy))
                d = _dist(x, y, nearest.cx, nearest.cy)
                if d < 60:
                    code = nearest.process_code
                    colour = CODE_COLOUR.get(code)
                    dist = round(d, 1)
            enriched.append({
                "sheet": key, "tag": inst.get("tag"),
                "prefix": inst.get("prefix"), "number": inst.get("number"),
                "meaning": inst.get("meaning"), "role": inst.get("role"),
                "line": code, "colour": colour, "x": x, "y": y,
                "layer": inst.get("layer"), "assign_dist_px": dist,
            })
    return enriched


def build_excel(enriched, dots, heat_loads):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # overview sheet
    ws0 = wb.active
    ws0.title = "00_Overview"
    ws0.append(["MINERVA QCELL/RFCELL P&ID - Component Catalog"])
    ws0.append(["Wave", "W003 + W004"])
    ws0.append(["Total instruments", len(enriched)])
    by_prefix = defaultdict(int)
    for e in enriched:
        by_prefix[e["prefix"]] += 1
    ws0.append([])
    ws0.append(["Prefix", "Count", "Meaning"])
    for pfx in sorted(by_prefix):
        sample = next((e for e in enriched if e["prefix"] == pfx), {})
        ws0.append([pfx, by_prefix[pfx], sample.get("meaning", "")])
    ws0["A1"].font = Font(bold=True, size=14)

    cols = ["Tag", "Prefix", "Number", "Line", "Colour", "Role",
            "Sheet", "X", "Y", "Layer", "AssignDist_px"]
    for pfx, title in CATEGORY_SHEETS.items():
        ws = wb.create_sheet(f"{pfx}_{title.replace(' ', '')[:20]}")
        ws.append(cols)
        rows = [e for e in enriched if e["prefix"] == pfx]
        for e in rows:
            ws.append([e["tag"], e["prefix"], e["number"], e["line"],
                       e["colour"] or "", e["role"], e["sheet"],
                       e["x"], e["y"], e["layer"], e["assign_dist_px"]])
        style_header(ws, len(cols))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(64+len(cols))}{ws.max_row}"

    # heat loads sheet
    wsh = wb.create_sheet("HeatLoads")
    wsh.append(["ID", "Colour", "Line", "Size_px", "ParentLine", "Energy_W"])
    for i, t in enumerate(heat_loads, 1):
        wsh.append([f"HL_{t['process_code']}_{i:02d}", t["colour"],
                    t["process_code"], t["size_px"], t["parent_line_code"], "TBD"])
    style_header(wsh, 6)
    wsh.freeze_panes = "A2"

    # spec dots sheet
    wsd = wb.create_sheet("SpecDots")
    wsd.append(["Line", "Count", "Type", "ExampleIDs"])
    for code, info in dots["dots_per_line"].items():
        wsd.append([code, info["count"], "P/T/Size spec change",
                    ", ".join(info["ids"][:5])])
    style_header(wsd, 4)

    os.makedirs(REPORTS, exist_ok=True)
    out = os.path.join(REPORTS, "COMPONENT_CATALOG.xlsx")
    wb.save(out)
    return out


def build_html(enriched):
    rows = []
    for e in enriched:
        sw = e["colour"] or "#cccccc"
        rows.append(
            f'<tr data-prefix="{e["prefix"]}" data-line="{e["line"]}">'
            f'<td>{e["tag"] or ""}</td><td>{e["prefix"] or ""}</td>'
            f'<td>{e["line"]}</td>'
            f'<td><span class="sw" style="background:{sw}"></span>{e["colour"] or "-"}</td>'
            f'<td>{e["role"] or ""}</td><td>{e["sheet"]}</td>'
            f'<td>{e["meaning"] or ""}</td><td>{e["layer"] or ""}</td></tr>')
    prefixes = sorted({e["prefix"] for e in enriched if e["prefix"]})
    opts = "".join(f'<option value="{p}">{p}</option>' for p in prefixes)
    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="utf-8"/>
<title>QCELL P&amp;ID - Component Catalog (W003/W004)</title>
<style>
 body{{font-family:Consolas,'DejaVu Sans Mono',monospace;margin:0;padding:18px;background:#f7f7f7;color:#111;}}
 h1{{font-size:20px;margin:0 0 4px;}}
 .bar{{margin:10px 0;font-size:13px;}}
 input,select{{font-family:inherit;padding:5px 8px;border:1px solid #aaa;}}
 table{{border-collapse:collapse;width:100%;background:#fff;font-size:12px;}}
 th,td{{border:1px solid #ddd;padding:5px 7px;text-align:left;}}
 th{{background:#305496;color:#fff;position:sticky;top:0;cursor:pointer;}}
 tr:nth-child(even){{background:#f4f7fb;}}
 .sw{{display:inline-block;width:12px;height:12px;border:1px solid #333;margin-right:5px;vertical-align:middle;}}
 .count{{color:#555;}}
</style></head><body>
<h1>MINERVA QCELL / RFCELL &mdash; Component Catalog</h1>
<div class="bar">Wave W003+W004 &middot; line/colour assigned by nearest-colour-line proximity (heuristic). Monospace = Consolas.</div>
<div class="bar">
 Search: <input id="q" placeholder="tag / meaning..." oninput="flt()"/>
 Prefix: <select id="pf" onchange="flt()"><option value="">all</option>{opts}</select>
 <span class="count" id="cnt"></span>
</div>
<table id="t"><thead><tr>
<th>Tag</th><th>Prefix</th><th>Line</th><th>Colour</th><th>Role</th><th>Sheet</th><th>Meaning</th><th>Layer</th>
</tr></thead><tbody>
{''.join(rows)}
</tbody></table>
<script>
function flt(){{
 var q=document.getElementById('q').value.toLowerCase();
 var pf=document.getElementById('pf').value;
 var n=0, rows=document.querySelectorAll('#t tbody tr');
 rows.forEach(function(r){{
   var okp = !pf || r.dataset.prefix===pf;
   var okq = !q || r.textContent.toLowerCase().indexOf(q)>=0;
   var show = okp&&okq; r.style.display=show?'':'none'; if(show)n++;
 }});
 document.getElementById('cnt').textContent=n+' / '+rows.length+' shown';
}}
flt();
</script></body></html>"""
    os.makedirs(PUBLISH, exist_ok=True)
    out = os.path.join(PUBLISH, "component_catalog.html")
    with open(out, "w") as fh:
        fh.write(html)
    return out


def run(models, instruments):
    enriched = assign_lines(models, instruments)
    with open(os.path.join(MODEL, "spec_dots_catalog.json")) as fh:
        dots = json.load(fh)
    with open(os.path.join(MODEL, "paired_elements.json")) as fh:
        heat_loads = json.load(fh)["triangles_to_lines"]
    xlsx = build_excel(enriched, dots, heat_loads)
    html = build_html(enriched)
    # save enriched assignment too
    with open(os.path.join(MODEL, "component_line_assignment.json"), "w") as fh:
        json.dump({"count": len(enriched), "components": enriched}, fh, indent=2)
    return {"xlsx": xlsx, "html": html, "components": len(enriched)}


def main():
    """CLI entrypoint: load models + instruments, build the catalog."""
    os.makedirs(REPORTS, exist_ok=True)
    os.makedirs(PUBLISH, exist_ok=True)
    models = B.load_models()
    instruments = B.load_instruments()
    result = run(models, instruments)
    print(json.dumps(result, indent=2))
    return result


if __name__ == "__main__":
    main()
