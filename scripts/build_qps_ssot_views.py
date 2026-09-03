#!/usr/bin/env python3
"""Build consolidated QPS Excel + HTML views from one normalized row dataset.

Input contract: JSON object with ``rows`` list, or a top-level JSON list. Each row is
an actionable Global_ID record from QPS_LKT_NEGO_RFI_COMPLIANCE_SSOT_v1.0.

This builder is presentation-only. It does not calculate or promote compliance.
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

VISIBLE_SHEETS = ["START_HERE", "MASTER_REVIEW", "NEG_RFI_RETURNS", "EVIDENCE_LINEAGE", "DASHBOARD"]
HIDDEN_SHEETS = ["LISTS_CONTROLS", "RAW_RELATIONS", "SOURCE_HASHES"]
IDENTITY = ["Global_ID", "Agenda_ID", "Priority", "Short_Topic", "OFFER_IDs", "RTM_IDs"]
MASTER = IDENTITY + [
    "Evidence_Class", "Evidence_Maturity", "Source_Type", "Compliance_State",
    "Reviewer_Assessment", "RFI_ID", "NEG_ID", "Disposition", "SCK_Position",
    "Requested_Return", "Owner", "Due_Date", "Reentry_State",
]
RETURNS = [
    "Global_ID", "Agenda_ID", "Short_Topic", "RFI_ID", "NEG_ID", "Meeting_Question",
    "Requested_Return", "Bidder_Return", "Minute_Reference", "Minute_Disposition",
    "BAFO_Evidence_Location", "Post_Award_Proof", "Canonical_Reentry_Test",
    "Reentry_State", "Owner", "Due_Date",
]
LINEAGE = [
    "Global_ID", "Short_Topic", "Source_File", "Source_Page_Table_Section", "Exact_Finding",
    "OFFER_IDs", "RTM_IDs", "COM_IDs", "GEN_IDs", "HSE_IDs", "TODO_IDs", "ADR_IDs",
    "OCD_IDs", "ICD_IDs", "Card_IDs", "BOM_IDs", "Utility_Building_IDs", "Source_SHA256",
]
LONG_TEXT = {"Exact_Finding", "Bidder_Position", "SCK_Position", "Meeting_Question", "Requested_Return", "Bidder_Return", "Canonical_Reentry_Test"}


def load_rows(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = data.get("rows", []) if isinstance(data, dict) else data
    if not isinstance(rows, list):
        raise ValueError("Input must be a JSON list or an object containing a rows list")
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for i, row in enumerate(rows, 1):
        if not isinstance(row, dict):
            raise ValueError(f"Row {i} is not an object")
        gid = str(row.get("Global_ID", "")).strip()
        if not gid:
            raise ValueError(f"Row {i} has no Global_ID")
        if gid in seen:
            raise ValueError(f"Duplicate Global_ID: {gid}")
        seen.add(gid)
        out.append(row)
    return out


def scalar(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (list, tuple, set)):
        return "; ".join(str(x) for x in v)
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False, sort_keys=True)
    return str(v)


def load_style(path: Path | None) -> dict[str, Any]:
    fallback = {
        "palette": {"tokens": {"define": "#562873", "warning": "#B5622A", "ok": "#1D7A5F", "critical": "#B42318", "surface": "#FFFFFF", "text": "#1F1F29"}},
        "typography": {"body": "Aptos", "fallback_body": "Arial"},
    }
    if path and path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return fallback
    return fallback


def hex6(value: str, fallback: str) -> str:
    value = (value or fallback).replace("#", "").upper()
    return value if len(value) == 6 else fallback.replace("#", "")


def apply_sheet(ws, columns: list[str], rows: list[dict[str, Any]], style: dict[str, Any]) -> None:
    tokens = style.get("palette", {}).get("tokens", {})
    purple = hex6(tokens.get("define", "#562873"), "562873")
    surface = hex6(tokens.get("surface", "#FFFFFF"), "FFFFFF")
    font_name = style.get("typography", {}).get("body", "Aptos")
    ws.append(columns)
    for c, name in enumerate(columns, 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=font_name, size=10, bold=True, color=surface)
        cell.fill = PatternFill("solid", fgColor=purple)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([scalar(row.get(c)) for c in columns])
    ws.freeze_panes = "D2" if "Global_ID" in columns else "A2"
    ws.auto_filter.ref = ws.dimensions
    for r in range(2, ws.max_row + 1):
        for c, name in enumerate(columns, 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=font_name, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=name in LONG_TEXT)
        ws.row_dimensions[r].height = 30 if any(scalar(rows[r-2].get(x)) for x in LONG_TEXT if x in columns) else 18
    for c, name in enumerate(columns, 1):
        letter = get_column_letter(c)
        width = 14
        if name in {"Global_ID", "Agenda_ID", "Priority", "RFI_ID", "NEG_ID", "Owner", "Due_Date"}:
            width = 13
        elif name in LONG_TEXT:
            width = 44
        elif name in {"Short_Topic", "OFFER_IDs", "RTM_IDs", "Compliance_State", "Reviewer_Assessment"}:
            width = 24
        ws.column_dimensions[letter].width = width
    # Group columns after the reader-facing core columns.
    if len(columns) > 10:
        ws.column_dimensions.group(get_column_letter(11), get_column_letter(len(columns)), hidden=True)


def build_xlsx(rows: list[dict[str, Any]], out: Path, source_hash: str, style: dict[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name in VISIBLE_SHEETS + HIDDEN_SHEETS:
        wb.create_sheet(name)
    start = wb["START_HERE"]
    start["A1"] = "QPS LKT NEG / RFI / COMPLIANCE — canonical review view"
    start["A2"] = "Authority"
    start["B2"] = "Derived working view; engineering/compliance authority remains cryoplant-project"
    start["A3"] = "Source row SHA256"
    start["B3"] = source_hash
    start["A4"] = "Formal completion"
    start["B4"] = "70/90 = 77.78% unless changed by accepted source-bound re-entry"
    start["A5"] = "Negotiation resolution"
    start["B5"] = "0/20 unless accepted bidder/minute re-entry changes it"
    start["A7"] = "Semantic rule"
    start["B7"] = "Deviation != NOK; PARTIAL retains both supported and residual conditions"
    start["A9"] = "Visible sheets"
    start["B9"] = ", ".join(VISIBLE_SHEETS)
    start.column_dimensions["A"].width = 24
    start.column_dimensions["B"].width = 95
    start.freeze_panes = "A2"
    font_name = style.get("typography", {}).get("body", "Aptos")
    for row in start.iter_rows():
        for cell in row:
            cell.font = Font(name=font_name, size=11, bold=cell.column == 1)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_sheet(wb["MASTER_REVIEW"], MASTER, rows, style)
    apply_sheet(wb["NEG_RFI_RETURNS"], RETURNS, rows, style)
    apply_sheet(wb["EVIDENCE_LINEAGE"], LINEAGE, rows, style)

    dash = wb["DASHBOARD"]
    dash.append(["Metric", "Value"])
    metrics = {
        "Canonical actionable rows": len(rows),
        "Rows with NEG": sum(bool(scalar(r.get("NEG_ID"))) for r in rows),
        "Rows with RFI": sum(bool(scalar(r.get("RFI_ID"))) for r in rows),
        "Rows unresolved": sum(scalar(r.get("Reentry_State")).upper() not in {"CLOSED", "VERIFIED"} for r in rows),
        "Rows with source locator": sum(bool(scalar(r.get("Source_Page_Table_Section"))) for r in rows),
    }
    for k, v in metrics.items():
        dash.append([k, v])
    dash.column_dimensions["A"].width = 34
    dash.column_dimensions["B"].width = 18

    raw = wb["RAW_RELATIONS"]
    raw.append(["Global_ID", "Row_JSON"])
    for r in rows:
        raw.append([r.get("Global_ID"), json.dumps(r, ensure_ascii=False, sort_keys=True)])
    src = wb["SOURCE_HASHES"]
    src.append(["source_row_sha256", source_hash])
    lists = wb["LISTS_CONTROLS"]
    lists.append(["source_type", "reviewer_assessment", "action_route", "evidence_maturity"])
    lists.append(["DEVIATION", "PARTIAL", "NEG", "PHASE_BOUND"])
    for name in HIDDEN_SHEETS:
        wb[name].sheet_state = "hidden"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def build_html(rows: list[dict[str, Any]], out: Path, source_hash: str, style: dict[str, Any]) -> None:
    tokens = style.get("palette", {}).get("tokens", {})
    purple = tokens.get("define", "#562873")
    warning = tokens.get("warning", "#B5622A")
    ok = tokens.get("ok", "#1D7A5F")
    critical = tokens.get("critical", "#B42318")
    text = tokens.get("text", "#1F1F29")
    font = style.get("typography", {}).get("body", "Aptos")
    data = json.dumps(rows, ensure_ascii=False).replace("</", "<\\/")
    columns = MASTER + [c for c in RETURNS + LINEAGE if c not in MASTER]
    doc = f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>QPS LKT NEG/RFI/Compliance SSOT</title><style>
:root{{--purple:{purple};--warn:{warning};--ok:{ok};--critical:{critical};--text:{text};}}
*{{box-sizing:border-box}}body{{font-family:{font},Arial,sans-serif;color:var(--text);margin:0;background:#fff;font-size:14px}}
header{{position:sticky;top:0;background:#fff;border-bottom:1px solid #ddd;padding:12px 18px;z-index:10}}h1{{font-size:20px;margin:0 0 8px;color:var(--purple)}}
.controls{{display:flex;gap:8px;flex-wrap:wrap;align-items:center}}input,select,button{{font:inherit;padding:6px 8px}}button.active{{background:var(--purple);color:#fff}}
.meta{{font-size:12px;color:#666;margin-top:7px}}main{{padding:12px 18px}}.table-wrap{{overflow:auto;max-height:calc(100vh - 150px);border-top:1px solid #ddd}}
table{{border-collapse:separate;border-spacing:0;width:max-content;min-width:100%}}th,td{{padding:7px 9px;border-bottom:1px solid #eee;vertical-align:top;max-width:420px}}
th{{position:sticky;top:0;background:var(--purple);color:#fff;text-align:left;z-index:2}}td.identity{{position:sticky;left:0;background:#fff;z-index:1;font-weight:600}}
tr:hover td{{background:#faf7fc}}.dense th,.dense td{{padding:3px 6px;font-size:12px}}.advanced{{display:none}}.show-advanced .advanced{{display:table-cell}}
.status{{font-weight:700}}.status.NOK,.status.MAJOR_DISCREPANCY{{color:var(--critical)}}.status.PARTIAL,.status.REVIEW_REQUIRED,.status.TBD{{color:var(--warn)}}.status.COMPLY,.status.ACCEPTED_VARIANCE{{color:var(--ok)}}
details{{max-width:520px}}code{{font-family:Consolas,monospace}}@media(max-width:900px){{header{{position:static}}.table-wrap{{max-height:none}}}}
</style></head><body><header><h1>QPS LKT NEG / RFI / Compliance — canonical view</h1><div class="controls">
<input id="q" type="search" placeholder="Global search">
<select id="status" multiple size="1"><option value="">All statuses</option></select>
<input id="agenda" placeholder="Agenda ID"><input id="offer" placeholder="OFFER"><input id="rtm" placeholder="RTM"><input id="rfi" placeholder="RFI"><input id="neg" placeholder="NEG">
<label><input id="unresolved" type="checkbox"> unresolved only</label><button id="advanced">Advanced columns</button><button id="dense">Dense mode</button>
<span id="count"></span></div><div class="meta">source row SHA256 <code>{source_hash}</code> · derived read-only view · Deviation ≠ NOK</div></header>
<main><div class="table-wrap"><table id="tbl"><thead><tr></tr></thead><tbody></tbody></table></div></main>
<script>const ROWS={data}; const COLS={json.dumps(columns)}; const CORE=new Set({json.dumps(MASTER)});
const $=s=>document.querySelector(s), esc=s=>String(s??'').replace(/[&<>"']/g,m=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[m]));
function val(r,k){{let v=r[k];return Array.isArray(v)?v.join('; '):(v&&typeof v==='object'?JSON.stringify(v):String(v??''));}}
function statuses(){{return [...new Set(ROWS.map(r=>val(r,'Compliance_State')).filter(Boolean))].sort()}} statuses().forEach(s=>{{let o=document.createElement('option');o.value=s;o.textContent=s;$('#status').append(o)}});
function selectedStatus(){{return [...$('#status').selectedOptions].map(o=>o.value).filter(Boolean)}}
function filt(r){{let q=$('#q').value.toLowerCase();if(q&&!Object.values(r).some(v=>String(v).toLowerCase().includes(q)))return false;let ss=selectedStatus();if(ss.length&&!ss.includes(val(r,'Compliance_State')))return false;
for(const k of ['agenda','offer','rtm','rfi','neg']){{let needle=$('#'+k).value.toLowerCase();let col={{agenda:'Agenda_ID',offer:'OFFER_IDs',rtm:'RTM_IDs',rfi:'RFI_ID',neg:'NEG_ID'}}[k];if(needle&&!val(r,col).toLowerCase().includes(needle))return false}}
if($('#unresolved').checked&&['CLOSED','VERIFIED'].includes(val(r,'Reentry_State').toUpperCase()))return false;return true}}
function render(){{let rows=ROWS.filter(filt);let tr=$('#tbl thead tr');tr.innerHTML=COLS.map((c,i)=>`<th class="${{CORE.has(c)?'':'advanced'}}">${{esc(c)}}</th>`).join('');
$('#tbl tbody').innerHTML=rows.map(r=>'<tr id="row-'+esc(val(r,'Global_ID'))+'">'+COLS.map((c,i)=>{{let x=val(r,c), cls=(CORE.has(c)?'':'advanced')+(i===0?' identity':''); if(c==='Compliance_State')return `<td class="${{cls}}"><span class="status ${{esc(x)}}">${{esc(x)}}</span></td>`;if(['Exact_Finding','SCK_Position','Requested_Return','Bidder_Return'].includes(c)&&x.length>100)return `<td class="${{cls}}"><details><summary>${{esc(x.slice(0,95))}}…</summary>${{esc(x)}}</details></td>`;return `<td class="${{cls}}">${{esc(x)}}</td>`}}).join('')+'</tr>').join('');
$('#count').textContent=`${{rows.length}} / ${{ROWS.length}} rows`; let id=location.hash.replace('#','');if(id)document.getElementById('row-'+id)?.scrollIntoView();}}
for(const id of ['q','status','agenda','offer','rtm','rfi','neg','unresolved'])$('#'+id).addEventListener('input',render);$('#advanced').onclick=()=>{{document.body.classList.toggle('show-advanced');$('#advanced').classList.toggle('active')}};$('#dense').onclick=()=>{{document.body.classList.toggle('dense');$('#dense').classList.toggle('active')}};render();</script></body></html>'''
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(doc, encoding="utf-8")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("input_json", type=Path)
    p.add_argument("--out-dir", type=Path, default=Path("out/qps_ssot_views"))
    p.add_argument("--style", type=Path, default=Path("ssot/ssot_style.json"))
    args = p.parse_args()
    raw = args.input_json.read_bytes()
    source_hash = hashlib.sha256(raw).hexdigest()
    rows = load_rows(args.input_json)
    style = load_style(args.style)
    stem = "QPS_LKT_NEGO_RFI_COMPLIANCE_SSOT_v1.0"
    build_xlsx(rows, args.out_dir / f"{stem}.xlsx", source_hash, style)
    build_html(rows, args.out_dir / f"{stem}.html", source_hash, style)
    manifest = {"document_id": "QPS_SSOT_VIEW_BUILD_RECEIPT", "source_sha256": source_hash, "row_count": len(rows), "outputs": [f"{stem}.xlsx", f"{stem}.html"], "formal_credit": 0}
    (args.out_dir / f"{stem}.build.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))

if __name__ == "__main__":
    main()
