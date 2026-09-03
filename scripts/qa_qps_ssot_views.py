#!/usr/bin/env python3
"""Governed QA for QPS canonical SSOT Excel/HTML views.

Runs structural workbook checks and Playwright browser interaction checks. It is
presentation QA only and never changes compliance or negotiation disposition.
"""
from __future__ import annotations

import argparse
import asyncio
import json
import zipfile
from pathlib import Path

from openpyxl import load_workbook

VISIBLE = ["START_HERE", "MASTER_REVIEW", "NEG_RFI_RETURNS", "EVIDENCE_LINEAGE", "DASHBOARD"]
HIDDEN = ["LISTS_CONTROLS", "RAW_RELATIONS", "SOURCE_HASHES"]


def check_xlsx(path: Path) -> list[dict[str, object]]:
    checks: list[dict[str, object]] = []
    def add(name: str, ok: bool, detail: str = "") -> None:
        checks.append({"check": name, "ok": bool(ok), "detail": detail})
    add("xlsx_zip_integrity", zipfile.is_zipfile(path))
    wb = load_workbook(path, data_only=False)
    visible = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
    add("five_visible_sheets", visible == VISIBLE, str(visible))
    add("hidden_support_sheets", all(x in wb.sheetnames and wb[x].sheet_state != "visible" for x in HIDDEN))
    for s in ["MASTER_REVIEW", "NEG_RFI_RETURNS", "EVIDENCE_LINEAGE"]:
        ws = wb[s]
        add(f"{s}_freeze_panes", ws.freeze_panes is not None, str(ws.freeze_panes))
        add(f"{s}_autofilter", bool(ws.auto_filter.ref), str(ws.auto_filter.ref))
        grouped = any((dim.outlineLevel or 0) > 0 for dim in ws.column_dimensions.values())
        add(f"{s}_grouped_columns", grouped)
    start = wb["START_HERE"]
    text = " ".join(str(c.value or "") for row in start.iter_rows() for c in row)
    add("source_hash_visible", "SHA256" in text or "sha256" in text)
    add("deviation_not_nok_visible", "Deviation != NOK" in text or "Deviation ≠ NOK" in text)
    formula_errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#") and cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"}:
                    formula_errors.append(f"{ws.title}!{cell.coordinate}:{cell.value}")
    add("no_formula_error_literals", not formula_errors, "; ".join(formula_errors[:10]))
    return checks


async def check_html(path: Path, screenshot_dir: Path) -> list[dict[str, object]]:
    from playwright.async_api import async_playwright
    checks: list[dict[str, object]] = []
    def add(name: str, ok: bool, detail: str = "") -> None:
        checks.append({"check": name, "ok": bool(ok), "detail": detail})
    screenshot_dir.mkdir(parents=True, exist_ok=True)
    errors: list[str] = []
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        for width, height in [(1440, 1000), (1920, 1080), (1280, 800)]:
            page = await browser.new_page(viewport={"width": width, "height": height})
            page.on("console", lambda msg: errors.append(f"console:{msg.type}:{msg.text}") if msg.type == "error" else None)
            page.on("pageerror", lambda exc: errors.append(f"pageerror:{exc}"))
            await page.goto(path.resolve().as_uri())
            await page.wait_for_timeout(250)
            overflow = await page.evaluate("document.documentElement.scrollWidth > document.documentElement.clientWidth")
            add(f"no_default_horizontal_overflow_{width}x{height}", not overflow)
            count_before = (await page.locator("#count").inner_text()).strip()
            await page.locator("#q").fill("__unlikely_qps_search_token__")
            await page.wait_for_timeout(100)
            count_after = (await page.locator("#count").inner_text()).strip()
            add(f"search_changes_count_{width}x{height}", count_before != count_after, f"{count_before}->{count_after}")
            await page.locator("#q").fill("")
            await page.locator("#advanced").click()
            add(f"advanced_toggle_{width}x{height}", await page.evaluate("document.body.classList.contains('show-advanced')"))
            await page.locator("#dense").click()
            add(f"dense_toggle_{width}x{height}", await page.evaluate("document.body.classList.contains('dense')"))
            counter = await page.locator("#count").inner_text()
            dom_rows = await page.locator("#tbl tbody tr").count()
            add(f"shown_counter_matches_dom_{width}x{height}", counter.split("/")[0].strip() == str(dom_rows), f"{counter}; dom={dom_rows}")
            await page.screenshot(path=str(screenshot_dir / f"default_{width}x{height}.png"), full_page=True)
            await page.close()
        add("no_console_or_page_errors", not errors, "; ".join(errors[:20]))
        await browser.close()
    return checks


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", type=Path)
    p.add_argument("html", type=Path)
    p.add_argument("--out", type=Path, default=Path("reports/qps_ssot_visual_qa.json"))
    p.add_argument("--screenshots", type=Path, default=Path("reports/qps_ssot_screenshots"))
    args = p.parse_args()
    checks = check_xlsx(args.xlsx)
    checks += asyncio.run(check_html(args.html, args.screenshots))
    report = {
        "document_id": "QPS_SSOT_VISUAL_QA_RECEIPT",
        "checks": checks,
        "passed": sum(bool(c["ok"]) for c in checks),
        "total": len(checks),
        "pass_rate": sum(bool(c["ok"]) for c in checks) / len(checks) if checks else 0.0,
        "formal_credit": 0,
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))
    if not all(c["ok"] for c in checks):
        raise SystemExit(1)

if __name__ == "__main__":
    main()
